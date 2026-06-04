import json, os, csv, io, re, requests, secrets, smtplib, threading, time, logging
import psycopg2, psycopg2.extras
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime, date, timedelta
from functools import wraps
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, request, jsonify, g, session, redirect, url_for
from markupsafe import escape as html_escape
from werkzeug.security import generate_password_hash, check_password_hash
from kirpi_frontend import AUTH_HTML, HTML, LEGAL_CSS, legal_page
from kirpi_utils import row_to_dict, today_str, month_range, calc_depreciation

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("kirpi")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
_is_prod = not os.environ.get("FLASK_DEBUG")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=_is_prod,
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["300 per minute"],
    storage_uri=os.environ.get("REDIS_URL", "memory://"),
)

# ── CSRF ──────────────────────────────────────────────────────────────────────

def get_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]

def _csrf_input():
    return f'<input type="hidden" name="csrf_token" value="{get_csrf_token()}">'

def validate_csrf():
    # Form, JSON body ve header'dan sırayla dene
    submitted = (request.form.get("csrf_token")
                 or (request.get_json(silent=True) or {}).get("csrf_token")
                 or request.headers.get("X-CSRF-Token")
                 or "")
    expected = session.get("csrf_token", "")
    return bool(expected and secrets.compare_digest(str(submitted), str(expected)))

@app.errorhandler(429)
def too_many_requests(e):
    return AUTH_HTML_render("login", "⛔ Çok fazla deneme yaptınız. Lütfen birkaç dakika bekleyin."), 429

def _error_page(code, title, msg):
    return f"""<!DOCTYPE html><html lang="tr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{code} — Kirpi</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#f2f2f7;color:#1c1c1e;font-family:'Inter',system-ui,sans-serif;
  min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px}}
.box{{background:#fff;border-radius:20px;padding:40px 36px;text-align:center;max-width:380px;
  box-shadow:0 4px 24px rgba(0,0,0,.08);border:1px solid #d1d1d6}}
.code{{font-size:3rem;font-weight:900;color:#007aff;margin-bottom:8px}}
h1{{font-size:1.2rem;font-weight:700;margin-bottom:10px}}
p{{font-size:.88rem;color:#6d6d72;margin-bottom:24px;line-height:1.6}}
a{{display:inline-block;background:#007aff;color:#fff;padding:11px 24px;
  border-radius:10px;text-decoration:none;font-weight:700;font-size:.9rem}}
a:hover{{background:#0062cc}}</style></head>
<body><div class="box"><div class="code">{code}</div>
<h1>{title}</h1><p>{msg}</p>
<a href="/">Ana Sayfaya Dön</a></div></body></html>""", code

@app.errorhandler(404)
def not_found(_):
    return _error_page(404, "Sayfa Bulunamadı", "Aradığınız sayfa mevcut değil veya taşınmış olabilir.")

@app.errorhandler(500)
def server_error(_):
    return _error_page(500, "Sunucu Hatası", "Beklenmedik bir hata oluştu. Lütfen daha sonra tekrar deneyin.")

@app.after_request
def add_security_headers(resp):
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['X-Frame-Options'] = 'DENY'
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    if request.is_secure:
        resp.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    resp.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "script-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: blob:; "
        "connect-src 'self'; "
        "frame-ancestors 'none'"
    )
    return resp
# ── DATABASE ──────────────────────────────────────────────────────────────────
def _pg_url():
    url = os.environ.get("DATABASE_URL", "")
    if url.startswith("postgres://"):
        url = "postgresql://" + url[11:]
    return url

class _PGCursor:
    """Thin wrapper: makes psycopg2 DictCursor behave like sqlite3 cursor."""
    def __init__(self, cur):
        self._cur = cur
        self.lastrowid = None

    def execute(self, sql, params=()):
        sql = sql.replace("?", "%s")
        is_insert = sql.strip().upper().startswith("INSERT") and "RETURNING" not in sql.upper()
        if is_insert:
            sql = sql.rstrip("; ") + " RETURNING id"
        self._cur.execute(sql, params)
        if is_insert:
            row = self._cur.fetchone()
            self.lastrowid = row[0] if row else None
        return self

    def fetchone(self):  return self._cur.fetchone()
    def fetchall(self): return self._cur.fetchall()

class _PGConn:
    """Wrapper that makes psycopg2 connection look like sqlite3 connection."""
    def __init__(self):
        self._conn = psycopg2.connect(_pg_url(),
            cursor_factory=psycopg2.extras.DictCursor)
        self._conn.autocommit = False

    def execute(self, sql, params=()):
        cur = self._conn.cursor()
        c = _PGCursor(cur)
        c.execute(sql, params)
        return c

    def commit(self):   self._conn.commit()
    def rollback(self): self._conn.rollback()
    def close(self):    self._conn.close()

    def __enter__(self): return self
    def __exit__(self, exc_type, *_):
        if exc_type is None: self._conn.commit()
        else: self._conn.rollback()
        self._conn.close()
        return False

def pg_connect(): return _PGConn()

# ── MAIL CONFIG ───────────────────────────────────────────────────────────────
MAIL_FROM       = os.environ.get("MAIL_FROM", "")
MAIL_PASSWORD   = os.environ.get("MAIL_PASSWORD", "")
BACKUP_EMAIL    = os.environ.get("BACKUP_EMAIL", MAIL_FROM)
APP_URL         = os.environ.get("APP_URL", "https://kirpifinans.com")
TELEGRAM_TOKEN  = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_BOT_USERNAME = os.environ.get("TELEGRAM_BOT_USERNAME", "KirpiNakitBot")

GELIR_CATS = ["Maaş", "Serbest Meslek", "Kira Geliri", "Yatırım Geliri / Satış",
               "Yatırım / Temettü", "Hediye / İkramiye", "Hesaplar Arası Transfer", "Diğer Gelir"]
GIDER_CATS = ["Kira / Mortgage", "Market / Gıda", "Faturalar", "Ulaşım", "Yemek / Restoran",
               "Eğlence", "Sağlık", "Giyim", "Eğitim", "Abonelikler", "Elektronik",
               "Sigorta", "Vergi / Harç",
               "Kredi Kartı Ödemesi", "Yemek Kartı Ödemesi",
               "Döviz Alımı", "Altın Alımı", "Yatırım Fonu", "Hisse Senedi",
               "Hesaplar Arası Transfer", "Diğer Gider"]
ALL_CATS   = GELIR_CATS + GIDER_CATS

# ── DB ────────────────────────────────────────────────────────────────────────

def get_db():
    if "db" not in g:
        g.db = pg_connect()
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db:
        try: db.commit()
        except Exception: pass
        db.close()

def init_db():
    stmts = [
        """CREATE TABLE IF NOT EXISTS users (
            id             SERIAL PRIMARY KEY,
            username       TEXT UNIQUE NOT NULL,
            display_name   TEXT NOT NULL DEFAULT '',
            password_hash  TEXT NOT NULL,
            email          TEXT NOT NULL DEFAULT '',
            email_verified INTEGER NOT NULL DEFAULT 0,
            verify_token   TEXT,
            created_at     TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL,
            token      TEXT UNIQUE NOT NULL,
            expires_at TEXT NOT NULL,
            used       INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS profiles (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL,
            name       TEXT NOT NULL,
            type       TEXT NOT NULL DEFAULT 'sahis',
            created_at TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS transactions (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL DEFAULT 1,
            profile_id  INTEGER NOT NULL DEFAULT 1,
            type        TEXT NOT NULL,
            amount      DOUBLE PRECISION NOT NULL,
            category    TEXT NOT NULL,
            description TEXT DEFAULT '',
            date        TEXT NOT NULL,
            created_at  TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS budgets (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL DEFAULT 1,
            profile_id INTEGER NOT NULL DEFAULT 1,
            category   TEXT NOT NULL,
            limit_     DOUBLE PRECISION NOT NULL,
            UNIQUE(profile_id, category)
        )""",
        """CREATE TABLE IF NOT EXISTS cards (
            id            SERIAL PRIMARY KEY,
            user_id       INTEGER NOT NULL DEFAULT 1,
            profile_id    INTEGER NOT NULL DEFAULT 1,
            bank_name     TEXT NOT NULL,
            card_name     TEXT NOT NULL DEFAULT '',
            owner         TEXT NOT NULL DEFAULT '',
            limit_        DOUBLE PRECISION NOT NULL DEFAULT 0,
            used_         DOUBLE PRECISION NOT NULL DEFAULT 0,
            due_day       INTEGER NOT NULL DEFAULT 1,
            min_pct       DOUBLE PRECISION NOT NULL DEFAULT 25,
            statement_day INTEGER NOT NULL DEFAULT 20,
            created_at    TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS investments (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL DEFAULT 1,
            profile_id  INTEGER NOT NULL DEFAULT 1,
            name        TEXT NOT NULL,
            itype       TEXT NOT NULL,
            symbol      TEXT NOT NULL DEFAULT '',
            quantity    DOUBLE PRECISION NOT NULL DEFAULT 0,
            buy_price   DOUBLE PRECISION NOT NULL DEFAULT 0,
            buy_date    TEXT NOT NULL,
            note        TEXT DEFAULT '',
            created_at  TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS recurring (
            id           SERIAL PRIMARY KEY,
            user_id      INTEGER NOT NULL DEFAULT 1,
            profile_id   INTEGER NOT NULL DEFAULT 1,
            type         TEXT NOT NULL,
            amount       DOUBLE PRECISION NOT NULL,
            category     TEXT NOT NULL,
            description  TEXT DEFAULT '',
            day_of_month INTEGER NOT NULL DEFAULT 1,
            active       INTEGER NOT NULL DEFAULT 1,
            created_at   TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS accounts (
            id              SERIAL PRIMARY KEY,
            user_id         INTEGER NOT NULL DEFAULT 1,
            profile_id      INTEGER NOT NULL DEFAULT 1,
            name            TEXT NOT NULL,
            bank            TEXT NOT NULL DEFAULT '',
            type            TEXT NOT NULL DEFAULT 'hesap',
            initial_balance DOUBLE PRECISION NOT NULL DEFAULT 0,
            color           TEXT NOT NULL DEFAULT '#007aff',
            active          INTEGER NOT NULL DEFAULT 1,
            created_at      TEXT NOT NULL
        )""",
        "CREATE INDEX IF NOT EXISTS idx_txn_profile_date ON transactions(profile_id, date)",
        """CREATE TABLE IF NOT EXISTS invoices (
            id                  SERIAL PRIMARY KEY,
            user_id             INTEGER NOT NULL DEFAULT 1,
            profile_id          INTEGER NOT NULL DEFAULT 1,
            project_name        TEXT NOT NULL,
            client_name         TEXT NOT NULL DEFAULT '',
            description         TEXT DEFAULT '',
            amount              DOUBLE PRECISION NOT NULL DEFAULT 0,
            currency            TEXT NOT NULL DEFAULT 'TRY',
            invoice_date        TEXT NOT NULL,
            due_date            TEXT NOT NULL,
            status              TEXT NOT NULL DEFAULT 'bekliyor',
            paid_date           TEXT DEFAULT '',
            early_discount_pct  DOUBLE PRECISION NOT NULL DEFAULT 0,
            early_discount_days INTEGER NOT NULL DEFAULT 0,
            late_penalty_pct    DOUBLE PRECISION NOT NULL DEFAULT 0,
            notes               TEXT DEFAULT '',
            created_at          TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS goals (
            id             SERIAL PRIMARY KEY,
            user_id        INTEGER NOT NULL DEFAULT 1,
            profile_id     INTEGER NOT NULL DEFAULT 1,
            name           TEXT NOT NULL,
            goal_type      TEXT NOT NULL DEFAULT 'diger',
            monthly_target DOUBLE PRECISION NOT NULL DEFAULT 0,
            note           TEXT DEFAULT '',
            created_at     TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS todos (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL DEFAULT 1,
            profile_id  INTEGER NOT NULL DEFAULT 1,
            text        TEXT NOT NULL,
            date        TEXT NOT NULL,
            done        INTEGER NOT NULL DEFAULT 0,
            archived    INTEGER NOT NULL DEFAULT 0,
            created_at  TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS invoices (
            id                  SERIAL PRIMARY KEY,
            user_id             INTEGER NOT NULL DEFAULT 1,
            profile_id          INTEGER NOT NULL DEFAULT 1,
            project_name        TEXT NOT NULL,
            client_name         TEXT NOT NULL DEFAULT '',
            description         TEXT DEFAULT '',
            amount              DOUBLE PRECISION NOT NULL DEFAULT 0,
            currency            TEXT NOT NULL DEFAULT 'TRY',
            invoice_date        TEXT NOT NULL,
            due_date            TEXT NOT NULL,
            status              TEXT NOT NULL DEFAULT 'bekliyor',
            paid_date           TEXT DEFAULT '',
            early_discount_pct  DOUBLE PRECISION NOT NULL DEFAULT 0,
            early_discount_days INTEGER NOT NULL DEFAULT 0,
            late_penalty_pct    DOUBLE PRECISION NOT NULL DEFAULT 0,
            notes               TEXT DEFAULT '',
            created_at          TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS suppliers (
            id                  SERIAL PRIMARY KEY,
            user_id             INTEGER NOT NULL DEFAULT 1,
            profile_id          INTEGER NOT NULL DEFAULT 1,
            name                TEXT NOT NULL,
            tax_no              TEXT DEFAULT '',
            contact             TEXT DEFAULT '',
            phone               TEXT DEFAULT '',
            email               TEXT DEFAULT '',
            address             TEXT DEFAULT '',
            payment_terms_days  INTEGER NOT NULL DEFAULT 30,
            notes               TEXT DEFAULT '',
            created_at          TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS supplier_invoices (
            id              SERIAL PRIMARY KEY,
            user_id         INTEGER NOT NULL DEFAULT 1,
            profile_id      INTEGER NOT NULL DEFAULT 1,
            supplier_id     INTEGER NOT NULL DEFAULT 0,
            supplier_name   TEXT NOT NULL DEFAULT '',
            invoice_no      TEXT DEFAULT '',
            description     TEXT DEFAULT '',
            amount          DOUBLE PRECISION NOT NULL DEFAULT 0,
            currency        TEXT NOT NULL DEFAULT 'TRY',
            invoice_date    TEXT NOT NULL,
            due_date        TEXT NOT NULL,
            status          TEXT NOT NULL DEFAULT 'bekliyor',
            paid_date       TEXT DEFAULT '',
            reference_rate  DOUBLE PRECISION NOT NULL DEFAULT 50,
            notes           TEXT DEFAULT '',
            created_at      TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS assets (
            id                  SERIAL PRIMARY KEY,
            user_id             INTEGER NOT NULL DEFAULT 1,
            profile_id          INTEGER NOT NULL DEFAULT 1,
            name                TEXT NOT NULL,
            asset_type          TEXT NOT NULL DEFAULT 'diger',
            plate_no            TEXT DEFAULT '',
            serial_no           TEXT DEFAULT '',
            purchase_date       TEXT NOT NULL,
            purchase_price      DOUBLE PRECISION NOT NULL DEFAULT 0,
            useful_life_years   INTEGER NOT NULL DEFAULT 5,
            depreciation_method TEXT NOT NULL DEFAULT 'normal',
            depreciation_rate   DOUBLE PRECISION NOT NULL DEFAULT 20,
            maintenance_date    TEXT DEFAULT '',
            insurance_date      TEXT DEFAULT '',
            insurance_company   TEXT DEFAULT '',
            notes               TEXT DEFAULT '',
            active              INTEGER NOT NULL DEFAULT 1,
            created_at          TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS asset_maintenance (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL DEFAULT 1,
            profile_id  INTEGER NOT NULL DEFAULT 1,
            asset_id    INTEGER NOT NULL,
            date        TEXT NOT NULL,
            mtype       TEXT NOT NULL DEFAULT 'bakim',
            description TEXT DEFAULT '',
            cost        DOUBLE PRECISION NOT NULL DEFAULT 0,
            odometer    DOUBLE PRECISION NOT NULL DEFAULT 0,
            next_date   TEXT DEFAULT '',
            notes       TEXT DEFAULT '',
            created_at  TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS card_daily_balance (
            id          SERIAL PRIMARY KEY,
            user_id     INTEGER NOT NULL DEFAULT 1,
            profile_id  INTEGER NOT NULL DEFAULT 1,
            card_id     INTEGER NOT NULL,
            date        TEXT NOT NULL,
            balance     DOUBLE PRECISION NOT NULL DEFAULT 0,
            notes       TEXT DEFAULT '',
            created_at  TEXT NOT NULL
        )""",
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_card_daily ON card_daily_balance(card_id, date)",
        """CREATE TABLE IF NOT EXISTS telegram_links (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL,
            profile_id INTEGER NOT NULL DEFAULT 1,
            tg_user_id TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS telegram_link_codes (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL,
            profile_id INTEGER NOT NULL DEFAULT 1,
            code       TEXT NOT NULL UNIQUE,
            expires_at TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS telegram_pending (
            id          SERIAL PRIMARY KEY,
            tg_user_id  TEXT NOT NULL,
            user_id     INTEGER NOT NULL,
            profile_id  INTEGER NOT NULL DEFAULT 1,
            type        TEXT NOT NULL,
            amount      DOUBLE PRECISION NOT NULL,
            category    TEXT NOT NULL,
            description TEXT NOT NULL DEFAULT '',
            tx_date     TEXT NOT NULL DEFAULT '',
            created_at  TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS subscriptions (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL UNIQUE,
            plan       TEXT NOT NULL DEFAULT 'premium',
            status     TEXT NOT NULL DEFAULT 'active',
            started_at TEXT NOT NULL,
            expires_at TEXT NOT NULL DEFAULT '',
            note       TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        )""",
    ]
    migrations = [
        ("users",        "email",          "TEXT NOT NULL DEFAULT ''"),
        ("users",        "email_verified",  "INTEGER NOT NULL DEFAULT 0"),
        ("users",        "verify_token",    "TEXT"),
        ("transactions", "user_id",         "INTEGER NOT NULL DEFAULT 1"),
        ("transactions", "profile_id",      "INTEGER NOT NULL DEFAULT 1"),
        ("budgets",      "user_id",         "INTEGER NOT NULL DEFAULT 1"),
        ("budgets",      "profile_id",      "INTEGER NOT NULL DEFAULT 1"),
        ("cards",        "user_id",         "INTEGER NOT NULL DEFAULT 1"),
        ("cards",        "profile_id",      "INTEGER NOT NULL DEFAULT 1"),
        ("investments",  "user_id",         "INTEGER NOT NULL DEFAULT 1"),
        ("investments",  "profile_id",      "INTEGER NOT NULL DEFAULT 1"),
        ("recurring",    "user_id",         "INTEGER NOT NULL DEFAULT 1"),
        ("recurring",    "profile_id",      "INTEGER NOT NULL DEFAULT 1"),
        ("recurring",    "days_of_month",   "TEXT NOT NULL DEFAULT ''"),
        ("profiles",     "avatar",          "TEXT NOT NULL DEFAULT ''"),
        ("transactions", "account_id",      "INTEGER DEFAULT NULL"),
        ("users",        "avatar",          "TEXT NOT NULL DEFAULT ''"),
        ("accounts",     "limit_",          "DOUBLE PRECISION NOT NULL DEFAULT 0"),
        ("users",        "trial_start",     "TEXT NOT NULL DEFAULT ''"),
        ("users",        "report_name",     "TEXT NOT NULL DEFAULT ''"),
        ("users",        "report_contact",  "TEXT NOT NULL DEFAULT ''"),
        ("users",        "report_logo",     "TEXT NOT NULL DEFAULT ''"),
        ("users",           "phone",        "TEXT NOT NULL DEFAULT ''"),
        ("telegram_pending","tx_date",      "TEXT NOT NULL DEFAULT ''"),
        ("cards",           "card_type",    "TEXT NOT NULL DEFAULT 'kredi'"),
        ("transactions",    "card_id",      "INTEGER DEFAULT NULL"),
        ("accounts",        "owner",        "TEXT NOT NULL DEFAULT ''"),
    ]
    with pg_connect() as con:
        for stmt in stmts:
            con.execute(stmt)
        for table, col, col_def in migrations:
            exists = con.execute(
                "SELECT 1 FROM information_schema.columns WHERE table_name=%s AND column_name=%s",
                (table, col)
            ).fetchone()
            if not exists:
                try:
                    con.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_def}")
                    log.info("Migration: added %s.%s", table, col)
                except Exception as exc:
                    log.warning("Migration skip %s.%s: %s", table, col, exc)
        users = con.execute("SELECT id FROM users").fetchall()
        for u in users:
            uid = u["id"]
            has_profile = con.execute("SELECT id FROM profiles WHERE user_id=%s", (uid,)).fetchone()
            if not has_profile:
                con.execute(
                    "INSERT INTO profiles (user_id,name,type,created_at) VALUES (%s,%s,%s,%s)",
                    (uid, "Şahıs", "sahis", datetime.now().isoformat())
                )


# ── EMAIL ─────────────────────────────────────────────────────────────────────

# ── ACCOUNT LOCKOUT ───────────────────────────────────────────────────────────
import threading as _threading
_login_attempts: dict = {}  # {username: [timestamp, ...]}
_lockout_lock = _threading.Lock()
_MAX_ATTEMPTS = 5
_LOCKOUT_SECS = 15 * 60  # 15 dakika

def _record_failed_login(username: str):
    now = time.time()
    with _lockout_lock:
        attempts = [t for t in _login_attempts.get(username, []) if now - t < _LOCKOUT_SECS]
        attempts.append(now)
        _login_attempts[username] = attempts

def _is_locked_out(username: str) -> int:
    """Returns seconds remaining, or 0 if not locked out."""
    now = time.time()
    with _lockout_lock:
        attempts = [t for t in _login_attempts.get(username, []) if now - t < _LOCKOUT_SECS]
        _login_attempts[username] = attempts
        if len(attempts) >= _MAX_ATTEMPTS:
            oldest = min(attempts)
            remaining = int(_LOCKOUT_SECS - (now - oldest))
            return max(remaining, 0)
    return 0

def _clear_login_attempts(username: str):
    with _lockout_lock:
        _login_attempts.pop(username, None)

SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.zoho.eu")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))

def _smtp_send(msg):
    if not MAIL_FROM or not MAIL_PASSWORD:
        log.warning("Mail credentials not configured — skipping email send")
        return False
    try:
        if SMTP_PORT == 465:
            import ssl
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=15, context=ctx) as s:
                s.login(MAIL_FROM, MAIL_PASSWORD)
                s.sendmail(MAIL_FROM, msg["To"], msg.as_string())
        else:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(MAIL_FROM, MAIL_PASSWORD)
                s.sendmail(MAIL_FROM, msg["To"], msg.as_string())
        log.info("Email sent to %s", msg["To"])
        return True
    except Exception as exc:
        log.error("Email send failed: %s", exc)
        return False

def send_verify_email(to_email, username, token):
    verify_url = f"{APP_URL}/verify-email/{token}"
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "🦔 Kirpi — Email Adresinizi Doğrulayın"
    msg["From"]    = f"Kirpi Nakit Akışı <{MAIL_FROM}>"
    msg["To"]      = to_email
    text = f"Merhaba {username},\n\nKirpi'ye hoş geldiniz! Hesabınızı aktif etmek için aşağıdaki bağlantıya tıklayın:\n{verify_url}\n\nBu bağlantı 24 saat geçerlidir."
    html = f"""<!DOCTYPE html>
<html lang="tr"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0a0c12;font-family:'Inter',system-ui,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#0a0c12;padding:40px 20px">
  <tr><td align="center">
    <table width="520" cellpadding="0" cellspacing="0" style="background:#111318;border:1px solid #1e2233;border-radius:16px;overflow:hidden">
      <tr><td style="background:linear-gradient(135deg,#6366f1,#a855f7);padding:32px;text-align:center">
        <div style="font-size:40px;margin-bottom:8px">🦔</div>
        <h1 style="color:#fff;font-size:1.4rem;font-weight:800;margin:0">Kirpi'ye Hoş Geldiniz!</h1>
        <p style="color:rgba(255,255,255,.7);font-size:.85rem;margin:4px 0 0">Nakit Akışı Takibi</p>
      </td></tr>
      <tr><td style="padding:36px 40px">
        <h2 style="color:#e2e8f0;font-size:1.1rem;font-weight:700;margin:0 0 12px">Merhaba, {username} 👋</h2>
        <p style="color:#94a3b8;font-size:.9rem;line-height:1.7;margin:0 0 28px">
          Kayıt olduğunuz için teşekkürler! Hesabınızı aktif etmek için<br>
          aşağıdaki butona tıklayarak email adresinizi doğrulayın.
        </p>
        <table cellpadding="0" cellspacing="0" style="margin:0 auto 28px">
          <tr><td style="background:#6366f1;border-radius:10px;text-align:center">
            <a href="{verify_url}" style="display:block;padding:14px 36px;color:#fff;font-size:.95rem;font-weight:700;text-decoration:none">Email Adresimi Doğrula</a>
          </td></tr>
        </table>
        <div style="background:#1a1d26;border:1px solid #2a2f45;border-radius:8px;padding:14px 16px;margin-bottom:24px">
          <p style="color:#64748b;font-size:.78rem;margin:0 0 6px">Ya da bu bağlantıyı tarayıcınıza yapıştırın:</p>
          <p style="color:#818cf8;font-size:.76rem;word-break:break-all;margin:0">{verify_url}</p>
        </div>
        <p style="color:#64748b;font-size:.78rem;line-height:1.6;margin:0">
          ⏰ Bu bağlantı <strong style="color:#94a3b8">24 saat</strong> geçerlidir.
        </p>
      </td></tr>
      <tr><td style="background:#0d0f18;border-top:1px solid #1e2233;padding:20px 40px;text-align:center">
        <p style="color:#475569;font-size:.75rem;margin:0">🦔 Kirpi Nakit Akışı • Otomatik mesaj, lütfen yanıtlamayın.</p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>"""
    msg.attach(MIMEText(text, "plain", "utf-8"))
    msg.attach(MIMEText(html,  "html",  "utf-8"))
    return _smtp_send(msg)

def send_reset_email(to_email, display_name, username, token):
    reset_url = f"{APP_URL}/reset-password/{token}"
    msg = MIMEMultipart("alternative")
    msg["Subject"] = "🦔 Kirpi — Şifre Sıfırlama"
    msg["From"]    = f"Kirpi Nakit Akışı <{MAIL_FROM}>"
    msg["To"]      = to_email

    text = f"Merhaba {display_name},\n\nKullanıcı adınız: {username}\n\nŞifrenizi sıfırlamak için aşağıdaki bağlantıya tıklayın:\n{reset_url}\n\nBu bağlantı 1 saat geçerlidir.\n\nEğer bu talebi siz yapmadıysanız bu maili görmezden gelebilirsiniz."

    html = f"""<!DOCTYPE html>
<html lang="tr"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0a0c12;font-family:'Inter',system-ui,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#0a0c12;padding:40px 20px">
  <tr><td align="center">
    <table width="520" cellpadding="0" cellspacing="0" style="background:#111318;border:1px solid #1e2233;border-radius:16px;overflow:hidden">
      <tr><td style="background:linear-gradient(135deg,#6366f1,#a855f7);padding:32px;text-align:center">
        <div style="font-size:40px;margin-bottom:8px">🦔</div>
        <h1 style="color:#fff;font-size:1.4rem;font-weight:800;margin:0">Kirpi</h1>
        <p style="color:rgba(255,255,255,.7);font-size:.85rem;margin:4px 0 0">Nakit Akışı Takibi</p>
      </td></tr>
      <tr><td style="padding:36px 40px">
        <h2 style="color:#e2e8f0;font-size:1.1rem;font-weight:700;margin:0 0 12px">Merhaba, {display_name} 👋</h2>
        <div style="background:#1a1d26;border:1px solid #2a2f45;border-radius:8px;padding:12px 16px;margin-bottom:20px">
          <p style="color:#64748b;font-size:.78rem;margin:0 0 4px">Giriş için kullanıcı adınız:</p>
          <p style="color:#818cf8;font-size:.92rem;font-weight:700;margin:0">👤 {username}</p>
        </div>
        <p style="color:#94a3b8;font-size:.9rem;line-height:1.7;margin:0 0 28px">
          Kirpi hesabınız için bir <strong style="color:#e2e8f0">şifre sıfırlama talebi</strong> aldık.<br>
          Şifrenizi sıfırlamak için aşağıdaki butona tıklayın.
        </p>
        <table cellpadding="0" cellspacing="0" style="margin:0 auto 28px">
          <tr><td style="background:#6366f1;border-radius:10px;text-align:center">
            <a href="{reset_url}" style="display:block;padding:14px 36px;color:#fff;font-size:.95rem;font-weight:700;text-decoration:none">Şifremi Sıfırla</a>
          </td></tr>
        </table>
        <div style="background:#1a1d26;border:1px solid #2a2f45;border-radius:8px;padding:14px 16px;margin-bottom:24px">
          <p style="color:#64748b;font-size:.78rem;margin:0 0 6px">Ya da bu bağlantıyı tarayıcınıza yapıştırın:</p>
          <p style="color:#818cf8;font-size:.76rem;word-break:break-all;margin:0">{reset_url}</p>
        </div>
        <p style="color:#64748b;font-size:.78rem;line-height:1.6;margin:0">
          ⏰ Bu bağlantı <strong style="color:#94a3b8">1 saat</strong> geçerlidir.<br>
          🔒 Eğer bu talebi siz yapmadıysanız bu maili görmezden gelebilirsiniz.
        </p>
      </td></tr>
      <tr><td style="background:#0d0f18;border-top:1px solid #1e2233;padding:20px 40px;text-align:center">
        <p style="color:#475569;font-size:.75rem;margin:0">🦔 Kirpi Nakit Akışı • Otomatik mesaj, lütfen yanıtlamayın.</p>
      </td></tr>
    </table>
  </td></tr>
</table>
</body></html>"""

    msg.attach(MIMEText(text, "plain", "utf-8"))
    msg.attach(MIMEText(html,  "html",  "utf-8"))
    return _smtp_send(msg)

def _pg_dump_csv():
    """Export all tables as CSV and return as a combined text string."""
    tables = ["users","profiles","transactions","budgets","cards","investments","recurring"]
    out = io.StringIO()
    with pg_connect() as con:
        for t in tables:
            rows = con.execute(f"SELECT * FROM {t}").fetchall()
            if not rows:
                continue
            cols = list(rows[0].keys())
            out.write(f"-- TABLE: {t} --\n")
            w = csv.writer(out)
            w.writerow(cols)
            for r in rows:
                w.writerow([r[c] for c in cols])
            out.write("\n")
    return out.getvalue()

def send_backup_email(to_email):
    """Attach DB export as CSV and email it."""
    try:
        dump = _pg_dump_csv()
        dump_bytes = dump.encode("utf-8")
    except Exception as exc:
        log.error("Backup dump failed: %s", exc)
        return False

    now_str  = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"kirpi_backup_{now_str}.sql"

    msg = MIMEMultipart()
    msg["Subject"] = f"🦔 Kirpi Yedek — {now_str}"
    msg["From"]    = f"Kirpi Yedekleme <{MAIL_FROM}>"
    msg["To"]      = to_email

    body = MIMEText(
        f"Kirpi veritabanı yedeği — {now_str}\n\n"
        "Bu mail otomatik olarak oluşturulmuştur.\n"
        "Yedeği geri yüklemek için .sql dosyasını bir SQLite istemcisine aktarabilirsiniz.",
        "plain", "utf-8"
    )
    msg.attach(body)

    part = MIMEBase("application", "octet-stream")
    part.set_payload(dump_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(part)

    return _smtp_send(msg)

def _daily_backup_loop():
    """Background thread: send a DB backup email every 24 hours."""
    time.sleep(30)  # wait for app to fully start
    while True:
        if BACKUP_EMAIL:
            log.info("Running scheduled daily backup...")
            send_backup_email(BACKUP_EMAIL)
        time.sleep(86400)  # 24 hours

# ── INPUT VALIDATION ──────────────────────────────────────────────────────────
_MAX_LEN = {
    "username": 40, "display_name": 80, "email": 120,
    "password": 128, "description": 500, "category": 60,
    "note": 500, "name": 100,
}

def check_max_len(value: str, field: str) -> bool:
    limit = _MAX_LEN.get(field, 300)
    return len(value) <= limit

# ── AUTH ──────────────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Giriş gerekli", "auth": False}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

# ── ABONELİK ──────────────────────────────────────────────────────────────────

TRIAL_DAYS = 30

def get_sub_status(user_id: int) -> dict:
    """Returns dict: status ('trial'|'premium'|'free'), is_premium (bool), days_left (int)"""
    with pg_connect() as con:
        user = con.execute("SELECT created_at FROM users WHERE id=%s", (user_id,)).fetchone()
        if not user:
            return {"status": "free", "is_premium": False, "days_left": 0}
        try:
            created = datetime.fromisoformat(user["created_at"])
            trial_days_left = max(0, TRIAL_DAYS - (datetime.now() - created).days)
        except Exception:
            trial_days_left = 0
        if trial_days_left > 0:
            return {"status": "trial", "is_premium": True, "days_left": trial_days_left}
        sub = con.execute(
            "SELECT * FROM subscriptions WHERE user_id=%s AND status='active'", (user_id,)
        ).fetchone()
        if sub and sub["expires_at"]:
            try:
                expires = datetime.fromisoformat(sub["expires_at"])
                if expires > datetime.now():
                    return {"status": "premium", "is_premium": True, "days_left": (expires - datetime.now()).days}
            except Exception:
                pass
    return {"status": "free", "is_premium": False, "days_left": 0}

def premium_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        uid = session.get("user_id")
        if not uid:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Giriş gerekli", "auth": False}), 401
            return redirect("/login")
        sub = get_sub_status(uid)
        if not sub["is_premium"]:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "error": "Bu özellik Premium üyelik gerektirir.", "premium_required": True}), 403
            return redirect("/premium?required=1")
        return f(*args, **kwargs)
    return decorated

ADMIN_USER = os.environ.get("ADMIN_USERNAME", "")

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect("/login")
        if not ADMIN_USER:
            return "Yetkisiz erişim — ADMIN_USERNAME tanımlanmamış", 403
        if session.get("username") != ADMIN_USER:
            return "Yetkisiz erişim", 403
        return f(*args, **kwargs)
    return decorated

@app.route("/register", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def register():
    if session.get("user_id"):
        return redirect("/")
    if request.method == "POST":
        if not validate_csrf():
            return AUTH_HTML_render("register", "Geçersiz istek. Lütfen sayfayı yenileyin."), 403
        username = request.form.get("username","").strip().lower()
        display  = request.form.get("display_name","").strip()
        email    = request.form.get("email","").strip().lower()
        password = request.form.get("password","")
        confirm  = request.form.get("confirm","")
        error = None
        if not username or not password or not email:
            error = "Kullanıcı adı, email ve şifre zorunlu"
        elif not check_max_len(username, "username") or not username.replace("_","").replace("-","").isalnum():
            error = "Kullanıcı adı en fazla 40 karakter, sadece harf/rakam/_ içerebilir"
        elif "@" not in email or "." not in email.split("@")[-1]:
            error = "Geçerli bir email adresi girin"
        elif not check_max_len(email, "email"):
            error = "Email adresi çok uzun"
        elif len(password) < 8:
            error = "Şifre en az 8 karakter olmalı"
        elif not check_max_len(password, "password"):
            error = "Şifre çok uzun"
        elif password != confirm:
            error = "Şifreler eşleşmiyor"
        if error:
            return AUTH_HTML_render("register", error)
        phone = request.form.get("phone","").strip().replace(" ","")
        ptype = request.form.get("profile_type","sahis")
        if ptype not in ("sahis","sirket"): ptype = "sahis"
        prof_name = display if display else ("Şirket" if ptype=="sirket" else "Şahıs")
        try:
            vtok = secrets.token_urlsafe(32)
            with pg_connect() as con:
                cur = con.execute(
                    "INSERT INTO users (username,display_name,password_hash,email,email_verified,verify_token,phone,created_at) VALUES (?,?,?,?,0,?,?,?)",
                    (username, display or username, generate_password_hash(password), email, vtok, phone, datetime.now().isoformat())
                )
                uid = cur.lastrowid
                con.execute(
                    "INSERT INTO profiles (user_id,name,type,created_at) VALUES (?,?,?,?)",
                    (uid, prof_name, ptype, datetime.now().isoformat())
                )
        except psycopg2.IntegrityError:
            return AUTH_HTML_render("register", "Bu kullanıcı adı veya email zaten kullanımda")
        send_verify_email(email, display or username, vtok)
        return redirect("/login?verify_sent=1")
    return AUTH_HTML_render("register")

@app.route("/login", methods=["GET", "POST"])
@limiter.limit("20 per minute", methods=["POST"])
def login():
    if session.get("user_id"):
        return redirect("/")
    if request.method == "POST":
        if not validate_csrf():
            return AUTH_HTML_render("login", "Geçersiz istek. Lütfen sayfayı yenileyin."), 403
        username = request.form.get("username","").strip().lower()
        password = request.form.get("password","")
        # Hesap kilitleme kontrolü
        remaining = _is_locked_out(username)
        if remaining > 0:
            mins = remaining // 60 + 1
            return AUTH_HTML_render("login", f"⛔ Çok fazla hatalı deneme. {mins} dakika sonra tekrar deneyin.")
        with pg_connect() as con:
            row = con.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        if row and check_password_hash(row["password_hash"], password):
            if not row["email_verified"]:
                return AUTH_HTML_render("login", "✉️ Email adresiniz doğrulanmamış. Şifreyi biliyorsanız <a href='/forgot-password' style='color:#818cf8'>şifre sıfırla</a> yaparak hem şifreyi sıfırlayın hem de emailinizi doğrulatın. Ya da <a href='/resend-verify' style='color:#818cf8'>yeni doğrulama maili al</a>.")
            _clear_login_attempts(username)
            remember = request.form.get("remember_me") == "1"
            session.permanent   = remember          # True → 30 gün, False → tarayıcı kapanınca
            session["user_id"]  = row["id"]
            session["username"] = row["username"]
            session["display"]  = row["display_name"]
            with pg_connect() as con2:
                prof = con2.execute("SELECT * FROM profiles WHERE user_id=? ORDER BY id LIMIT 1",(row["id"],)).fetchone()
                if prof:
                    session["profile_id"]   = prof["id"]
                    session["profile_name"] = prof["name"]
                    session["profile_type"] = prof["type"]
            log.info("Login success: %s", username)
            return redirect("/")
        _record_failed_login(username)
        log.warning("Login failed: %s from %s", username, request.remote_addr)
        return AUTH_HTML_render("login", "Kullanıcı adı veya şifre hatalı")
    msg = ""
    if request.args.get("verify_sent"):
        msg = "✅ Doğrulama linki email adresinize gönderildi. Lütfen emailinizi kontrol edin."
    elif request.args.get("verified"):
        msg = "✅ Email doğrulandı! Giriş yapabilirsiniz."
    elif request.args.get("reset"):
        msg = "Şifreniz başarıyla güncellendi. Giriş yapabilirsiniz."
    return AUTH_HTML_render("login", msg)

@app.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def forgot_password():
    if request.method == "POST":
        if not validate_csrf():
            return AUTH_HTML_render("forgot", "Geçersiz istek. Lütfen sayfayı yenileyin."), 403
        identifier = request.form.get("identifier","").strip()
        id_lower   = identifier.lower()
        # Telefon numarasını normalleştir (başındaki 0 veya +90 kaldır)
        phone_clean = identifier.replace(" ","").replace("-","").replace("(","").replace(")","")
        if phone_clean.startswith("+90"): phone_clean = phone_clean[3:]
        if phone_clean.startswith("90") and len(phone_clean) > 10: phone_clean = phone_clean[2:]
        if phone_clean.startswith("0"): phone_clean = phone_clean[1:]
        with pg_connect() as con:
            row = con.execute(
                "SELECT * FROM users WHERE username=%s OR email=%s OR "
                "(phone != '' AND (phone=%s OR phone=%s OR phone=%s))",
                (id_lower, id_lower, identifier, phone_clean, "0"+phone_clean)
            ).fetchone()
        if row:
            token = secrets.token_urlsafe(32)
            expires = (datetime.now() + timedelta(hours=1)).isoformat()
            with pg_connect() as con:
                con.execute(
                    "INSERT INTO password_reset_tokens (user_id,token,expires_at,created_at) VALUES (?,?,?,?)",
                    (row["id"], token, expires, datetime.now().isoformat())
                )
            sent = send_reset_email(row["email"], row["display_name"] or row["username"], row["username"], token)
            log.info("Reset email sent=%s for user=%s", sent, row["username"])
        # Always show success to prevent user enumeration
        return AUTH_HTML_render("forgot_sent", "Eğer bu hesap varsa, şifre sıfırlama linki email adresinize gönderildi.")
    return AUTH_HTML_render("forgot")

@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    with pg_connect() as con:
        tok = con.execute(
            "SELECT * FROM password_reset_tokens WHERE token=? AND used=0", (token,)
        ).fetchone()
    if not tok or datetime.fromisoformat(tok["expires_at"]) < datetime.now():
        return AUTH_HTML_render("reset_invalid", "Bu link geçersiz veya süresi dolmuş.")
    if request.method == "POST":
        if not validate_csrf():
            return AUTH_HTML_render("reset", "Geçersiz istek. Lütfen sayfayı yenileyin.", token=token), 403
        password = request.form.get("password","")
        confirm  = request.form.get("confirm","")
        if len(password) < 6:
            return AUTH_HTML_render("reset", "Şifre en az 6 karakter olmalı", token=token)
        if password != confirm:
            return AUTH_HTML_render("reset", "Şifreler eşleşmiyor", token=token)
        with pg_connect() as con:
            con.execute(
                "UPDATE users SET password_hash=?, email_verified=1, verify_token=NULL WHERE id=?",
                (generate_password_hash(password), tok["user_id"])
            )
            con.execute("UPDATE password_reset_tokens SET used=1 WHERE token=?", (token,))
        return redirect("/login?reset=1")
    return AUTH_HTML_render("reset", token=token)

@app.route("/verify-email/<token>")
def verify_email(token):
    with pg_connect() as con:
        row = con.execute("SELECT * FROM users WHERE verify_token=? AND email_verified=0", (token,)).fetchone()
    if not row:
        return AUTH_HTML_render("reset_invalid", "Bu doğrulama linki geçersiz veya zaten kullanılmış.")
    with pg_connect() as con:
        con.execute("UPDATE users SET email_verified=1, verify_token=NULL WHERE id=?", (row["id"],))
    return redirect("/login?verified=1")

@app.route("/resend-verify", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def resend_verify():
    if request.method == "POST":
        if not validate_csrf():
            return AUTH_HTML_render("resend_verify", "Geçersiz istek. Lütfen sayfayı yenileyin."), 403
        email = request.form.get("email","").strip().lower()
        with pg_connect() as con:
            row = con.execute("SELECT * FROM users WHERE email=? AND email_verified=0", (email,)).fetchone()
        if row:
            vtok = secrets.token_urlsafe(32)
            with pg_connect() as con:
                con.execute("UPDATE users SET verify_token=? WHERE id=?", (vtok, row["id"]))
            send_verify_email(email, row["display_name"] or row["username"], vtok)
        return AUTH_HTML_render("forgot_sent", "Eğer bu email doğrulanmamış bir hesaba aitse, yeni doğrulama linki gönderildi.")
    return AUTH_HTML_render("resend_verify")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

# ── Mobil uygulama JSON API ──────────────────────────────────────────────────

@app.route("/api/mobile/login", methods=["POST"])
@limiter.limit("20 per minute")
def mobile_login():
    data     = request.get_json(silent=True) or {}
    username = data.get("username", "").strip().lower()
    password = data.get("password", "")
    if not username or not password:
        return jsonify({"ok": False, "error": "Kullanıcı adı ve şifre gerekli"}), 400
    remaining = _is_locked_out(username)
    if remaining > 0:
        mins = remaining // 60 + 1
        return jsonify({"ok": False, "error": f"Çok fazla hatalı deneme. {mins} dakika sonra tekrar deneyin."}), 429
    with pg_connect() as con:
        row = con.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not (row and check_password_hash(row["password_hash"], password)):
        _record_failed_login(username)
        return jsonify({"ok": False, "error": "Kullanıcı adı veya şifre hatalı"}), 401
    if not row["email_verified"]:
        return jsonify({"ok": False, "error": "E-posta adresiniz doğrulanmamış"}), 403
    _clear_login_attempts(username)
    session.permanent  = True
    session["user_id"] = row["id"]
    session["username"] = row["username"]
    session["display"]  = row["display_name"]
    with pg_connect() as con2:
        prof = con2.execute(
            "SELECT * FROM profiles WHERE user_id=? ORDER BY id LIMIT 1", (row["id"],)
        ).fetchone()
        profiles = con2.execute(
            "SELECT * FROM profiles WHERE user_id=? ORDER BY id", (row["id"],)
        ).fetchall()
        if prof:
            session["profile_id"]   = prof["id"]
            session["profile_name"] = prof["name"]
            session["profile_type"] = prof["type"]
    sub = get_sub_status(row["id"])
    return jsonify({
        "ok":         True,
        "id":         row["id"],
        "username":   row["username"],
        "email":      row["email"],
        "is_premium": sub.get("active", False),
        "active_profile_id": prof["id"] if prof else None,
        "profiles": [
            {"id": p["id"], "name": p["name"], "type": p["type"]}
            for p in profiles
        ],
    })

@app.route("/api/mobile/logout", methods=["POST"])
def mobile_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/mobile/me")
@login_required
def mobile_me():
    uid = session["user_id"]
    with pg_connect() as con:
        row      = con.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
        profiles = con.execute(
            "SELECT * FROM profiles WHERE user_id=? ORDER BY id", (uid,)
        ).fetchall()
    sub = get_sub_status(uid)
    return jsonify({
        "ok":         True,
        "id":         uid,
        "username":   row["username"] if row else session.get("username"),
        "email":      row["email"] if row else "",
        "is_premium": sub.get("active", False),
        "active_profile_id": session.get("profile_id"),
        "profiles": [
            {"id": p["id"], "name": p["name"], "type": p["type"]}
            for p in profiles
        ],
    })

@app.route("/api/me")
@login_required
def api_me():
    uid = session.get("user_id")
    db  = get_db()
    row = db.execute("SELECT display_name,username,email,avatar,phone FROM users WHERE id=?", (uid,)).fetchone()
    avatar = row["avatar"] if row else ""
    sub    = get_sub_status(uid)
    return jsonify({
        "username":     session.get("username"),
        "display":      session.get("display") or (row["display_name"] if row else ""),
        "profile_id":   session.get("profile_id"),
        "profile_name": session.get("profile_name"),
        "profile_type": session.get("profile_type"),
        "email":        row["email"] if row else "",
        "phone":        (row["phone"] if row else "") or "",
        "avatar":       avatar or "",
        "subscription": sub,
    })

@app.route("/api/me/update", methods=["POST"])
@login_required
def api_me_update():
    data    = request.get_json(force=True)
    display = (data.get("display_name") or "").strip()
    avatar  = data.get("avatar", None)
    phone   = (data.get("phone") or "").strip().replace(" ","")[:20]
    uid     = session["user_id"]
    db      = get_db()
    if display:
        db.execute("UPDATE users SET display_name=? WHERE id=?", (display, uid))
        session["display"] = display
    if phone is not None:
        db.execute("UPDATE users SET phone=? WHERE id=?", (phone, uid))
    if avatar is not None:
        safe_prefixes = ("data:image/png;base64,", "data:image/jpeg;base64,",
                         "data:image/jpg;base64,", "data:image/webp;base64,")
        if avatar and not any(avatar.startswith(p) for p in safe_prefixes):
            return jsonify({"ok": False, "error": "Geçersiz resim formatı"}), 400
        db.execute("UPDATE users SET avatar=? WHERE id=?", (avatar, uid))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/me/password", methods=["POST"])
@login_required
def api_me_password():
    data    = request.get_json(force=True)
    current = data.get("current", "")
    new_pw  = data.get("new", "")
    confirm = data.get("confirm", "")
    uid     = session["user_id"]
    db      = get_db()
    user    = db.execute("SELECT password_hash FROM users WHERE id=?", (uid,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], current):
        return jsonify({"ok": False, "error": "Mevcut şifre yanlış"})
    if len(new_pw) < 6:
        return jsonify({"ok": False, "error": "Şifre en az 6 karakter olmalı"})
    if new_pw != confirm:
        return jsonify({"ok": False, "error": "Şifreler eşleşmiyor"})
    db.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_pw), uid))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/me/delete", methods=["POST"])
@login_required
def api_me_delete():
    uid  = session["user_id"]
    data = request.get_json(force=True)
    pw   = data.get("password", "")
    db   = get_db()
    user = db.execute("SELECT password_hash FROM users WHERE id=?", (uid,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], pw):
        return jsonify({"ok": False, "error": "Şifre hatalı"}), 403
    user_tables = [
        "telegram_links", "telegram_link_codes", "telegram_pending",
        "card_daily_balance", "asset_maintenance", "assets",
        "supplier_invoices", "suppliers", "invoices", "accounts",
        "investments", "recurring", "todos", "goals", "budgets",
        "transactions", "cards", "profiles", "password_reset_tokens",
    ]
    with pg_connect() as con:
        for tbl in user_tables:
            con.execute(f"DELETE FROM {tbl} WHERE user_id=%s", (uid,))
        con.execute("DELETE FROM users WHERE id=%s", (uid,))
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me/report-settings", methods=["GET"])
@login_required
def api_report_settings_get():
    uid = session["user_id"]
    db  = get_db()
    row = db.execute("SELECT report_name,report_contact,report_logo FROM users WHERE id=?", (uid,)).fetchone()
    if not row:
        return jsonify({"report_name":"","report_contact":"","report_logo":""})
    return jsonify({"report_name":row["report_name"],"report_contact":row["report_contact"],"report_logo":row["report_logo"]})

@app.route("/api/me/report-settings", methods=["POST"])
@login_required
def api_report_settings_save():
    uid  = session["user_id"]
    data = request.get_json(force=True)
    name    = (data.get("report_name") or "")[:120]
    contact = (data.get("report_contact") or "")[:200]
    logo    = data.get("report_logo") or ""
    if logo:
        safe_prefixes = ("data:image/png;base64,","data:image/jpeg;base64,","data:image/jpg;base64,","data:image/webp;base64,","data:image/svg+xml;base64,")
        if not any(logo.startswith(p) for p in safe_prefixes):
            return jsonify({"ok":False,"error":"Geçersiz logo formatı"}),400
        if len(logo) > 500_000:
            return jsonify({"ok":False,"error":"Logo dosyası çok büyük (max 375 KB)"}),400
    db = get_db()
    db.execute("UPDATE users SET report_name=%s,report_contact=%s,report_logo=%s WHERE id=%s",(name,contact,logo,uid))
    db.commit()
    return jsonify({"ok":True})

def get_pid():
    """Return active profile_id for the current user — always validates ownership."""
    uid = session.get("user_id")
    if not uid:
        return None
    pid = session.get("profile_id")
    with pg_connect() as con:
        if pid:
            # Verify this pid actually belongs to the current user
            valid = con.execute(
                "SELECT id FROM profiles WHERE id=%s AND user_id=%s", (pid, uid)
            ).fetchone()
            if valid:
                return pid
        # Fallback: use first profile of this user
        row = con.execute(
            "SELECT id FROM profiles WHERE user_id=%s ORDER BY id LIMIT 1", (uid,)
        ).fetchone()
        if row:
            session["profile_id"] = row["id"]
            return row["id"]
    return None

# ── PROFILES API ──────────────────────────────────────────────────────────────

@app.route("/api/profiles", methods=["GET"])
@login_required
def list_profiles():
    uid = session["user_id"]
    rows = get_db().execute("SELECT * FROM profiles WHERE user_id=? ORDER BY id",(uid,)).fetchall()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/profiles", methods=["POST"])
@login_required
def add_profile():
    uid = session["user_id"]
    d = request.get_json(force=True)
    name  = d.get("name","").strip()
    ptype = d.get("type","sahis")
    if not name or ptype not in ("sahis","sirket"):
        return jsonify({"ok":False,"error":"Geçersiz veri"}),400
    db = get_db()
    cur = db.execute(
        "INSERT INTO profiles (user_id,name,type,created_at) VALUES (?,?,?,?)",
        (uid, name, ptype, datetime.now().isoformat()))
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid,"name":name,"type":ptype})

@app.route("/api/profiles/<int:pid>/avatar", methods=["PUT"])
@login_required
def update_profile_avatar(pid):
    uid = session["user_id"]
    data = request.get_json(force=True)
    avatar = data.get("avatar", "")
    safe_prefixes = ("data:image/png;base64,", "data:image/jpeg;base64,",
                     "data:image/jpg;base64,", "data:image/webp;base64,")
    if avatar and not any(avatar.startswith(p) for p in safe_prefixes):
        return jsonify({"ok": False, "error": "Geçersiz resim formatı"}), 400
    db = get_db()
    db.execute("UPDATE profiles SET avatar=? WHERE id=? AND user_id=?", (avatar, pid, uid))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/profiles/<int:pid>/switch", methods=["POST"])
@login_required
def switch_profile(pid):
    uid = session["user_id"]
    db  = get_db()
    prof = db.execute("SELECT * FROM profiles WHERE id=? AND user_id=?",(pid,uid)).fetchone()
    if not prof: return jsonify({"ok":False,"error":"Profil bulunamadı"}),404
    session["profile_id"]   = prof["id"]
    session["profile_name"] = prof["name"]
    session["profile_type"] = prof["type"]
    return jsonify({"ok":True,"name":prof["name"],"type":prof["type"]})

@app.route("/api/profiles/<int:pid>", methods=["DELETE"])
@login_required
def del_profile(pid):
    uid = session["user_id"]
    db  = get_db()
    count = db.execute("SELECT COUNT(*) AS n FROM profiles WHERE user_id=?",(uid,)).fetchone()["n"]
    if count <= 1: return jsonify({"ok":False,"error":"En az bir profil olmalı"}),400
    db.execute("DELETE FROM profiles WHERE id=? AND user_id=?",(pid,uid))
    db.commit()
    if session.get("profile_id") == pid:
        first = db.execute("SELECT * FROM profiles WHERE user_id=? ORDER BY id LIMIT 1",(uid,)).fetchone()
        if first:
            session["profile_id"]   = first["id"]
            session["profile_name"] = first["name"]
            session["profile_type"] = first["type"]
    return jsonify({"ok":True})

# ── CRUD ──────────────────────────────────────────────────────────────────────

@app.route("/api/transactions", methods=["GET"])
@login_required
def list_transactions():
    db    = get_db()
    year  = request.args.get("year",  type=int)
    month = request.args.get("month", type=int)
    pid   = get_pid()
    q     = "SELECT * FROM transactions WHERE profile_id=?"
    params = [pid]
    if year and month:
        s, e = month_range(year, month)
        q += " AND date BETWEEN ? AND ?"; params += [s, e]
    elif year:
        q += " AND date LIKE ?"; params.append(f"{year}-%")
    q += " ORDER BY date DESC, id DESC"
    return jsonify([row_to_dict(r) for r in db.execute(q, params).fetchall()])

@app.route("/api/transactions", methods=["POST"])
@login_required
def add_transaction():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    ttype = d.get("type"); amount = float(d.get("amount", 0))
    cat = d.get("category",""); desc = d.get("description",""); dt = d.get("date", today_str())
    account_id = d.get("account_id") or None
    card_id    = d.get("card_id") or None
    if ttype not in ("gelir","gider") or amount <= 0 or not cat:
        return jsonify({"ok": False, "error": "Geçersiz veri"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,account_id,card_id,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (uid, pid, ttype, amount, cat, desc, dt, account_id, card_id, datetime.now().isoformat()))
    # Kart borcu otomatik güncelle
    if card_id:
        if ttype == "gider":
            db.execute("UPDATE cards SET used_=used_+? WHERE id=? AND profile_id=?", (amount, card_id, pid))
        elif ttype == "gelir":
            db.execute("UPDATE cards SET used_=GREATEST(0,used_-?) WHERE id=? AND profile_id=?", (amount, card_id, pid))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/transactions/<int:tid>", methods=["PUT"])
@login_required
def update_transaction(tid):
    pid = get_pid(); db = get_db()
    d = request.get_json(force=True)
    # Eski işlemi oku — kart borcu senkronizasyonu için
    old = db.execute("SELECT * FROM transactions WHERE id=? AND profile_id=?", (tid, pid)).fetchone()
    if not old: return jsonify({"ok": False, "error": "İşlem bulunamadı"}), 404

    fields, params = [], []
    for col in ("type","amount","category","description","date","account_id","card_id"):
        if col in d:
            fields.append(f"{col}=?")
            if col == "amount":
                params.append(float(d[col]))
            elif col in ("account_id","card_id"):
                params.append(int(d[col]) if d[col] else None)
            else:
                params.append(d[col])
    if not fields: return jsonify({"ok": False}), 400
    params += [tid, pid]
    db.execute(f"UPDATE transactions SET {','.join(fields)} WHERE id=? AND profile_id=?", params)

    # Kart borcu senkronizasyonu: eski etkiyi geri al, yeni etkiyi uygula
    old_card_id = old["card_id"]
    new_card_id = int(d["card_id"]) if "card_id" in d and d["card_id"] else old_card_id
    old_amount  = float(old["amount"])
    new_amount  = float(d["amount"]) if "amount" in d else old_amount
    old_type    = old["type"]
    new_type    = d.get("type", old_type)

    def _card_delta(card_id, amount, ttype, reverse=False):
        if not card_id: return
        sign = -1 if reverse else 1
        if ttype == "gider":
            db.execute("UPDATE cards SET used_=GREATEST(0,used_+?) WHERE id=? AND profile_id=?",
                       (sign * amount, card_id, pid))
        elif ttype == "gelir":
            db.execute("UPDATE cards SET used_=GREATEST(0,used_+?) WHERE id=? AND profile_id=?",
                       (-sign * amount, card_id, pid))

    if old_card_id or new_card_id:
        _card_delta(old_card_id, old_amount, old_type, reverse=True)   # eski etkiyi sil
        _card_delta(new_card_id, new_amount, new_type, reverse=False)  # yeni etkiyi uygula

    db.commit()
    return jsonify({"ok": True})

@app.route("/api/transactions/<int:tid>", methods=["DELETE"])
@login_required
def del_transaction(tid):
    pid = get_pid(); db = get_db()
    tx = db.execute("SELECT * FROM transactions WHERE id=? AND profile_id=?", (tid, pid)).fetchone()
    if tx and tx.get("card_id"):
        cid = tx["card_id"]
        amt = float(tx["amount"])
        has_account = bool(tx.get("account_id"))
        if tx["type"] == "gider":
            if has_account:
                # Kart ödemesi silindi → borcu geri yükle (ödeme geri alındı)
                db.execute("UPDATE cards SET used_=used_+? WHERE id=? AND profile_id=?", (amt, cid, pid))
            else:
                # Kart harcaması silindi → borcu düşür (harcama geri alındı)
                db.execute("UPDATE cards SET used_=GREATEST(0,used_-?) WHERE id=? AND profile_id=?", (amt, cid, pid))
        elif tx["type"] == "gelir":
            # İade silindi → borcu geri yükle
            db.execute("UPDATE cards SET used_=used_+? WHERE id=? AND profile_id=?", (amt, cid, pid))
    db.execute("DELETE FROM transactions WHERE id=? AND profile_id=?", (tid, pid))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/transactions/bulk-delete", methods=["POST"])
@login_required
def bulk_delete():
    pid = get_pid()
    ids = request.get_json(force=True)
    if not isinstance(ids, list): return jsonify({"ok": False}), 400
    placeholders = ",".join("?" * len(ids))
    get_db().execute(f"DELETE FROM transactions WHERE id IN ({placeholders}) AND profile_id=?", ids + [pid])
    get_db().commit()
    return jsonify({"ok": True, "deleted": len(ids)})

@app.route("/api/summary")
@login_required
def summary():
  try:
    pid   = get_pid()
    db    = get_db()
    today_d = date.today()
    year  = request.args.get("year",  today_d.year,  type=int)
    month = request.args.get("month", today_d.month, type=int)
    period = request.args.get("period", "month")  # month | year | all
    rstart = request.args.get("start","")
    rend   = request.args.get("end","")
    if rstart and rend:
        start, end = rstart, rend
    elif period == "year":
        year_param = request.args.get("year", today_d.year, type=int)
        start = f"{year_param}-01-01"
        end   = f"{year_param}-12-31"
    elif period == "all":
        start = "0001-01-01"
        end   = "9999-12-31"
    else:  # month (default)
        start, end = month_range(year, month)
    rows = db.execute(
        "SELECT type,category,SUM(amount) as total FROM transactions "
        "WHERE profile_id=? AND date BETWEEN ? AND ? GROUP BY type,category", (pid, start, end)).fetchall()
    gelir_total = gider_total = 0.0
    gelir_cats  = {}; gider_cats = {}
    for r in rows:
        if r["type"]=="gelir": gelir_total+=r["total"]; gelir_cats[r["category"]]=round(r["total"],2)
        else:                   gider_total+=r["total"]; gider_cats[r["category"]]=round(r["total"],2)
    bar = []
    for m in range(1,13):
        s,e2 = month_range(today_d.year, m)
        totals = db.execute("SELECT type,SUM(amount) as t FROM transactions WHERE profile_id=? AND date BETWEEN ? AND ? GROUP BY type",(pid,s,e2)).fetchall()
        g_val  = float(next((r["t"] for r in totals if r["type"]=="gelir"),0) or 0)
        ex_val = float(next((r["t"] for r in totals if r["type"]=="gider"),0) or 0)
        bar.append({"month":m,"gelir":round(g_val,2),"gider":round(ex_val,2)})
    bal = db.execute("SELECT SUM(CASE WHEN type='gelir' THEN amount ELSE -amount END) as b FROM transactions WHERE profile_id=?",(pid,)).fetchone()
    budgets = {r["category"]:r["limit_"] for r in db.execute("SELECT * FROM budgets WHERE profile_id=?",(pid,)).fetchall()}

    # ── Net Kullanılabilir Nakit ──────────────────────────────────────────────
    # Bu ay gelir - bu ay gider - kart asgari ödemeler (henüz ödenmemişse)
    cards = db.execute("SELECT used_, min_pct, limit_ FROM cards WHERE profile_id=?", (pid,)).fetchall()
    toplam_kart_borcu   = sum(float(c["used_"] or 0) for c in cards)
    toplam_kart_limit   = sum(float(c["limit_"] or 0) for c in cards)
    toplam_asgari       = sum(
        round(float(c["used_"] or 0) * float(c["min_pct"] or 25) / 100, 2)
        for c in cards if float(c["used_"] or 0) > 0
    )
    # Recurring zorunlu giderler (aktif)
    recurring_gider = db.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM recurring WHERE profile_id=? AND type='gider' AND active=1",
        (pid,)
    ).fetchone()["t"] or 0
    # Kullanılabilir nakit = bu ay net - asgari ödemeler
    net_this_month   = round(gelir_total - gider_total, 2)
    kullanilabilir   = round(net_this_month - toplam_asgari, 2)
    # Kart kullanılabilir limit
    kart_kullanilabilir = round(toplam_kart_limit - toplam_kart_borcu, 2)

    return jsonify({
        "gelir": round(gelir_total,2), "gider": round(gider_total,2),
        "net":   round(gelir_total-gider_total,2),
        "balance": round(bal["b"] or 0, 2),
        "gelir_cats": gelir_cats, "gider_cats": gider_cats,
        "bar": bar, "budgets": budgets, "period": period,
        "kart_borcu":        round(toplam_kart_borcu, 2),
        "kart_limit":        round(toplam_kart_limit, 2),
        "kart_kullanilabilir": kart_kullanilabilir,
        "asgari_odeme":      round(toplam_asgari, 2),
        "recurring_gider":   round(recurring_gider, 2),
        "kullanilabilir_nakit": kullanilabilir,
    })
  except Exception as _e:
    log.error("summary error: %s", _e, exc_info=True)
    return jsonify({"error": str(_e), "gelir":0,"gider":0,"net":0,"balance":0,
                    "gelir_cats":{},"gider_cats":{},"bar":[],"budgets":{},"period":"month",
                    "kart_borcu":0,"kart_limit":0,"kart_kullanilabilir":0,
                    "asgari_odeme":0,"recurring_gider":0,"kullanilabilir_nakit":0})

@app.route("/api/budgets", methods=["POST"])
@login_required
def set_budget():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    cat = d.get("category",""); limit = float(d.get("limit",0))
    if not cat or limit<=0: return jsonify({"ok":False}),400
    db = get_db()
    db.execute("INSERT INTO budgets(user_id,profile_id,category,limit_) VALUES(?,?,?,?) ON CONFLICT(profile_id,category) DO UPDATE SET limit_=excluded.limit_",(uid,pid,cat,limit))
    db.commit(); return jsonify({"ok":True})

@app.route("/api/goals", methods=["GET"])
@login_required
def list_goals():
    pid = get_pid(); db = get_db()
    rows = db.execute("SELECT * FROM goals WHERE profile_id=? ORDER BY created_at", (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/goals", methods=["POST"])
@login_required
def add_goal():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    name = d.get("name","").strip()
    gtype = d.get("goal_type","diger")
    monthly = float(d.get("monthly_target",0))
    note = d.get("note","").strip()
    if not name or monthly <= 0:
        return jsonify({"ok":False,"error":"İsim ve aylık hedef zorunlu"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO goals (user_id,profile_id,name,goal_type,monthly_target,note,created_at) VALUES (?,?,?,?,?,?,?)",
        (uid, pid, name, gtype, monthly, note, datetime.now().isoformat())
    )
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/goals/<int:gid>", methods=["DELETE"])
@login_required
def del_goal(gid):
    pid = get_pid(); db = get_db()
    db.execute("DELETE FROM goals WHERE id=? AND profile_id=?", (gid, pid))
    db.commit()
    return jsonify({"ok":True})

# ── TODOS ────────────────────────────────────────────────────────────────────

@app.route("/api/todos", methods=["GET"])
@login_required
def list_todos():
    pid = get_pid(); db = get_db()
    dt = request.args.get("date", date.today().isoformat())
    rows = db.execute(
        "SELECT * FROM todos WHERE profile_id=? AND date=? ORDER BY id", (pid, dt)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/todos", methods=["POST"])
@login_required
def add_todo():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    text = d.get("text","").strip()
    dt   = d.get("date", date.today().isoformat())
    if not text: return jsonify({"ok":False,"error":"Görev metni gerekli"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO todos (user_id,profile_id,text,date,done,archived,created_at) VALUES (?,?,?,?,0,0,?)",
        (uid, pid, text, dt, datetime.now().isoformat())
    )
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/todos/<int:tid>", methods=["PUT"])
@login_required
def update_todo(tid):
    pid = get_pid(); db = get_db()
    d = request.get_json(force=True)
    fields, params = [], []
    for col in ("text","done","archived","date"):
        if col in d:
            fields.append(f"{col}=?")
            params.append(d[col])
    if not fields: return jsonify({"ok":False}), 400
    params += [tid, pid]
    db.execute(f"UPDATE todos SET {','.join(fields)} WHERE id=? AND profile_id=?", params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/todos/<int:tid>", methods=["DELETE"])
@login_required
def del_todo(tid):
    pid = get_pid(); db = get_db()
    db.execute("DELETE FROM todos WHERE id=? AND profile_id=?", (tid, pid))
    db.commit()
    return jsonify({"ok":True})

# ── SUPPLIERS ────────────────────────────────────────────────────────────────

@app.route("/api/suppliers", methods=["GET"])
@premium_required
def list_suppliers():
    pid = get_pid(); db = get_db()
    rows = db.execute("SELECT * FROM suppliers WHERE profile_id=? ORDER BY name", (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/suppliers", methods=["POST"])
@premium_required
def add_supplier():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    name = d.get("name","").strip()
    if not name: return jsonify({"ok":False,"error":"Tedarikçi adı zorunlu"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO suppliers (user_id,profile_id,name,tax_no,contact,phone,email,address,payment_terms_days,notes,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (uid, pid, name, d.get("tax_no",""), d.get("contact",""), d.get("phone",""),
         d.get("email",""), d.get("address",""), int(d.get("payment_terms_days",30)),
         d.get("notes",""), datetime.now().isoformat())
    )
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/suppliers/<int:sid>", methods=["PUT"])
@premium_required
def update_supplier(sid):
    pid = get_pid(); db = get_db()
    d = request.get_json(force=True)
    fields, params = [], []
    for col in ("name","tax_no","contact","phone","email","address","payment_terms_days","notes"):
        if col in d:
            fields.append(f"{col}=?")
            params.append(d[col])
    if not fields: return jsonify({"ok":False}), 400
    params += [sid, pid]
    db.execute(f"UPDATE suppliers SET {','.join(fields)} WHERE id=? AND profile_id=?", params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/suppliers/<int:sid>", methods=["DELETE"])
@premium_required
def del_supplier(sid):
    pid = get_pid(); db = get_db()
    db.execute("DELETE FROM suppliers WHERE id=? AND profile_id=?", (sid, pid))
    db.commit()
    return jsonify({"ok":True})

# ── SUPPLIER INVOICES ─────────────────────────────────────────────────────────

@app.route("/api/supplier-invoices", methods=["GET"])
@premium_required
def list_supplier_invoices():
    pid = get_pid(); db = get_db()
    status = request.args.get("status","")
    if status:
        rows = db.execute(
            "SELECT * FROM supplier_invoices WHERE profile_id=? AND status=? ORDER BY due_date", (pid, status)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM supplier_invoices WHERE profile_id=? ORDER BY due_date DESC", (pid,)
        ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/supplier-invoices", methods=["POST"])
@premium_required
def add_supplier_invoice():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    sup_name = d.get("supplier_name","").strip()
    amount   = float(d.get("amount",0))
    inv_date = d.get("invoice_date", date.today().isoformat())
    due_date = d.get("due_date","")
    if not sup_name or amount <= 0 or not due_date:
        return jsonify({"ok":False,"error":"Tedarikçi, tutar ve vade zorunlu"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO supplier_invoices (user_id,profile_id,supplier_id,supplier_name,invoice_no,description,amount,currency,invoice_date,due_date,status,paid_date,reference_rate,notes,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (uid, pid, int(d.get("supplier_id",0)), sup_name, d.get("invoice_no",""),
         d.get("description",""), amount, d.get("currency","TRY"), inv_date, due_date,
         "bekliyor", "", float(d.get("reference_rate",50)), d.get("notes",""),
         datetime.now().isoformat())
    )
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/supplier-invoices/<int:iid>", methods=["PUT"])
@premium_required
def update_supplier_invoice(iid):
    pid = get_pid(); db = get_db()
    d = request.get_json(force=True)
    fields, params = [], []
    for col in ("supplier_name","invoice_no","description","amount","currency","invoice_date","due_date","status","paid_date","reference_rate","notes"):
        if col in d:
            fields.append(f"{col}=?")
            params.append(float(d[col]) if col in ("amount","reference_rate") else d[col])
    if not fields: return jsonify({"ok":False}), 400
    params += [iid, pid]
    db.execute(f"UPDATE supplier_invoices SET {','.join(fields)} WHERE id=? AND profile_id=?", params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/supplier-invoices/<int:iid>", methods=["DELETE"])
@premium_required
def del_supplier_invoice(iid):
    pid = get_pid(); db = get_db()
    db.execute("DELETE FROM supplier_invoices WHERE id=? AND profile_id=?", (iid, pid))
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/supplier-invoices/<int:iid>/pay", methods=["POST"])
@premium_required
def pay_supplier_invoice(iid):
    pid = get_pid(); db = get_db()
    d = request.get_json(force=True)
    paid_date = d.get("paid_date", date.today().isoformat())
    db.execute(
        "UPDATE supplier_invoices SET status='odendi',paid_date=? WHERE id=? AND profile_id=?",
        (paid_date, iid, pid)
    )
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/supplier-invoices/aging")
@premium_required
def supplier_invoices_aging():
    pid = get_pid(); db = get_db()
    today_d = date.today()
    rows = db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? AND status='bekliyor' ORDER BY due_date",
        (pid,)
    ).fetchall()
    buckets = {"0_30":[],"31_60":[],"61_90":[],"90plus":[]}
    totals  = {"0_30":0.0,"31_60":0.0,"61_90":0.0,"90plus":0.0}
    for r in rows:
        try:
            due = date.fromisoformat(r["due_date"])
            days_overdue = (today_d - due).days
        except:
            days_overdue = 0
        item = dict(r)
        item["days_overdue"] = days_overdue
        if days_overdue <= 0:
            buckets["0_30"].append(item); totals["0_30"] += float(r["amount"])
        elif days_overdue <= 30:
            buckets["0_30"].append(item); totals["0_30"] += float(r["amount"])
        elif days_overdue <= 60:
            buckets["31_60"].append(item); totals["31_60"] += float(r["amount"])
        elif days_overdue <= 90:
            buckets["61_90"].append(item); totals["61_90"] += float(r["amount"])
        else:
            buckets["90plus"].append(item); totals["90plus"] += float(r["amount"])
    grand_total = sum(totals.values())
    return jsonify({"ok":True,"buckets":buckets,"totals":totals,"grand_total":grand_total})

@app.route("/api/supplier-invoices/float-gain")
@premium_required
def supplier_invoices_float_gain():
    pid = get_pid(); db = get_db()
    today_d = date.today()
    rows = db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? AND status='odendi' AND paid_date!=''",
        (pid,)
    ).fetchall()
    results = []
    total_gain = 0.0
    for r in rows:
        try:
            due  = date.fromisoformat(r["due_date"])
            paid = date.fromisoformat(r["paid_date"])
            delayed_days = (paid - due).days
            if delayed_days > 0:
                rate = float(r["reference_rate"] or 50)
                gain = float(r["amount"]) * (delayed_days / 365) * (rate / 100)
                total_gain += gain
                results.append({
                    "supplier_name": r["supplier_name"],
                    "amount": float(r["amount"]),
                    "due_date": r["due_date"],
                    "paid_date": r["paid_date"],
                    "delayed_days": delayed_days,
                    "reference_rate": rate,
                    "float_gain": round(gain, 2)
                })
        except:
            continue
    results.sort(key=lambda x: x["float_gain"], reverse=True)
    return jsonify({"ok":True,"items":results,"total_gain":round(total_gain,2)})

# ── ASSETS & DEPRECIATION ─────────────────────────────────────────────────────

DEPRECIATION_RATES = {
    "arac":       {"rate": 20.0, "life": 5},
    "bilgisayar": {"rate": 33.33, "life": 3},
    "makine":     {"rate": 20.0, "life": 5},
    "mobilya":    {"rate": 20.0, "life": 5},
    "bina":       {"rate": 2.0,  "life": 50},
    "diger":      {"rate": 20.0, "life": 5},
}


@app.route("/api/assets", methods=["GET"])
@premium_required
def list_assets():
    pid = get_pid(); db = get_db()
    rows = db.execute("SELECT * FROM assets WHERE profile_id=? AND active=1 ORDER BY name", (pid,)).fetchall()
    result = []
    for r in rows:
        item = dict(r)
        dep = calc_depreciation(float(r["purchase_price"]), r["purchase_date"],
                                float(r["depreciation_rate"]), r["depreciation_method"])
        item["book_value"]   = dep["book_value"]
        item["accumulated"]  = dep["accumulated"]
        item["annual_dep"]   = dep["annual"]
        result.append(item)
    return jsonify(result)

@app.route("/api/assets", methods=["POST"])
@premium_required
def add_asset():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    name = d.get("name","").strip()
    if not name: return jsonify({"ok":False,"error":"Varlık adı zorunlu"}), 400
    atype = d.get("asset_type","diger")
    defaults = DEPRECIATION_RATES.get(atype, DEPRECIATION_RATES["diger"])
    rate = float(d.get("depreciation_rate", defaults["rate"]))
    life = int(d.get("useful_life_years", defaults["life"]))
    db = get_db()
    cur = db.execute(
        "INSERT INTO assets (user_id,profile_id,name,asset_type,plate_no,serial_no,purchase_date,purchase_price,useful_life_years,depreciation_method,depreciation_rate,maintenance_date,insurance_date,insurance_company,notes,active,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?)",
        (uid, pid, name, atype, d.get("plate_no",""), d.get("serial_no",""),
         d.get("purchase_date", date.today().isoformat()), float(d.get("purchase_price",0)),
         life, d.get("depreciation_method","normal"), rate,
         d.get("maintenance_date",""), d.get("insurance_date",""), d.get("insurance_company",""),
         d.get("notes",""), datetime.now().isoformat())
    )
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/assets/<int:aid>", methods=["PUT"])
@premium_required
def update_asset(aid):
    pid = get_pid(); db = get_db()
    d = request.get_json(force=True)
    fields, params = [], []
    for col in ("name","asset_type","plate_no","serial_no","purchase_date","purchase_price",
                "useful_life_years","depreciation_method","depreciation_rate",
                "maintenance_date","insurance_date","insurance_company","notes","active"):
        if col in d:
            fields.append(f"{col}=?")
            params.append(float(d[col]) if col in ("purchase_price","depreciation_rate") else d[col])
    if not fields: return jsonify({"ok":False}), 400
    params += [aid, pid]
    db.execute(f"UPDATE assets SET {','.join(fields)} WHERE id=? AND profile_id=?", params)
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/assets/<int:aid>", methods=["DELETE"])
@premium_required
def del_asset(aid):
    pid = get_pid(); db = get_db()
    db.execute("UPDATE assets SET active=0 WHERE id=? AND profile_id=?", (aid, pid))
    db.commit()
    return jsonify({"ok":True})

@app.route("/api/assets/<int:aid>/maintenance", methods=["GET"])
@premium_required
def list_asset_maintenance(aid):
    pid = get_pid(); db = get_db()
    rows = db.execute(
        "SELECT * FROM asset_maintenance WHERE profile_id=? AND asset_id=? ORDER BY date DESC",
        (pid, aid)
    ).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route("/api/assets/<int:aid>/maintenance", methods=["POST"])
@premium_required
def add_asset_maintenance(aid):
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    dt = d.get("date", date.today().isoformat())
    db = get_db()
    cur = db.execute(
        "INSERT INTO asset_maintenance (user_id,profile_id,asset_id,date,mtype,description,cost,odometer,next_date,notes,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (uid, pid, aid, dt, d.get("mtype","bakim"), d.get("description",""),
         float(d.get("cost",0)), float(d.get("odometer",0)),
         d.get("next_date",""), d.get("notes",""), datetime.now().isoformat())
    )
    db.commit()
    # Update asset maintenance_date
    db.execute("UPDATE assets SET maintenance_date=? WHERE id=? AND profile_id=?",
               (d.get("next_date",""), aid, pid))
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/assets/maintenance/<int:mid>", methods=["DELETE"])
@premium_required
def del_asset_maintenance(mid):
    pid = get_pid(); db = get_db()
    db.execute("DELETE FROM asset_maintenance WHERE id=? AND profile_id=?", (mid, pid))
    db.commit()
    return jsonify({"ok":True})

# ── CARD DAILY BALANCE ────────────────────────────────────────────────────────

@app.route("/api/cards/daily-report", methods=["GET"])
@login_required
def card_daily_report():
    pid = get_pid(); db = get_db()
    cards = db.execute("SELECT * FROM cards WHERE profile_id=? ORDER BY bank_name", (pid,)).fetchall()
    today_str = date.today().isoformat()
    yesterday_str = (date.today() - __import__("datetime").timedelta(days=1)).isoformat()
    result = []
    for c in cards:
        cid = c["id"]
        today_row = db.execute(
            "SELECT balance FROM card_daily_balance WHERE profile_id=? AND card_id=? AND date=?",
            (pid, cid, today_str)
        ).fetchone()
        yest_row = db.execute(
            "SELECT balance FROM card_daily_balance WHERE profile_id=? AND card_id=? AND date=?",
            (pid, cid, yesterday_str)
        ).fetchone()
        today_bal = float(today_row["balance"]) if today_row else None
        yest_bal  = float(yest_row["balance"])  if yest_row  else None
        spent = None
        if today_bal is not None and yest_bal is not None:
            spent = today_bal - yest_bal
        result.append({
            "card_id": cid,
            "bank_name": c["bank_name"],
            "card_name": c["card_name"] or "",
            "today_balance": today_bal,
            "yesterday_balance": yest_bal,
            "spent_today": round(spent, 2) if spent is not None else None
        })
    return jsonify({"ok":True,"cards":result,"date":today_str})

@app.route("/api/cards/daily-balance", methods=["POST"])
@login_required
def save_card_daily_balance():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    cid = int(d.get("card_id",0))
    bal = float(d.get("balance",0))
    dt  = d.get("date", date.today().isoformat())
    if not cid: return jsonify({"ok":False,"error":"Kart gerekli"}), 400
    db = get_db()
    # Upsert
    existing = db.execute(
        "SELECT id FROM card_daily_balance WHERE profile_id=? AND card_id=? AND date=?",
        (pid, cid, dt)
    ).fetchone()
    if existing:
        db.execute("UPDATE card_daily_balance SET balance=?,notes=? WHERE id=?",
                   (bal, d.get("notes",""), existing["id"]))
    else:
        db.execute(
            "INSERT INTO card_daily_balance (user_id,profile_id,card_id,date,balance,notes,created_at) VALUES (?,?,?,?,?,?,?)",
            (uid, pid, cid, dt, bal, d.get("notes",""), datetime.now().isoformat())
        )
    db.commit()
    return jsonify({"ok":True})

# ── TELEGRAM BOT ──────────────────────────────────────────────────────────────

_TG_MAIN_KEYBOARD = {
    "keyboard": [
        [{"text": "💰 Bakiye"}, {"text": "📅 Bugün"}, {"text": "📆 Bu Ay"}],
        [{"text": "💳 Kartlar"}, {"text": "📊 Bütçe"}, {"text": "📈 Yatırım"}],
        [{"text": "🗒️ Son 5"},   {"text": "🎯 Hedefler"}, {"text": "❓ Yardım"}],
        [{"text": "🗑️ Son İşlemi Sil"}],
    ],
    "resize_keyboard": True,
    "persistent": True,
}

def tg_send(chat_id, text, parse_mode="HTML", reply_markup=None, show_keyboard=True):
    if not TELEGRAM_TOKEN: return
    payload = {"chat_id": chat_id, "text": text, "parse_mode": parse_mode}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    elif show_keyboard:
        payload["reply_markup"] = _TG_MAIN_KEYBOARD
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                      json=payload, timeout=6)
    except Exception: pass

def tg_answer_cb(cb_id, text=""):
    if not TELEGRAM_TOKEN: return
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/answerCallbackQuery",
                      json={"callback_query_id": cb_id, "text": text}, timeout=5)
    except Exception: pass

def tg_edit_msg(chat_id, msg_id, text):
    if not TELEGRAM_TOKEN: return
    try:
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/editMessageText",
                      json={"chat_id": chat_id, "message_id": msg_id,
                            "text": text, "parse_mode": "HTML"}, timeout=5)
    except Exception: pass

_TG_CAT_KW = [
    # Gider kategorileri (uzun anahtar kelimeler önce — öncelikli eşleşme için)
    ("süpermarket","Market / Gıda"),("supermarket","Market / Gıda"),
    ("carrefour","Market / Gıda"),("migros","Market / Gıda"),("bim","Market / Gıda"),
    ("a101","Market / Gıda"),("şok market","Market / Gıda"),("market","Market / Gıda"),("gıda","Market / Gıda"),("manav","Market / Gıda"),
    ("restoran","Yemek / Restoran"),("starbucks","Yemek / Restoran"),("mcdonalds","Yemek / Restoran"),
    ("burger king","Yemek / Restoran"),("pizza","Yemek / Restoran"),("döner","Yemek / Restoran"),
    ("kahvaltı","Yemek / Restoran"),("yemek","Yemek / Restoran"),("kahve","Yemek / Restoran"),("kafe","Yemek / Restoran"),
    ("elektrik","Faturalar"),("doğalgaz","Faturalar"),("su faturası","Faturalar"),
    ("internet faturası","Faturalar"),("gsm","Faturalar"),("turkcell","Faturalar"),
    ("vodafone","Faturalar"),("türk telekom","Faturalar"),("fatura","Faturalar"),
    ("benzin","Ulaşım"),("akaryakıt","Ulaşım"),("shell","Ulaşım"),("opet","Ulaşım"),
    ("taksi","Ulaşım"),("uber","Ulaşım"),("metro","Ulaşım"),("otobüs","Ulaşım"),
    ("bilet","Ulaşım"),("ulaşım","Ulaşım"),("araç","Ulaşım"),
    ("konut kredisi","Konut Kredisi Taksiti"),("eminevim","Konut Kredisi Taksiti"),
    ("finansevim","Konut Kredisi Taksiti"),("mortgage","Konut Kredisi Taksiti"),
    ("araç kredisi","Araç Kredisi Taksiti"),("araba kredisi","Araç Kredisi Taksiti"),
    ("ihtiyaç kredisi","İhtiyaç Kredisi"),("kredi taksit","İhtiyaç Kredisi"),
    ("sigorta","Sigorta"),("kasko","Sigorta"),("trafik sigortası","Sigorta"),
    ("vergi","Vergi / Harç"),("kdv","Vergi / Harç"),("sgk","Vergi / Harç"),
    ("kira","Kira / Mortgage"),
    ("eczane","Sağlık"),("hastane","Sağlık"),("doktor","Sağlık"),("ilaç","Sağlık"),("sağlık","Sağlık"),
    ("netflix","Abonelikler"),("spotify","Abonelikler"),("youtube","Abonelikler"),
    ("apple","Abonelikler"),("abonelik","Abonelikler"),
    ("sinema","Eğlence"),("tiyatro","Eğlence"),("eğlence","Eğlence"),("oyun","Eğlence"),
    ("kıyafet","Giyim"),("ayakkabı","Giyim"),("giyim","Giyim"),("mağaza","Giyim"),("zara","Giyim"),
    ("telefon","Elektronik"),("laptop","Elektronik"),("elektronik","Elektronik"),
    ("kurs","Eğitim"),("kitap","Eğitim"),("eğitim","Eğitim"),
    # Gelir kategorileri
    ("maaş","Maaş"),("maas","Maaş"),
    ("kira geliri","Kira Geliri"),
    ("temettü","Yatırım / Temettü"),("faiz","Yatırım / Temettü"),("yatırım","Yatırım / Temettü"),
    ("ikramiye","Hediye / İkramiye"),("prim","Hediye / İkramiye"),("hediye","Hediye / İkramiye"),
    ("freelance","Serbest Meslek"),("serbest","Serbest Meslek"),("danışmanlık","Serbest Meslek"),
]
_GELIR_CATS = set(GELIR_CATS)  # sunucu listesiyle senkron

def _parse_tg_amount(text_l):
    """Parse Turkish amount: '10 bin' → 10000, '2.5 milyon' → 2500000, '500' → 500"""
    # X bin Y (örn: 10 bin 500 → 10500)
    m = re.search(r'(\d[\d.,]*)\s*bin\s+(\d+)', text_l)
    if m:
        try: return float(m.group(1).replace(',','.')) * 1000 + float(m.group(2))
        except: pass
    # X bin (örn: 10 bin → 10000)
    m = re.search(r'(\d[\d.,]*)\s*bin\b', text_l)
    if m:
        try: return float(m.group(1).replace(',','.')) * 1000
        except: pass
    # X milyon
    m = re.search(r'(\d[\d.,]*)\s*milyon\b', text_l)
    if m:
        try: return float(m.group(1).replace(',','.')) * 1_000_000
        except: pass
    # Xk (örn: 10k → 10000)
    m = re.search(r'(\d[\d.,]*)\s*k\b', text_l)
    if m:
        try: return float(m.group(1).replace(',','.')) * 1000
        except: pass
    # Noktalı/virgüllü sayı (12.000 veya 12,000 → 12000)
    m = re.search(r'\b(\d{1,3}(?:[.,]\d{3})+)\s*(?:tl|₺|lira)?\b', text_l)
    if m:
        try: return float(re.sub(r'[.,](?=\d{3}\b)', '', m.group(1)).replace(',','.'))
        except: pass
    # Normal sayı
    m = re.search(r'\b(\d[\d]*(?:[.,]\d+)?)\s*(?:tl|₺|lira)?\b', text_l)
    if m:
        try: return float(m.group(1).replace(',','.'))
        except: pass
    return None

def _parse_tg_msg(text):
    text_l = text.lower().strip()
    amount = _parse_tg_amount(text_l)
    if not amount or amount <= 0: return None
    # Miktarı ve birim kelimelerini temizle
    no_amt = re.sub(r'\d[\d.,]*\s*(?:bin|milyon|k|tl|₺|lira)?\b', '', text_l).strip()
    category = None
    for kw, cat in _TG_CAT_KW:
        if kw in no_amt or kw in text_l:
            category = cat; break
    if not category:
        category = "Diğer Gider"
    ttype = "gelir" if category in _GELIR_CATS else "gider"
    # Açıklama: sayılar hariç orijinal kelimeler
    stop = {'tl','₺','lira','bin','milyon','k'}
    desc_words = [w for w in text.split() if not re.match(r'^\d[\d.,]*$', w) and w.lower() not in stop]
    return {"type": ttype, "amount": amount, "category": category,
            "description": ' '.join(desc_words) or text}

TELEGRAM_WEBHOOK_SECRET = os.environ.get("TELEGRAM_WEBHOOK_SECRET", "")

# ── TELEGRAM YARDIMCI FONKSİYONLARI ──────────────────────────────────────────

def _tg_fmt_amount(n):
    return f"₺{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def _tg_parse_date(text_l):
    """'dün', 'bugün', '2 gün önce' gibi ifadeleri tarihe çevirir. None → bugün."""
    today = date.today()
    if "dün" in text_l or "dun" in text_l:
        return today - timedelta(days=1)
    if "evvelsi" in text_l:
        return today - timedelta(days=2)
    m = re.search(r'(\d+)\s*gün\s*önce', text_l)
    if m:
        return today - timedelta(days=int(m.group(1)))
    return today

def _tg_strip_date_words(text_l):
    text_l = re.sub(r'\d+\s*gün\s*önce', '', text_l)
    return text_l.replace("dün","").replace("dun","").replace("bugün","").replace("evvelsi gün","").strip()

def _tg_parse_category_override(text):
    """'#kategori' ifadesini bulur ve döner (None yoksa)."""
    m = re.search(r'#([^\s]+)', text)
    return m.group(1) if m else None

def _tg_match_category(kw_text):
    """Keyword'den kategori bul."""
    for kw, cat in _TG_CAT_KW:
        if kw in kw_text:
            return cat
    return None

_MONTHS_TR_MAP = {
    "ocak":1,"january":1,"jan":1,
    "şubat":2,"subat":2,"february":2,"feb":2,
    "mart":3,"march":3,"mar":3,
    "nisan":4,"april":4,"apr":4,
    "mayıs":5,"mayis":5,"may":5,
    "haziran":6,"june":6,"jun":6,
    "temmuz":7,"july":7,"jul":7,
    "ağustos":8,"agustos":8,"august":8,"aug":8,
    "eylül":9,"eylul":9,"september":9,"sep":9,
    "ekim":10,"october":10,"oct":10,
    "kasım":11,"kasim":11,"november":11,"nov":11,
    "aralık":12,"aralik":12,"december":12,"dec":12,
}

def _tg_parse_month(text_l):
    """'haziran', 'geçen ay', 'mayıs 2024' gibi ifadelerden (year,month) döner. None→bugün."""
    today = date.today()
    if "geçen ay" in text_l or "gecen ay" in text_l:
        if today.month == 1:
            return today.year - 1, 12
        return today.year, today.month - 1
    if "önceki ay" in text_l or "onceki ay" in text_l:
        if today.month == 1:
            return today.year - 1, 12
        return today.year, today.month - 1
    if "bu ay" in text_l:
        return today.year, today.month
    # "haziran 2024" veya "haziran" gibi
    for mn, mi in _MONTHS_TR_MAP.items():
        if mn in text_l:
            # Yıl var mı?
            ym = re.search(r'\b(20\d{2})\b', text_l)
            yr = int(ym.group(1)) if ym else today.year
            # Gelecek yılsa geçen yıla çek (tahmin)
            if not ym and yr == today.year and mi > today.month:
                yr -= 1
            return yr, mi
    return None

def _tg_summary_msg(uid, pid, db, period="bugun", year=None, month=None):
    """Bakiye veya özet mesajı oluştur."""
    today = date.today()
    if period == "bugun":
        rows = db.execute(
            "SELECT type,amount,category,description FROM transactions WHERE profile_id=%s AND date=%s ORDER BY id DESC",
            (pid, today.isoformat())
        ).fetchall()
        if not rows:
            return "📅 <b>Bugün henüz işlem yok.</b>"
        gelir = sum(r["amount"] for r in rows if r["type"]=="gelir")
        gider = sum(r["amount"] for r in rows if r["type"]=="gider")
        lines = [f"📅 <b>Bugün — {today.strftime('%d.%m.%Y')}</b>\n"]
        for r in rows:
            ico = "📈" if r["type"]=="gelir" else "📉"
            lines.append(f"{ico} {r['category']} — <b>{_tg_fmt_amount(r['amount'])}</b>"
                        + (f" <i>({r['description']})</i>" if r["description"] else ""))
        lines.append(f"\n📊 Gelir: <b>{_tg_fmt_amount(gelir)}</b>  |  Gider: <b>{_tg_fmt_amount(gider)}</b>  |  Net: <b>{_tg_fmt_amount(gelir-gider)}</b>")
        return "\n".join(lines)

    elif period == "ay":
        import calendar as _cal
        yr  = year  or today.year
        mo  = month or today.month
        s   = f"{yr}-{mo:02d}-01"
        e   = f"{yr}-{mo:02d}-{_cal.monthrange(yr, mo)[1]:02d}"
        rows = db.execute(
            "SELECT type,SUM(amount) as t,category FROM transactions WHERE profile_id=%s AND date BETWEEN %s AND %s GROUP BY type,category ORDER BY t DESC",
            (pid, s, e)
        ).fetchall()
        gelir = sum(r["t"] for r in rows if r["type"]=="gelir")
        gider = sum(r["t"] for r in rows if r["type"]=="gider")
        MONTHS_TR = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran","Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]
        net = gelir - gider
        is_current = (yr == today.year and mo == today.month)
        period_lbl = MONTHS_TR[mo] + (" " + str(yr) if yr != today.year else "")
        lines = [f"📆 <b>{period_lbl} Özeti</b>"
                 + (f" ({today.day}. güne kadar)" if is_current else "") + "\n"]
        lines.append(f"📈 Toplam Gelir: <b>{_tg_fmt_amount(gelir)}</b>")
        lines.append(f"📉 Toplam Gider: <b>{_tg_fmt_amount(gider)}</b>")
        lines.append(f"{'✅' if net>=0 else '⚠️'} Net Kalan: <b>{_tg_fmt_amount(net)}</b>")
        if gelir > 0:
            tasarruf_pct = round(net/gelir*100) if gelir else 0
            lines.append(f"💡 Tasarruf oranı: %{tasarruf_pct}")
        gider_cats = [(r["category"], r["t"]) for r in rows if r["type"]=="gider"]
        gider_cats.sort(key=lambda x: -x[1])
        if gider_cats:
            lines.append("\n<b>En Büyük Giderler:</b>")
            for cat, amt in gider_cats[:5]:
                pct = round(amt/gider*100) if gider else 0
                lines.append(f"  • {cat}: {_tg_fmt_amount(amt)} (%{pct})")
        if not rows:
            lines.append("\n(Bu dönemde işlem bulunamadı)")
        return "\n".join(lines)

    elif period == "bakiye":
        # Hesap bakiyesi
        acc = db.execute(
            "SELECT COALESCE(SUM(initial_balance),0) as ib FROM accounts WHERE profile_id=%s AND active=1", (pid,)
        ).fetchone()
        tx_bal = db.execute(
            "SELECT COALESCE(SUM(CASE WHEN type='gelir' THEN amount ELSE -amount END),0) as b FROM transactions WHERE profile_id=%s AND date<=%s",
            (pid, today.isoformat())
        ).fetchone()
        card_debt = db.execute(
            "SELECT COALESCE(SUM(used_),0) as d FROM cards WHERE profile_id=%s", (pid,)
        ).fetchone()
        invest = db.execute(
            "SELECT COALESCE(SUM(quantity*buy_price),0) as v FROM investments WHERE profile_id=%s", (pid,)
        ).fetchone()
        hesap = float((acc["ib"] or 0)) + float((tx_bal["b"] or 0))
        borç  = float(card_debt["d"] or 0)
        yat   = float(invest["v"] or 0)
        likid = hesap - borç
        net_w = hesap + yat - borç
        lines = ["💰 <b>Anlık Finansal Durum</b>\n",
                 f"🏦 Hesap Bakiyesi: <b>{_tg_fmt_amount(hesap)}</b>",
                 f"💳 Kart Borcu: <b>-{_tg_fmt_amount(borç)}</b>",
                 f"{'✅' if likid>=0 else '⚠️'} Net Likidite: <b>{_tg_fmt_amount(likid)}</b>"]
        if yat > 0:
            lines.append(f"📈 Yatırım (maliyet): <b>{_tg_fmt_amount(yat)}</b>")
            lines.append(f"📊 Net Varlık: <b>{_tg_fmt_amount(net_w)}</b>")
        return "\n".join(lines)

    return "❓"

def _tg_last_tx(uid, pid, db, n=5):
    rows = db.execute(
        "SELECT id,type,amount,category,description,date FROM transactions WHERE profile_id=%s ORDER BY date DESC,id DESC LIMIT %s",
        (pid, n)
    ).fetchall()
    if not rows:
        return "📋 Henüz işlem yok."
    lines = [f"🗒️ <b>Son {len(rows)} İşlem</b>\n"]
    for r in rows:
        ico = "📈" if r["type"]=="gelir" else "📉"
        lines.append(f"{ico} <code>#{r['id']}</code> {r['date'][5:]} — {r['category']} <b>{_tg_fmt_amount(r['amount'])}</b>"
                    + (f" <i>{r['description']}</i>" if r["description"] else ""))
    return "\n".join(lines)

def _tg_budget_msg(pid, db):
    today = date.today()
    s = f"{today.year}-{today.month:02d}-01"
    import calendar; last = calendar.monthrange(today.year, today.month)[1]
    e = f"{today.year}-{today.month:02d}-{last:02d}"
    budgets = {r["category"]:r["limit_"] for r in db.execute("SELECT category,limit_ FROM budgets WHERE profile_id=%s",(pid,)).fetchall()}
    if not budgets:
        return "📊 Bütçe tanımlanmamış.\nKirpi uygulamasından kategori bütçesi belirleyin."
    spending = {}
    for r in db.execute("SELECT category,SUM(amount) as t FROM transactions WHERE profile_id=%s AND type='gider' AND date BETWEEN %s AND %s GROUP BY category",(pid,s,e)).fetchall():
        spending[r["category"]] = r["t"]
    lines = [f"📊 <b>Bütçe Durumu — {today.month:02d}/{today.year}</b>\n"]
    for cat, limit in sorted(budgets.items()):
        spent = spending.get(cat, 0)
        pct = round(spent/limit*100) if limit else 0
        bar = "█"*(pct//10) + "░"*(10-min(pct//10,10))
        warn = "🔴" if pct>=100 else "🟡" if pct>=80 else "🟢"
        lines.append(f"{warn} <b>{cat}</b>\n   {bar} %{pct}\n   {_tg_fmt_amount(spent)} / {_tg_fmt_amount(limit)}")
    return "\n".join(lines)

_CTYPE_ICO = {"kredi": "💳", "banka": "🏧", "yemek": "🍽️", "hediye": "🎁"}

def _tg_parse_card_payment(text_l):
    """Kart ödeme mesajını ayrıştır. None döner veya dict."""
    pay_kws = ["ödeme", "ödendi", "öde", "odeme", "odendi", "odedim", "ödedim"]
    if not any(k in text_l for k in pay_kws):
        return None
    # Kart/kredi ipucu
    card_kws = ["kart", "kredi", "card"]
    if not any(k in text_l for k in card_kws):
        return None
    is_asgari = any(k in text_l for k in ["asgari","asgarisi","minimum","min ","min."])
    is_tam    = any(k in text_l for k in ["tam ödeme","tüm borç","tum borc","tam borç","tamamını","tamamini","hepsini","hepsi"])
    amount = None
    m = re.search(r'(\d[\d.]*(?:[.,]\d+)?)\s*(?:tl|₺|lira)?', text_l)
    if m:
        try: amount = float(m.group(1).replace('.','').replace(',','.'))
        except: pass
    return {"is_asgari": is_asgari, "is_tam": is_tam, "amount": amount, "raw": text_l}

def _tg_find_card(db, pid, text_l):
    """Metinden kart bul (banka adı fuzzy eşleme)."""
    cards = db.execute(
        "SELECT * FROM cards WHERE profile_id=%s AND card_type IN ('kredi','banka') ORDER BY bank_name", (pid,)
    ).fetchall()
    if not cards: return None
    import unicodedata
    def norm(s): return unicodedata.normalize('NFKD',s.lower()).encode('ascii','ignore').decode()
    text_n = norm(text_l)
    best, best_score = None, 0
    for c in cards:
        bank_n = norm(c['bank_name'] or '')
        cname_n = norm(c['card_name'] or '')
        score = sum(2 for w in bank_n.split() if len(w)>2 and w in text_n)
        score += sum(1 for w in cname_n.split() if len(w)>2 and w in text_n)
        if score > best_score:
            best_score, best = score, c
    return best if best_score > 0 else None

def _tg_cards_msg(pid, db):
    cards = db.execute("SELECT * FROM cards WHERE profile_id=%s ORDER BY bank_name",(pid,)).fetchall()
    if not cards:
        return "💳 Kayıtlı kart yok."
    today = date.today()
    # Türe göre grupla
    groups = {"kredi": [], "banka": [], "yemek": [], "hediye": []}
    for c in cards:
        groups.setdefault(c.get("card_type","kredi"), []).append(c)
    lines = ["💳 <b>Kartlarım</b>\n"]
    for ctype, group in groups.items():
        if not group: continue
        ico = _CTYPE_ICO.get(ctype, "💳")
        type_name = {"kredi":"Kredi Kartları","banka":"Banka Kartları","yemek":"Yemek Kartları","hediye":"Hediye Kartları"}.get(ctype,"Kartlar")
        lines.append(f"\n{ico} <b>{type_name}</b>")
        for c in group:
            used = c["used_"] or 0
            lim  = c["limit_"] or 0
            pct  = round(used/lim*100) if lim else 0
            warn = "🔴" if pct>=90 else "🟡" if pct>=70 else "🟢"
            card_label = f"{c['bank_name']}" + (f" {c['card_name']}" if c.get('card_name') else "")
            if ctype in ("yemek","hediye"):
                lines.append(f"{warn} <b>{card_label}</b>\n"
                             f"   Kalan: {_tg_fmt_amount(lim-used)} / {_tg_fmt_amount(lim)} (%{100-pct})")
            else:
                due_day = c["due_day"] or 1
                if today.day <= due_day:
                    due_date = today.replace(day=due_day)
                else:
                    if today.month == 12:
                        due_date = date(today.year+1, 1, due_day)
                    else:
                        import calendar
                        last = calendar.monthrange(today.year, today.month+1)[1]
                        due_date = date(today.year, today.month+1, min(due_day, last))
                days_left = (due_date - today).days
                lines.append(f"{warn} <b>{card_label}</b>\n"
                             f"   Borç: {_tg_fmt_amount(used)} / {_tg_fmt_amount(lim)} (%{pct})\n"
                             f"   Son ödeme: {due_date.strftime('%d.%m')} ({days_left} gün)")
    return "\n".join(lines)

@app.route("/api/telegram/webhook", methods=["POST"])
def telegram_webhook():
    if not TELEGRAM_TOKEN: return "ok"
    if TELEGRAM_WEBHOOK_SECRET:
        incoming = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
        if not secrets.compare_digest(incoming, TELEGRAM_WEBHOOK_SECRET):
            return "Forbidden", 403
    data = request.get_json(force=True, silent=True) or {}

    # ── Callback query (Evet/Hayır butonları) ──
    cb = data.get("callback_query")
    if cb:
        cb_id   = cb["id"]
        cdata   = cb.get("data","")
        chat_id = cb["message"]["chat"]["id"]
        msg_id  = cb["message"]["message_id"]
        tg_uid  = str(cb["from"]["id"])
        db = pg_connect()
        try:
            if cdata.startswith("confirm:"):
                parts = cdata.split(":")
                pid_pending = int(parts[1])
                card_id = int(parts[3]) if len(parts) >= 4 and parts[2] == "c" else None
                row = db.execute(
                    "SELECT * FROM telegram_pending WHERE id=%s AND tg_user_id=%s",
                    (pid_pending, tg_uid)
                ).fetchone()
                if row:
                    # Kart bilgisini önceden al (display için)
                    card_info = None
                    if card_id:
                        card_info = db.execute(
                            "SELECT bank_name,card_name FROM cards WHERE id=%s AND profile_id=%s",
                            (card_id, row["profile_id"])
                        ).fetchone()
                        if not card_info: card_id = None  # güvenlik: başka profil
                    tx_date_val = row.get("tx_date") or date.today().isoformat()
                    db.execute(
                        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,card_id,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (row["user_id"], row["profile_id"], row["type"], row["amount"],
                         row["category"], row["description"], tx_date_val,
                         card_id, datetime.now().isoformat())
                    )
                    if card_id:
                        if row["type"] == "gider":
                            db.execute("UPDATE cards SET used_=used_+%s WHERE id=%s", (row["amount"], card_id))
                        else:
                            db.execute("UPDATE cards SET used_=GREATEST(0,used_-%s) WHERE id=%s", (row["amount"], card_id))
                    db.execute("DELETE FROM telegram_pending WHERE id=%s", (pid_pending,))
                    db.commit()
                    icon = "📈" if row["type"] == "gelir" else "📉"
                    pay_str = ""
                    if card_info:
                        pay_str = f"\n💳 Kart: {card_info['bank_name']}" + (f" {card_info['card_name']}" if card_info.get("card_name") else "")
                    else:
                        pay_str = "\n💵 Nakit"
                    tg_edit_msg(chat_id, msg_id,
                        f"{icon} <b>Kaydedildi!</b>\n\n"
                        f"{'Gelir' if row['type']=='gelir' else 'Gider'}: <b>₺{row['amount']:,.2f}</b>\n"
                        f"Kategori: {row['category']}\n"
                        f"Not: {row['description']}"
                        + pay_str
                    )
                    tg_answer_cb(cb_id, "✅ Kaydedildi!")
                else:
                    tg_answer_cb(cb_id, "Bu işlem zaten işlendi.")
            elif cdata.startswith("cancel:"):
                pid_pending = int(cdata.split(":")[1])
                db.execute("DELETE FROM telegram_pending WHERE id=%s AND tg_user_id=%s", (pid_pending, tg_uid))
                db.commit()
                tg_edit_msg(chat_id, msg_id, "❌ İptal edildi.")
                tg_answer_cb(cb_id, "İptal edildi.")
            elif cdata.startswith("cardpay:"):
                parts = cdata.split(":")
                if parts[1] == "cancel":
                    tg_edit_msg(chat_id, msg_id, "❌ Kart ödemesi iptal edildi.")
                    tg_answer_cb(cb_id, "İptal.")
                else:
                    card_id = int(parts[1]); pay_amount = float(parts[2])
                    link2 = db.execute("SELECT user_id,profile_id FROM telegram_links WHERE tg_user_id=%s",(tg_uid,)).fetchone()
                    if link2:
                        card = db.execute("SELECT * FROM cards WHERE id=%s AND profile_id=%s",(card_id,link2["profile_id"])).fetchone()
                        if card:
                            cur_debt = float(card["used_"] or 0)
                            paid = round(min(pay_amount, cur_debt), 2)
                            new_debt = round(max(0, cur_debt - paid), 2)
                            cat = "Yemek Kartı Ödemesi" if card.get("card_type")=="yemek" else "Kredi Kartı Ödemesi"
                            desc = f"{card['bank_name']} {card.get('card_name') or ''} ödemesi".strip()
                            db.execute(
                                "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                                (link2["user_id"],link2["profile_id"],"gider",paid,cat,desc,date.today().isoformat(),datetime.now().isoformat())
                            )
                            db.execute("UPDATE cards SET used_=%s WHERE id=%s",(new_debt,card_id))
                            db.commit()
                            clbl = card['bank_name']+(f" {card['card_name']}" if card.get('card_name') else "")
                            tg_edit_msg(chat_id,msg_id,
                                f"✅ <b>Kart ödemesi kaydedildi!</b>\n\n"
                                f"💳 {clbl}\n"
                                f"💸 Ödenen: {_tg_fmt_amount(paid)}\n"
                                f"📉 Kalan borç: {_tg_fmt_amount(new_debt)}"
                            )
                            tg_answer_cb(cb_id,"✅ Kaydedildi!")
                        else:
                            tg_answer_cb(cb_id,"Kart bulunamadı.")
                    else:
                        tg_answer_cb(cb_id,"Yetki hatası.")
            elif cdata.startswith("deltx:"):
                val = cdata.split(":")[1]
                if val == "cancel":
                    tg_edit_msg(chat_id, msg_id, "↩️ Silme iptal edildi.")
                    tg_answer_cb(cb_id, "İptal.")
                else:
                    tx_id = int(val)
                    # Güvenlik: bu işlem kullanıcıya ait mi?
                    link2 = db.execute("SELECT user_id,profile_id FROM telegram_links WHERE tg_user_id=%s",(tg_uid,)).fetchone()
                    if link2:
                        row = db.execute("SELECT * FROM transactions WHERE id=%s AND profile_id=%s",(tx_id,link2["profile_id"])).fetchone()
                        if row:
                            db.execute("DELETE FROM transactions WHERE id=%s",(tx_id,))
                            db.commit()
                            tg_edit_msg(chat_id, msg_id, f"🗑 <b>Silindi!</b>\n{row['category']} — {_tg_fmt_amount(row['amount'])}")
                            tg_answer_cb(cb_id, "Silindi!")
                        else:
                            tg_answer_cb(cb_id, "İşlem bulunamadı.")
                    else:
                        tg_answer_cb(cb_id, "Yetki hatası.")
        except Exception as e:
            log.error("Telegram callback error: %s", e)
            tg_answer_cb(cb_id, "Hata oluştu.")
        finally:
            db.close()
        return "ok"

    # ── Normal mesaj ──
    msg = data.get("message") or data.get("edited_message")
    if not msg: return "ok"
    chat_id = msg["chat"]["id"]
    text    = (msg.get("text") or "").strip()
    tg_uid  = str(msg["from"]["id"])
    db = pg_connect()
    try:
        if text.lower().startswith("/link "):
            code = text[6:].strip().upper()
            pending = db.execute(
                "SELECT user_id, profile_id FROM telegram_link_codes WHERE code=%s AND expires_at > %s",
                (code, datetime.now().isoformat())
            ).fetchone()
            if pending:
                uid, pid = pending["user_id"], pending["profile_id"]
                existing = db.execute("SELECT id FROM telegram_links WHERE tg_user_id=%s", (tg_uid,)).fetchone()
                if existing:
                    db.execute("UPDATE telegram_links SET user_id=%s, profile_id=%s WHERE tg_user_id=%s", (uid, pid, tg_uid))
                else:
                    db.execute("INSERT INTO telegram_links (user_id,profile_id,tg_user_id,created_at) VALUES (%s,%s,%s,%s)",
                               (uid, pid, tg_uid, datetime.now().isoformat()))
                db.execute("DELETE FROM telegram_link_codes WHERE code=%s", (code,))
                db.commit()
                tg_send(chat_id,
                    "✅ <b>Hesabınız bağlandı!</b>\n\n"
                    "Aşağıdaki tuşları kullanabilir ya da doğrudan yazabilirsin:\n\n"
                    "• <code>market 250</code> → gider\n"
                    "• <code>maaş 15000</code> → gelir\n"
                    "• <code>dün benzin 500</code> → tarihli kayıt\n\n"
                    "❓ Yardım için <b>yardım</b> yaz.")
            else:
                tg_send(chat_id, "❌ Geçersiz veya süresi dolmuş kod.\nKirpi Ayarlar sayfasından yeni kod alın.")
            return "ok"

        text_l = text.lower().strip()

        if text_l == "/start":
            tg_send(chat_id,
                "🦔 <b>Kirpi Finans Bot'a hoş geldiniz!</b>\n\n"
                "<b>Önce hesabınızı bağlayın:</b>\n"
                "1. Kirpi → Ayarlar → Telegram → Bağlantı Kodu Al\n"
                "2. Buraya <code>/link KOD</code> gönderin\n\n"
                "Bağlandıktan sonra /help yazın.")
            return "ok"

        if text_l in ("/help", "/yardim", "yardim", "yardım", "ne yapabilirsin",
                      "komutlar", "ne yazabilirim", "❓ yardım"):
            tg_send(chat_id,
                "📖 <b>Kirpi Bot — Ne Yapabilirsin?</b>\n\n"
                "<b>💸 İşlem Kaydetme</b> (sadece yaz, slash gerekmez):\n"
                "  <code>market 250</code>\n"
                "  <code>maaş 15000</code>\n"
                "  <code>dün benzin 500</code>\n"
                "  <code>2 gün önce kira 8000</code>\n"
                "  <code>yemek 350 #eğlence</code> → kategori belirle\n"
                "  <code>gelir 5000 serbest iş</code>\n"
                "  <code>gider 300 eczane</code>\n\n"
                "<b>📊 Sorgular</b> (slash'sız da yazabilirsin):\n"
                "  <code>bakiye</code> → anlık finansal durum\n"
                "  <code>bugun</code> → bugünkü işlemler\n"
                "  <code>son</code> → son 5 işlem\n"
                "  <code>ay</code> → bu ayın özeti\n"
                "  <code>butce</code> → bütçe durumu\n"
                "  <code>kartlar</code> → kart borçları\n"
                "  <code>yatirim</code> → portföy özeti\n"
                "  <code>hedef</code> → hedefler\n\n"
                "<b>💳 Kart Ödeme:</b>\n"
                "  <code>yapı kredi asgari ödeme</code>\n"
                "  <code>garanti gold 5000 ödendi</code>\n"
                "  <code>akbank kart tam borç ödeme</code>\n\n"
                "<b>✏️ Yönetim:</b>\n"
                "  <code>sil</code> → son işlemi sil\n"
                "  <code>iptal</code> → bekleyen işlemi iptal et")
            return "ok"

        link = db.execute("SELECT user_id, profile_id FROM telegram_links WHERE tg_user_id=%s", (tg_uid,)).fetchone()
        if not link:
            tg_send(chat_id, "⚠️ Hesabınız bağlı değil.\n/start yazarak başlayın.")
            return "ok"

        uid, pid = link["user_id"], link["profile_id"]

        # ── KOMUTLAR ───────────────────────────────────────────────────────────

        if text_l in ("/bakiye", "/balance", "bakiye", "balance", "💰 bakiye",
                      "ne kadar param var", "param ne kadar", "kalan ne", "elimde ne kadar var",
                      "finansal durum", "durum"):
            tg_send(chat_id, _tg_summary_msg(uid, pid, db, "bakiye"))
            return "ok"

        if text_l in ("/bugun", "/bugün", "bugun", "bugün", "📅 bugün",
                      "bugün ne harcadım", "bugun ne harcadim",
                      "bugünün özeti", "bugunun ozeti", "bugun ozet"):
            tg_send(chat_id, _tg_summary_msg(uid, pid, db, "bugun"))
            return "ok"

        # Belirli ay sorgusu: "haziran", "geçen ay", "mayıs 2024 özeti" vb.
        _parsed_month = _tg_parse_month(text_l)
        _is_month_query = (
            text_l in ("/ay", "/aylik", "/aylık", "ay", "aylik", "aylık", "📆 bu ay",
                       "bu ayın özeti", "bu ayin ozeti", "aylık özet", "bu ay ne harcadım",
                       "bu ay ozet", "aylik ozet", "geçen ay", "gecen ay", "önceki ay")
            or (_parsed_month and any(kw in text_l for kw in
                ["özet","ozet","harca","kalanı","kalan","ne kadar","para","para kaldı",
                 "ayı","ayi","ayinda","ayında","bütçe","butce","gelir","gider","ay",
                 "kalem","kalemler","kalemlerim","giderler","giderlerim","gelirler","gelirlerim","detay","rapor"]))
        )
        if _is_month_query:
            if _parsed_month:
                yr, mo = _parsed_month
            else:
                yr, mo = date.today().year, date.today().month
            tg_send(chat_id, _tg_summary_msg(uid, pid, db, "ay", year=yr, month=mo))
            return "ok"

        if text_l.startswith(("/son", "son", "🗒️ son 5")) and (len(text_l) <= 3 or text_l[3:].strip().isdigit()):
            parts = text_l.split()
            n = int(parts[1]) if len(parts)>1 and parts[1].isdigit() else 5
            n = min(n, 20)
            tg_send(chat_id, _tg_last_tx(uid, pid, db, n))
            return "ok"

        if text_l in ("/butce", "/bütçe", "butce", "bütçe", "📊 bütçe",
                      "bütçem nasıl", "butce durum", "butce ozet"):
            tg_send(chat_id, _tg_budget_msg(pid, db))
            return "ok"

        if text_l in ("/kartlar", "/kart", "kartlar", "kart", "💳 kartlar",
                      "kart borçları", "kart borclarim", "kartlarim",
                      "yemek kart", "yemek kartı", "yemek kartim"):
            tg_send(chat_id, _tg_cards_msg(pid, db))
            return "ok"

        # ── Kart ödeme algılama ──
        _cp = _tg_parse_card_payment(text_l)
        if _cp:
            found_card = _tg_find_card(db, pid, text_l)
            if found_card:
                cur_debt = float(found_card["used_"] or 0)
                if cur_debt <= 0:
                    tg_send(chat_id, f"💳 {found_card['bank_name']} kartında ödenecek borç yok.")
                    return "ok"
                min_pay = round(cur_debt * float(found_card["min_pct"] or 25) / 100, 2)
                if _cp["is_asgari"]:
                    pay_amount = min_pay
                elif _cp["is_tam"]:
                    pay_amount = cur_debt
                elif _cp["amount"]:
                    pay_amount = min(_cp["amount"], cur_debt)
                else:
                    pay_amount = min_pay  # varsayılan: asgari
                clbl = found_card['bank_name']+(f" {found_card['card_name']}" if found_card.get('card_name') else "")
                pay_type = "asgari ödeme" if _cp["is_asgari"] or (not _cp["is_tam"] and not _cp["amount"]) else "tam borç" if _cp["is_tam"] else "ödeme"
                tg_send(chat_id,
                    f"💳 <b>{clbl}</b> için {pay_type} onayı?\n\n"
                    f"💸 Ödenecek: <b>{_tg_fmt_amount(pay_amount)}</b>\n"
                    f"📉 Mevcut borç: {_tg_fmt_amount(cur_debt)}\n"
                    f"📉 Ödeme sonrası: {_tg_fmt_amount(max(0,cur_debt-pay_amount))}",
                    reply_markup={"inline_keyboard": [[
                        {"text": "✅ Evet, kaydet", "callback_data": f"cardpay:{found_card['id']}:{pay_amount}"},
                        {"text": "❌ İptal", "callback_data": "cardpay:cancel"}
                    ]]}
                )
                return "ok"

        if text_l in ("/hesaplar", "hesaplar", "hesabım", "hesaplarim",
                      "🏦 hesaplar", "hesap bakiye", "bakiyeler"):
            accs = db.execute(
                "SELECT name,bank,type,initial_balance FROM accounts WHERE profile_id=%s AND active=1 ORDER BY name", (pid,)
            ).fetchall()
            if not accs:
                tg_send(chat_id, "🏦 Kayıtlı hesap yok.")
            else:
                _ACC_ICO_TG = {"vadesiz":"🏦","tasarruf":"📅","vadeli":"📅","kredi_karti":"💳","kmh":"🔄","konut_kredisi":"🏠","arac_kredisi":"🚗","ihtiyac_kredisi":"💼","diger":"📋"}
                lines = ["🏦 <b>Hesaplarım</b>\n"]
                for a in accs:
                    ico = _ACC_ICO_TG.get(a["type"],"🏦")
                    lines.append(f"{ico} <b>{a['bank']} — {a['name']}</b>\n"
                                 f"   Bakiye: {_tg_fmt_amount(a['initial_balance'] or 0)}")
                tg_send(chat_id, "\n".join(lines))
            return "ok"

        if text_l in ("/yatirim", "/yatırım", "yatirim", "yatırım", "📈 yatırım",
                      "yatırımlarım", "portfoy", "portföy", "yatirimlarim"):
            rows = db.execute(
                "SELECT name,itype,quantity,buy_price FROM investments WHERE profile_id=%s ORDER BY quantity*buy_price DESC LIMIT 10",
                (pid,)
            ).fetchall()
            if not rows:
                tg_send(chat_id, "📈 Kayıtlı yatırım yok.")
            else:
                total = sum(r["quantity"]*r["buy_price"] for r in rows)
                lines = ["📈 <b>Yatırım Portföyü</b> (maliyet bazlı)\n"]
                for r in rows:
                    val = r["quantity"] * r["buy_price"]
                    lines.append(f"  • {r['name']} ({r['itype']}): <b>{_tg_fmt_amount(val)}</b>")
                lines.append(f"\n💼 Toplam: <b>{_tg_fmt_amount(total)}</b>")
                tg_send(chat_id, "\n".join(lines))
            return "ok"

        if text_l in ("/hedef", "/hedefler", "hedef", "hedefler", "hedeflerim", "🎯 hedefler"):
            rows = db.execute(
                "SELECT name,goal_type,monthly_target FROM goals WHERE profile_id=%s ORDER BY id LIMIT 10",
                (pid,)
            ).fetchall()
            if not rows:
                tg_send(chat_id, "🎯 Kayıtlı hedef yok.")
            else:
                lines = ["🎯 <b>Hedefler</b>\n"]
                for r in rows:
                    lines.append(f"  • {r['name']} — Aylık hedef: <b>{_tg_fmt_amount(r['monthly_target'])}</b>")
                tg_send(chat_id, "\n".join(lines))
            return "ok"

        if text_l in ("/sil", "/son_sil", "sil", "son islemi sil", "son işlemi sil", "sil son", "🗑️ son i̇şlemi sil", "🗑️ son işlemi sil"):
            last = db.execute(
                "SELECT id,type,amount,category,date FROM transactions WHERE profile_id=%s ORDER BY id DESC LIMIT 1",
                (pid,)
            ).fetchone()
            if not last:
                tg_send(chat_id, "❌ Silinecek işlem yok.")
            else:
                icon = "📈" if last["type"]=="gelir" else "📉"
                tg_send(chat_id,
                    f"🗑 Son işlemi silmek istiyor musunuz?\n\n"
                    f"{icon} {last['category']} — <b>{_tg_fmt_amount(last['amount'])}</b>\n"
                    f"Tarih: {last['date']}",
                    reply_markup={"inline_keyboard": [[
                        {"text": "✅ Evet, sil", "callback_data": f"deltx:{last['id']}"},
                        {"text": "❌ İptal",     "callback_data": "deltx:cancel"}
                    ]]}
                )
            return "ok"

        if text_l in ("/iptal", "iptal", "vazgec", "vazgeç", "hayir", "hayır"):
            db.execute("DELETE FROM telegram_pending WHERE tg_user_id=%s", (tg_uid,))
            db.commit()
            tg_send(chat_id, "✅ Bekleyen tüm işlemler iptal edildi.")
            return "ok"

        # ── /gelir ve /gider komutları ──
        forced_type = None
        if text_l.startswith("/gelir ") or text_l.startswith("gelir "):
            forced_type = "gelir"
            text = re.sub(r'^/?gelir\s+', '', text, flags=re.IGNORECASE)
            text_l = text.lower()
        elif text_l.startswith("/gider ") or text_l.startswith("gider "):
            forced_type = "gider"
            text = re.sub(r'^/?gider\s+', '', text, flags=re.IGNORECASE)
            text_l = text.lower()

        # ── Doğal dil: tarih + kategori override + işlem ──
        tx_date  = _tg_parse_date(text_l)
        clean_text = _tg_strip_date_words(text_l)
        cat_override = _tg_parse_category_override(text)
        if cat_override:
            clean_text = re.sub(r'#[^\s]+', '', clean_text).strip()

        parsed = _parse_tg_msg(clean_text if clean_text else text)
        if not parsed:
            tg_send(chat_id,
                "❓ Anlamadım.\n\n"
                "Örnekler:\n"
                "  <code>market 250</code>\n"
                "  <code>maaş 15000</code>\n"
                "  <code>dün benzin 500</code>\n"
                "  <code>yemek 350 #eğlence</code>\n\n"
                "/help — tüm komutlar")
            return "ok"

        # Kategori override uygula
        if cat_override:
            for kw, cat in _TG_CAT_KW:
                if kw == cat_override.lower():
                    parsed["category"] = cat; break
            else:
                # Direkt kategori adı dene
                parsed["category"] = cat_override.capitalize()

        # Zorunlu tür uygula
        if forced_type:
            parsed["type"] = forced_type

        parsed["date"] = tx_date.isoformat()

        # Pending olarak kaydet
        cur = db.execute(
            "INSERT INTO telegram_pending (tg_user_id,user_id,profile_id,type,amount,category,description,tx_date,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (tg_uid, uid, pid, parsed["type"], parsed["amount"],
             parsed["category"], parsed["description"],
             parsed.get("date", date.today().isoformat()), datetime.now().isoformat())
        )
        db.commit()
        pending_id = cur.lastrowid

        icon = "📈" if parsed["type"] == "gelir" else "📉"
        date_str = f" ({tx_date.strftime('%d.%m')})" if tx_date != date.today() else ""

        # Gider ise kartları göster
        keyboard_rows = []
        if parsed["type"] == "gider":
            all_cards = db.execute(
                "SELECT id,bank_name,card_name,card_type,used_,limit_ FROM cards WHERE profile_id=%s ORDER BY bank_name",
                (pid,)
            ).fetchall()
            cat_l = parsed["category"].lower()
            is_yemek = "yemek" in cat_l or "restoran" in cat_l
            # Yemek kategorisiyse yemek kartlarını öne al, sonra kredi; diğerlerinde kredi önce
            if is_yemek:
                relevant = ([c for c in all_cards if c["card_type"] == "yemek"] +
                            [c for c in all_cards if c["card_type"] == "kredi"])
            else:
                relevant = ([c for c in all_cards if c["card_type"] == "kredi"] +
                            [c for c in all_cards if c["card_type"] == "yemek"])
            # Hiç kredi/yemek yoksa banka kartını da ekle
            if not relevant:
                relevant = [c for c in all_cards if c["card_type"] == "banka"]
            row = []
            for c in relevant[:8]:
                ico = "🍽️" if c["card_type"] == "yemek" else ("🏧" if c["card_type"] == "banka" else "💳")
                lbl = f"{ico} {c['bank_name']}"
                if c.get("card_name"): lbl += f" {c['card_name']}"
                if c["limit_"]:
                    avail = round(float(c["limit_"]) - float(c["used_"] or 0), 2)
                    lbl += f" ·{_tg_fmt_amount(avail)}"
                if len(lbl) > 30: lbl = lbl[:29] + "…"
                row.append({"text": lbl, "callback_data": f"confirm:{pending_id}:c:{c['id']}"})
                if len(row) == 2:
                    keyboard_rows.append(row); row = []
            if row: keyboard_rows.append(row)

        keyboard_rows.append([
            {"text": "💵 Nakit Kaydet", "callback_data": f"confirm:{pending_id}"},
            {"text": "❌  İptal",       "callback_data": f"cancel:{pending_id}"}
        ])

        tg_send(chat_id,
            f"{icon} <b>{'Gelir' if parsed['type']=='gelir' else 'Gider'}: {_tg_fmt_amount(parsed['amount'])}</b>{date_str}\n"
            f"📂 Kategori: {parsed['category']}\n"
            f"📝 Not: {parsed['description']}\n\n"
            + ("💳 Kartla kaydet ya da nakit seç:" if parsed["type"]=="gider" else "Kayıt edilsin mi?"),
            reply_markup={"inline_keyboard": keyboard_rows}
        )
    except Exception as e:
        log.error("Telegram webhook error: %s", e)
        tg_send(chat_id, "⚠️ Bir hata oluştu, lütfen tekrar deneyin.")
    finally:
        db.close()
    return "ok"

@app.route("/api/telegram/link-code", methods=["POST"])
@premium_required
def telegram_link_code():
    uid = session["user_id"]; pid = get_pid()
    code = secrets.token_hex(3).upper()
    expires = (datetime.now() + timedelta(minutes=15)).isoformat()
    db = get_db()
    db.execute("DELETE FROM telegram_link_codes WHERE user_id=%s", (uid,))
    db.execute("INSERT INTO telegram_link_codes (user_id,profile_id,code,expires_at) VALUES (%s,%s,%s,%s)",
               (uid, pid, code, expires))
    db.commit()
    return jsonify({"ok": True, "code": code, "bot": TELEGRAM_BOT_USERNAME})

@app.route("/api/telegram/status", methods=["GET"])
@premium_required
def telegram_status():
    uid = session["user_id"]
    links = get_db().execute(
        "SELECT id, tg_user_id, created_at FROM telegram_links WHERE user_id=%s ORDER BY created_at", (uid,)
    ).fetchall()
    return jsonify({"count": len(links), "links": [row_to_dict(r) for r in links]})

@app.route("/api/telegram/unlink", methods=["DELETE"])
@premium_required
def telegram_unlink():
    uid = session["user_id"]
    data = request.get_json(force=True, silent=True) or {}
    lid = data.get("id")
    db = get_db()
    if lid:
        db.execute("DELETE FROM telegram_links WHERE id=%s AND user_id=%s", (lid, uid))
    else:
        db.execute("DELETE FROM telegram_links WHERE user_id=%s", (uid,))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/telegram/setup-webhook", methods=["POST"])
@premium_required
def setup_telegram_webhook():
    if not TELEGRAM_TOKEN:
        return jsonify({"ok": False, "error": "TELEGRAM_BOT_TOKEN ayarlanmamış"}), 400
    payload = {"url": f"{APP_URL}/api/telegram/webhook"}
    if TELEGRAM_WEBHOOK_SECRET:
        payload["secret_token"] = TELEGRAM_WEBHOOK_SECRET
    r = requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/setWebhook",
        json=payload, timeout=10
    )
    return jsonify(r.json())

# ── ACCOUNTS (Banka Ürünleri) ─────────────────────────────────────────────────

@app.route("/api/accounts", methods=["GET"])
@login_required
def list_accounts():
    pid = get_pid(); db = get_db()
    # Tek sorguda hesap + işlem toplamları (N+1 sorunu giderildi)
    rows = db.execute(
        """SELECT a.*,
             COALESCE(SUM(CASE WHEN t.type='gelir' THEN t.amount ELSE 0 END),0) AS tagged_gelir,
             COALESCE(SUM(CASE WHEN t.type='gider' THEN t.amount ELSE 0 END),0) AS tagged_gider
           FROM accounts a
           LEFT JOIN transactions t ON t.account_id=a.id AND t.profile_id=a.profile_id
           WHERE a.profile_id=? AND a.active=1
           GROUP BY a.id
           ORDER BY a.bank, a.name""",
        (pid,)
    ).fetchall()
    result = []
    for r in rows:
        acc = row_to_dict(r)
        gelir  = float(acc.pop("tagged_gelir", 0) or 0)
        gider  = float(acc.pop("tagged_gider", 0) or 0)
        acc["tagged_gelir"] = gelir
        acc["tagged_gider"] = gider
        init   = float(acc.get("initial_balance") or 0)
        limit_ = float(acc.get("limit_") or 0)
        atype  = acc.get("type", "vadesiz")
        if atype in ("kredi_karti", "kmh"):
            # Kredi ürünlerinde borç = gider − gelir (ters)
            debt = init + gider - gelir
            acc["computed_balance"] = round(debt, 2)
            acc["available"] = round(limit_ - debt, 2) if limit_ > 0 else None
        else:
            bal = init + gelir - gider
            acc["computed_balance"] = round(bal, 2)
            acc["available"] = None
        result.append(acc)
    return jsonify(result)

@app.route("/api/accounts", methods=["POST"])
@login_required
def add_account():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    name = d.get("name","").strip()
    bank = d.get("bank","").strip()
    atype = d.get("type","vadesiz")
    initial = float(d.get("initial_balance",0))
    limit_ = float(d.get("limit_",0))
    color = d.get("color","#007aff")
    owner = d.get("owner","").strip()
    if not name or not bank:
        return jsonify({"ok":False,"error":"Banka ve ürün adı gerekli"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO accounts (user_id,profile_id,name,bank,type,initial_balance,limit_,color,owner,active,created_at) VALUES (?,?,?,?,?,?,?,?,?,1,?)",
        (uid, pid, name, bank, atype, initial, limit_, color, owner, datetime.now().isoformat())
    )
    db.commit()
    return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/accounts/<int:aid>", methods=["PUT"])
@login_required
def update_account(aid):
    pid = get_pid()
    d = request.get_json(force=True)
    fields = []; params = []
    for col in ("name","bank","type","color","owner"):
        if col in d:
            fields.append(f"{col}=?"); params.append(d[col])
    for col in ("initial_balance","limit_"):
        if col in d:
            fields.append(f"{col}=?"); params.append(float(d[col]))
    if "active" in d:
        fields.append("active=?"); params.append(int(d["active"]))
    if not fields: return jsonify({"ok":False}), 400
    params += [aid, pid]
    get_db().execute(f"UPDATE accounts SET {','.join(fields)} WHERE id=? AND profile_id=?", params)
    get_db().commit()
    return jsonify({"ok":True})

@app.route("/api/accounts/<int:aid>", methods=["DELETE"])
@login_required
def delete_account(aid):
    pid = get_pid()
    get_db().execute("DELETE FROM accounts WHERE id=? AND profile_id=?", (aid, pid))
    get_db().commit()
    return jsonify({"ok":True})

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

@app.route("/api/export/excel")
@premium_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from collections import defaultdict

    pid = get_pid(); db = get_db()
    today_d = date.today()

    # ── Renk paleti (iOS/Apple HIG uyumlu) ───────────────────────────────────
    NAVY    = "0D1B2A"   # koyu lacivert — başlıklar
    BLUE    = "0A84FF"   # mavi — vurgu
    GREEN   = "1A6B3A"   # koyu yeşil — gelir
    RED     = "C0392B"   # kırmızı — gider
    AMBER   = "B7610A"   # turuncu/amber — uyarı
    LGRAY   = "F7F8FA"   # açık gri — alt satırlar
    DGRAY   = "E5E5EA"   # orta gri — ayırıcılar
    MGRAY   = "6D6D72"   # metin gri
    WHITE   = "FFFFFF"
    G_LIGHT = "EBF7F0"   # açık yeşil bg
    R_LIGHT = "FDEDEC"   # açık kırmızı bg

    MONTHS_TR = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
                 "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

    def F(bold=False, size=10, color="1C1C1E"):
        return Font(name="Calibri", bold=bold, size=size, color=color)
    def fill(hex_):
        return PatternFill("solid", fgColor=hex_)
    def bdr(style="thin", color="D1D1D6"):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)
    def bdr_bottom(color="B0B0B8"):
        s = Side(style="medium", color=color)
        return Border(bottom=s)
    def AL(wrap=False): return Alignment(horizontal="left",   vertical="top" if wrap else "center", wrap_text=wrap)
    def AC():           return Alignment(horizontal="center", vertical="center")
    def AR():           return Alignment(horizontal="right",  vertical="center")
    def AWrap():        return Alignment(horizontal="left",   vertical="top", wrap_text=True)

    def sc(ws, r, c, val, bold=False, size=10, color="1C1C1E",
           bg=None, align="L", fmt=None, wrap=False):
        """Set cell — merkezi yardımcı."""
        cell = ws.cell(row=r, column=c, value=val)
        cell.font   = F(bold=bold, size=size, color=color)
        cell.border = bdr()
        if align == "C":   cell.alignment = AC()
        elif align == "R": cell.alignment = AR()
        elif wrap:         cell.alignment = AWrap()
        else:              cell.alignment = AL()
        if bg:  cell.fill = fill(bg)
        if fmt: cell.number_format = fmt
        return cell

    def title_row(ws, text, cols, row=1, bg=NAVY, size=14):
        ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
        c = ws[f"A{row}"]
        c.value = text; c.font = Font(name="Calibri", bold=True, size=size, color=WHITE)
        c.fill = fill(bg); c.alignment = AL()
        ws.row_dimensions[row].height = 38

    def sub_row(ws, text, cols, row=2, bg=NAVY):
        ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
        c = ws[f"A{row}"]
        c.value = text; c.font = Font(name="Calibri", size=9, color="90A4AE")
        c.fill = fill(bg); c.alignment = AL()
        ws.row_dimensions[row].height = 20

    def hdr_row(ws, headers, row, bg=BLUE, sizes=None):
        for ci, h in enumerate(headers, 1):
            c = sc(ws, row, ci, h, bold=True, size=9, color=WHITE, bg=bg, align="C")
        ws.row_dimensions[row].height = 24

    # ── Veri çekme ───────────────────────────────────────────────────────────
    profile_row  = db.execute("SELECT name FROM profiles WHERE id=?", (pid,)).fetchone()
    profile_name = profile_row["name"] if profile_row else ""

    # Ödeme yöntemi bilgisiyle birlikte işlemler
    txs = db.execute(
        """SELECT t.type, t.amount, t.category, t.description, t.date,
                  c.bank_name AS card_bank, c.card_name, c.card_type,
                  a.bank AS acc_bank, a.name AS acc_name, a.type AS acc_type
           FROM transactions t
           LEFT JOIN cards    c ON t.card_id    = c.id
           LEFT JOIN accounts a ON t.account_id = a.id
           WHERE t.profile_id=? ORDER BY t.date, t.id""",
        (pid,)
    ).fetchall()

    def pay_method(r):
        if r["card_bank"]:
            ico = "🍽" if r["card_type"] == "yemek" else "🏧" if r["card_type"] == "banka" else "💳"
            lbl = f"{ico} {r['card_bank']}"
            if r["card_name"]: lbl += f" {r['card_name']}"
            return lbl
        if r["acc_bank"]:
            ico = "🔄" if r["acc_type"] == "kmh" else "🏦"
            return f"{ico} {r['acc_bank']} · {r['acc_name']}"
        return "💵 Nakit"

    # Ay bazlı gruplama
    by_month  = defaultdict(list)
    cat_pivot = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    gelir_total = gider_total = 0.0
    active_months = set()
    cat_totals = defaultdict(lambda: defaultdict(float))
    for r in txs:
        ym = r["date"][:7]
        by_month[ym].append(r)
        active_months.add(ym)
        cat_pivot[r["type"]][r["category"]][ym] += float(r["amount"])
        cat_totals[r["type"]][r["category"]]     += float(r["amount"])
        if r["type"] == "gelir": gelir_total += float(r["amount"])
        else:                    gider_total += float(r["amount"])

    sorted_months = sorted(active_months)
    assets = db.execute(
        "SELECT * FROM assets WHERE profile_id=? AND active=1 ORDER BY name", (pid,)
    ).fetchall()

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws0 = wb.active
    ws0.title = "📊 Özet"

    N_COLS = 8
    title_row(ws0, f"   🦔  KİRPİ — FİNANSAL RAPOR", N_COLS, row=1)
    sub_row(ws0,   f"   {profile_name}  ·  Tüm Dönem  ·  {today_d.strftime('%d %B %Y')}", N_COLS, row=2)
    ws0.row_dimensions[3].height = 10

    # ── KPI BANNER ───────────────────────────────────────────────────────────
    net_ = gelir_total - gider_total
    kpis = [
        ("TOPLAM GELİR",  f"₺{gelir_total:,.2f}", GREEN),
        ("TOPLAM GİDER",  f"₺{gider_total:,.2f}", RED),
        ("NET BAKİYE",    f"₺{net_:,.2f}", GREEN if net_>=0 else RED),
        ("İŞLEM SAYISI",  str(len(txs)), BLUE),
    ]
    for ki, (lbl, val, clr) in enumerate(kpis):
        col1 = ki * 2 + 1; col2 = col1 + 1
        c1 = get_column_letter(col1); c2 = get_column_letter(col2)
        ws0.merge_cells(f"{c1}4:{c2}4")
        ws0.merge_cells(f"{c1}5:{c2}5")
        ws0.merge_cells(f"{c1}6:{c2}6")
        for rr in [4,5,6]:
            for cc in [col1, col2]:
                ws0.cell(row=rr, column=cc).fill = fill(clr)
        lc = ws0[f"{c1}4"]; lc.value = lbl
        lc.font = Font(name="Calibri", bold=True, size=8, color="FFFFFF99")
        lc.alignment = AC()
        vc = ws0[f"{c1}5"]; vc.value = val
        vc.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        vc.fill = fill(clr); vc.alignment = AC()
    ws0.row_dimensions[4].height = 16
    ws0.row_dimensions[5].height = 32
    ws0.row_dimensions[6].height = 16
    ws0.row_dimensions[7].height = 12

    # ── AYLIK ÖZET ───────────────────────────────────────────────────────────
    mo_summary = defaultdict(lambda: {"gelir": 0.0, "gider": 0.0})
    for r in txs:
        mo_summary[r["date"][:7]][r["type"]] += float(r["amount"])

    # Bölüm başlığı
    ws0.merge_cells("A8:H8")
    bh = ws0["A8"]
    bh.value = "   AYLIK GELİR / GİDER ÖZETİ"
    bh.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    bh.fill = fill(BLUE); bh.alignment = AL()
    ws0.row_dimensions[8].height = 22

    for ci, h in enumerate(["YIL","AY","GELİR","GİDER","NET","DEĞİŞİM (NET)"], 1):
        sc(ws0, 9, ci, h, bold=True, size=9, color=WHITE, bg=NAVY, align="C")
    ws0.row_dimensions[9].height = 22

    prev_net = None
    for i, ym in enumerate(sorted(mo_summary.keys(), reverse=True)):
        yr_, mo_ = int(ym[:4]), int(ym[5:7])
        g_ = mo_summary[ym]["gelir"]; z_ = mo_summary[ym]["gider"]
        net_mo = g_ - z_
        change_val = net_mo - prev_net if prev_net is not None else None
        prev_net = net_mo
        dr = 10 + i
        bg_ = LGRAY if i % 2 == 0 else WHITE
        sc(ws0, dr, 1, yr_,                                           bg=bg_, align="C", size=9)
        sc(ws0, dr, 2, MONTHS_TR[mo_] if mo_ <= 12 else mo_,         bg=bg_, align="C", size=9)
        sc(ws0, dr, 3, round(g_,2),  fmt='#,##0.00 ₺', color=GREEN,  bg=bg_, align="R", bold=True, size=9)
        sc(ws0, dr, 4, round(z_,2),  fmt='#,##0.00 ₺', color=RED,    bg=bg_, align="R", bold=True, size=9)
        sc(ws0, dr, 5, round(net_mo,2), fmt='#,##0.00 ₺',
           color=GREEN if net_mo>=0 else RED, bg=bg_, align="R", bold=True, size=9)
        change_str = "" if change_val is None else f"{'▲' if change_val>=0 else '▼'} ₺{abs(change_val):,.0f}"
        sc(ws0, dr, 6, change_str, color=GREEN if (change_val or 0)>=0 else RED, bg=bg_, align="C", size=9)
        ws0.row_dimensions[dr].height = 18

    # ── TOP GİDER KATEGORİLERİ ───────────────────────────────────────────────
    last_mo_row = 10 + len(mo_summary)
    sep_row = last_mo_row + 2
    ws0.row_dimensions[sep_row-1].height = 12
    ws0.merge_cells(f"A{sep_row}:H{sep_row}")
    bh2 = ws0[f"A{sep_row}"]
    bh2.value = "   EN BÜYÜK GİDER KATEGORİLERİ (Tüm Dönem)"
    bh2.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    bh2.fill = fill("8E4545"); bh2.alignment = AL()
    ws0.row_dimensions[sep_row].height = 22

    for ci, h in enumerate(["KATEGORİ","TOPLAM GİDER","%"], 1):
        sc(ws0, sep_row+1, ci, h, bold=True, size=9, color=WHITE, bg=NAVY, align="C")
    ws0.row_dimensions[sep_row+1].height = 20

    top_gider = sorted(cat_totals["gider"].items(), key=lambda x: -x[1])[:12]
    for i, (cat, amt) in enumerate(top_gider):
        pct = round(amt/gider_total*100, 1) if gider_total else 0
        dr = sep_row + 2 + i
        bg_ = R_LIGHT if i % 2 == 0 else WHITE
        sc(ws0, dr, 1, cat,         bg=bg_, size=9)
        sc(ws0, dr, 2, round(amt,2), fmt='#,##0.00 ₺', color=RED, bg=bg_, align="R", bold=True, size=9)
        sc(ws0, dr, 3, f"%{pct}",   bg=bg_, align="C", size=9, color=MGRAY)
        ws0.row_dimensions[dr].height = 17

    for ci, w in enumerate([10, 14, 14, 14, 14, 18, 12, 12], 1):
        ws0.column_dimensions[get_column_letter(ci)].width = w
    ws0.freeze_panes = "A10"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2/3 — KATEGORİ PİVOT (Gelir + Gider)
    # ══════════════════════════════════════════════════════════════════════════
    def make_pivot_sheet(wb_, title, ttype, color):
        ws = wb_.create_sheet(title)
        cats = sorted(cat_pivot[ttype].keys())
        months = sorted_months
        N = len(months) + 3  # Kategori + aylar + Toplam + %

        title_row(ws, f"   🦔  {title.upper()}", N, row=1, bg=NAVY)
        sub_row(ws,   f"   {profile_name}  ·  {today_d.strftime('%d.%m.%Y')}", N, row=2, bg=NAVY)

        # Sütun başlıkları
        sc(ws, 3, 1, "KATEGORİ", bold=True, size=9, color=WHITE, bg=color, align="C")
        for mi, ym in enumerate(months, 2):
            yr_, mo_ = int(ym[:4]), int(ym[5:7])
            cell = ws.cell(row=3, column=mi)
            cell.value = f"{MONTHS_TR[mo_]}\n{yr_}"
            cell.font  = Font(name="Calibri", bold=True, size=8, color=WHITE)
            cell.fill  = fill(color); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr()
            ws.column_dimensions[get_column_letter(mi)].width = 12
        sc(ws, 3, len(months)+2, "TOPLAM",  bold=True, size=9, color=WHITE, bg=NAVY, align="C")
        sc(ws, 3, len(months)+3, "PAY %",   bold=True, size=9, color=WHITE, bg=NAVY, align="C")
        ws.row_dimensions[3].height = 30
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions[get_column_letter(len(months)+2)].width = 14
        ws.column_dimensions[get_column_letter(len(months)+3)].width = 10

        # Kategori satırları
        grand = 0.0
        for ci_, cat in enumerate(cats):
            dr = 4 + ci_
            bg_ = G_LIGHT if ttype=="gelir" and ci_%2==0 else (R_LIGHT if ttype=="gider" and ci_%2==0 else (LGRAY if ci_%2==0 else WHITE))
            sc(ws, dr, 1, cat, bold=True, size=9, bg=bg_)
            row_total = 0.0
            for mi, ym in enumerate(months, 2):
                v = cat_pivot[ttype][cat].get(ym, 0)
                row_total += v
                if v:
                    sc(ws, dr, mi, round(v,2), fmt='#,##0.00', color=color, bg=bg_, align="R", size=9)
                else:
                    ws.cell(row=dr, column=mi).fill = fill(bg_)
                    ws.cell(row=dr, column=mi).border = bdr()
            sc(ws, dr, len(months)+2, round(row_total,2), bold=True, fmt='#,##0.00', color=color, bg=DGRAY, align="R", size=9)
            grand_pct = f"%{row_total/cat_totals[ttype].get(cat,1)*100:.1f}" if gider_total else ""
            total_all = sum(cat_totals[ttype].values())
            grand_pct = f"%{round(row_total/total_all*100,1)}" if total_all else ""
            sc(ws, dr, len(months)+3, grand_pct, size=9, color=MGRAY, bg=DGRAY, align="C")
            grand += row_total
            ws.row_dimensions[dr].height = 18

        # Toplam satırı
        total_row = 4 + len(cats)
        sc(ws, total_row, 1, "TOPLAM", bold=True, size=10, color=WHITE, bg=NAVY, align="C")
        for mi, ym in enumerate(months, 2):
            col_sum = sum(cat_pivot[ttype][c].get(ym,0) for c in cats)
            sc(ws, total_row, mi, round(col_sum,2), bold=True, fmt='#,##0.00', color=WHITE, bg=NAVY, align="R", size=9)
        sc(ws, total_row, len(months)+2, round(grand,2), bold=True, size=11, fmt='#,##0.00', color=WHITE, bg=color, align="R")
        sc(ws, total_row, len(months)+3, "%100", bold=True, size=9, color=WHITE, bg=NAVY, align="C")
        ws.row_dimensions[total_row].height = 28
        ws.freeze_panes = "B4"
        return ws

    make_pivot_sheet(wb, "💚 Gelir Pivot", "gelir", GREEN)
    make_pivot_sheet(wb, "❤️ Gider Pivot", "gider", RED)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — HESAP HAREKETLERİ (yeniden tasarım)
    # ══════════════════════════════════════════════════════════════════════════
    ws_tx = wb.create_sheet("📋 Hareketler")
    TX_COLS = 7
    title_row(ws_tx, "   🦔  HESAP HAREKETLERİ", TX_COLS, row=1)
    sub_row(ws_tx, f"   {profile_name}  ·  {today_d.strftime('%d.%m.%Y')}", TX_COLS, row=2)

    # Sütun genişlikleri: Tarih | Tür | Kategori | Açıklama | Ödeme Yöntemi | Tutar | Bakiye
    col_cfg = [("Tarih",11), ("Tür",10), ("Kategori",20), ("Açıklama",40),
               ("Ödeme Yöntemi",22), ("Tutar (₺)",14), ("Aylık Net",12)]
    for ci_, (_, w) in enumerate(col_cfg, 1):
        ws_tx.column_dimensions[get_column_letter(ci_)].width = w

    cur_row = 3
    for ym in sorted(by_month.keys(), reverse=True):
        yr_, mo_ = int(ym[:4]), int(ym[5:7])
        mo_txs = sorted(by_month[ym], key=lambda x: x["date"])
        mo_g  = sum(float(r["amount"]) for r in mo_txs if r["type"]=="gelir")
        mo_z  = sum(float(r["amount"]) for r in mo_txs if r["type"]=="gider")
        mo_net = mo_g - mo_z

        # ── Ay başlık şeridi ────────────────────────────────────────────────
        ws_tx.merge_cells(f"A{cur_row}:E{cur_row}")
        mc = ws_tx[f"A{cur_row}"]
        mc.value = f"   {MONTHS_TR[mo_]} {yr_}  ·  {len(mo_txs)} işlem"
        mc.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
        mc.fill  = fill(BLUE); mc.alignment = AL()
        # Gelir/Gider toplamı sağda
        sc(ws_tx, cur_row, 6, round(mo_g,2), bold=True, fmt='#,##0.00 ₺', color=WHITE, bg=GREEN, align="R", size=10)
        sc(ws_tx, cur_row, 7, round(mo_z,2), bold=True, fmt='#,##0.00 ₺', color=WHITE, bg=RED,   align="R", size=10)
        ws_tx.row_dimensions[cur_row].height = 26
        cur_row += 1

        # ── Sütun başlıkları ────────────────────────────────────────────────
        for ci_, (h, _) in enumerate(col_cfg, 1):
            sc(ws_tx, cur_row, ci_, h, bold=True, size=9, color=WHITE, bg=NAVY, align="C")
        ws_tx.row_dimensions[cur_row].height = 22
        cur_row += 1

        # ── İşlem satırları ─────────────────────────────────────────────────
        for i, r in enumerate(mo_txs):
            is_g  = r["type"] == "gelir"
            desc  = (r["description"] or "").strip()
            # Açıklamaya göre dinamik satır yüksekliği
            lines = max(1, (len(desc) // 38) + 1) if desc else 1
            row_h = min(60, 18 + (lines - 1) * 14)
            # Satır rengi
            bg_ = (G_LIGHT if is_g else R_LIGHT) if i % 2 == 0 else WHITE

            sc(ws_tx, cur_row, 1, r["date"],               bg=bg_, align="C", size=9)
            # Tür badge
            type_cell = ws_tx.cell(row=cur_row, column=2)
            type_cell.value = "▲ Gelir" if is_g else "▼ Gider"
            type_cell.font  = Font(name="Calibri", bold=True, size=9,
                                   color=GREEN if is_g else RED)
            type_cell.fill  = fill(bg_)
            type_cell.alignment = AC()
            type_cell.border = bdr()

            sc(ws_tx, cur_row, 3, r["category"],            bg=bg_, size=9)
            # Açıklama — wrap_text aktif
            desc_cell = ws_tx.cell(row=cur_row, column=4)
            desc_cell.value     = desc
            desc_cell.font      = Font(name="Calibri", size=9, color="3D3D3D")
            desc_cell.fill      = fill(bg_)
            desc_cell.alignment = AWrap()
            desc_cell.border    = bdr()

            sc(ws_tx, cur_row, 5, pay_method(r),            bg=bg_, size=8, color=MGRAY)
            sc(ws_tx, cur_row, 6, float(r["amount"]),
               bold=True, fmt='#,##0.00 ₺',
               color=GREEN if is_g else RED,
               bg=bg_, align="R", size=10)
            ws_tx.row_dimensions[cur_row].height = row_h
            cur_row += 1

        # ── Ay net özet satırı ──────────────────────────────────────────────
        ws_tx.merge_cells(f"A{cur_row}:E{cur_row}")
        nc = ws_tx[f"A{cur_row}"]
        nc.value = f"   {'▲' if mo_net>=0 else '▼'} {MONTHS_TR[mo_]} Net"
        nc.font  = Font(name="Calibri", bold=True, size=9, color=GREEN if mo_net>=0 else RED)
        nc.fill  = fill(DGRAY); nc.alignment = AL()
        sc(ws_tx, cur_row, 6, round(mo_net,2), bold=True, fmt='#,##0.00 ₺',
           color=GREEN if mo_net>=0 else RED, bg=DGRAY, align="R", size=10)
        ws_tx.row_dimensions[cur_row].height = 20
        cur_row += 2  # boşluk satırı

    ws_tx.freeze_panes = "A3"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — TEDARİKÇİ FATURALARI
    # ══════════════════════════════════════════════════════════════════════════
    ws_sup = wb.create_sheet("🏭 Tedarikçi")
    sup_hdrs = [("Tedarikçi",22),("Fatura No",14),("Tutar",14),("Döviz",10),
                ("Fatura Tar.",13),("Vade",13),("Durum",14)]
    title_row(ws_sup, "   🦔  TEDARİKÇİ FATURALARI", len(sup_hdrs), row=1)
    for ci_, (h,w) in enumerate(sup_hdrs, 1):
        sc(ws_sup, 2, ci_, h, bold=True, size=9, color=WHITE, bg=BLUE, align="C")
        ws_sup.column_dimensions[get_column_letter(ci_)].width = w
    ws_sup.row_dimensions[2].height = 22

    for i, row in enumerate(db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? ORDER BY due_date", (pid,)
    ).fetchall()):
        dr = 3 + i
        try:
            due_d   = date.fromisoformat(row["due_date"])
            is_late = due_d < today_d and row["status"] == "bekliyor"
            days_ov = (today_d - due_d).days if is_late else 0
        except: is_late = False; days_ov = 0
        bg_ = LGRAY if i % 2 == 0 else WHITE
        if is_late: bg_ = R_LIGHT
        elif row["status"] == "odendi": bg_ = G_LIGHT
        status_lbl = ("✅ Ödendi" if row["status"]=="odendi"
                      else f"⚠️ {days_ov} gün gecikmiş" if is_late else "🕐 Bekliyor")
        sc(ws_sup, dr, 1, row["supplier_name"],     bg=bg_, size=9)
        sc(ws_sup, dr, 2, row["invoice_no"] or "",  bg=bg_, align="C", size=9)
        sc(ws_sup, dr, 3, float(row["amount"]),     fmt='#,##0.00', bg=bg_, align="R", size=9)
        sc(ws_sup, dr, 4, row["currency"],          bg=bg_, align="C", size=9)
        sc(ws_sup, dr, 5, row["invoice_date"],      bg=bg_, align="C", size=9)
        sc(ws_sup, dr, 6, row["due_date"],          bg=bg_, align="C", size=9,
           color=RED if is_late else "1C1C1E")
        sc(ws_sup, dr, 7, status_lbl, bold=True,    bg=bg_, align="C", size=9,
           color=GREEN if row["status"]=="odendi" else (RED if is_late else AMBER))
        ws_sup.row_dimensions[dr].height = 18

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — VARLIKLAR & AMORTİSMAN
    # ══════════════════════════════════════════════════════════════════════════
    ws_ast = wb.create_sheet("🚗 Varlıklar")
    ast_hdrs = [("Varlık Adı",26),("Tür",14),("Alış Tarihi",13),("Alış Bedeli",16),
                ("Amortisman %",12),("Yıllık Amort.",16),("Defter Değeri",16)]
    title_row(ws_ast, "   🦔  VARLIKLAR & AMORTİSMAN", len(ast_hdrs), row=1)
    for ci_, (h,w) in enumerate(ast_hdrs, 1):
        sc(ws_ast, 2, ci_, h, bold=True, size=9, color=WHITE, bg=BLUE, align="C")
        ws_ast.column_dimensions[get_column_letter(ci_)].width = w
    ws_ast.row_dimensions[2].height = 22

    total_cost = total_book = 0.0
    for i, row in enumerate(assets):
        dep = calc_depreciation(float(row["purchase_price"]), row["purchase_date"],
                                float(row["depreciation_rate"]))
        dr  = 3 + i
        bg_ = LGRAY if i % 2 == 0 else WHITE
        sc(ws_ast, dr, 1, row["name"],                     bg=bg_, size=9)
        sc(ws_ast, dr, 2, row["asset_type"],               bg=bg_, align="C", size=9)
        sc(ws_ast, dr, 3, row["purchase_date"],            bg=bg_, align="C", size=9)
        sc(ws_ast, dr, 4, float(row["purchase_price"]),    fmt='#,##0.00 ₺', bg=bg_, align="R", size=9)
        sc(ws_ast, dr, 5, f"%{float(row['depreciation_rate']):.0f}", bg=bg_, align="C", size=9)
        sc(ws_ast, dr, 6, dep["annual"],    fmt='#,##0.00 ₺', bold=True, color=RED,  bg=bg_, align="R", size=9)
        sc(ws_ast, dr, 7, dep["book_value"],fmt='#,##0.00 ₺', bold=True, color=BLUE, bg=bg_, align="R", size=9)
        ws_ast.row_dimensions[dr].height = 18
        total_cost += float(row["purchase_price"]); total_book += dep["book_value"]
    if assets:
        dr_t = 3 + len(assets)
        sc(ws_ast, dr_t, 1, "TOPLAM", bold=True, size=10, color=WHITE, bg=NAVY, align="C")
        sc(ws_ast, dr_t, 4, round(total_cost,2), fmt='#,##0.00 ₺', bold=True, color=WHITE, bg=NAVY, align="R")
        sc(ws_ast, dr_t, 7, round(total_book,2), fmt='#,##0.00 ₺', bold=True, color=WHITE, bg=NAVY, align="R")
        ws_ast.row_dimensions[dr_t].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"kirpi_rapor_{today_d.strftime('%Y%m%d')}.xlsx"
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def write_sheet_header(ws, title, subtitle=""):
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = f"🦔 KIRPI — {title}"
        c.font  = Font(name="Calibri", bold=True, size=16, color=WHITE)
        c.fill  = navy_fill()
        c.alignment = center()
        ws.row_dimensions[1].height = 36
        if subtitle:
            ws.merge_cells("A2:H2")
            c2 = ws["A2"]
            c2.value = subtitle
            c2.font  = Font(name="Calibri", size=10, color="A0AEC0")
            c2.fill  = navy_fill()
            c2.alignment = center()
            ws.row_dimensions[2].height = 20
            return 3
        return 2

    def write_col_headers(ws, row, headers, col_widths=None):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill = orange_fill()
            c.alignment = center()
            c.border = thin_border()
        ws.row_dimensions[row].height = 22
        if col_widths:
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

    def alt_fill(row_idx):
        return gray_fill() if row_idx % 2 == 0 else PatternFill("solid", fgColor=WHITE)

    wb = Workbook()

    # ── Sheet 1: Hesap Hareketleri ────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Hesap Hareketleri"
    txs = db.execute(
        "SELECT * FROM transactions WHERE profile_id=? ORDER BY date DESC, id DESC", (pid,)
    ).fetchall()
    profile_row = db.execute("SELECT name FROM profiles WHERE id=?", (pid,)).fetchone()
    profile_name = profile_row["name"] if profile_row else ""

    ws0.merge_cells("A1:F1")
    c = ws0["A1"]
    c.value = f"🦔 KIRPI — HESAP HAREKETLERİ"
    c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.fill = navy_fill()
    c.alignment = center()
    ws0.row_dimensions[1].height = 36

    ws0.merge_cells("A2:F2")
    c2 = ws0["A2"]
    c2.value = f"{profile_name}  •  {today_d.strftime('%d.%m.%Y')}  •  {len(txs)} işlem"
    c2.font = Font(name="Calibri", size=10, color="A0AEC0")
    c2.fill = navy_fill()
    c2.alignment = center()
    ws0.row_dimensions[2].height = 20

    write_col_headers(ws0, 3, ["Tarih","Tür","Kategori","Açıklama","Tutar","Not"],
                      [14, 10, 18, 32, 14, 20])
    ws0.freeze_panes = "A4"
    ws0.auto_filter.ref = "A3:F3"

    gelir_total = 0.0
    gider_total = 0.0
    GREEN_LIGHT = "D6F5E3"
    RED_LIGHT   = "FDE8E8"

    for ri, row in enumerate(txs):
        dr = 4 + ri
        is_gelir = row["type"] == "gelir"
        amt = float(row["amount"])
        if is_gelir: gelir_total += amt
        else:        gider_total += amt

        row_fill = PatternFill("solid", fgColor=(GREEN_LIGHT if is_gelir else RED_LIGHT) if ri%2==0
                               else ("EDFAF3" if is_gelir else "FAF0F0"))
        vals = [row["date"],
                "Gelir" if is_gelir else "Gider",
                row["category"] or "",
                row["description"] or "",
                amt,
                row["note"] or ""]
        for ci, v in enumerate(vals, 1):
            c = ws0.cell(row=dr, column=ci, value=v)
            c.fill   = row_fill
            c.border = thin_border()
            if ci == 1:
                c.font = body_font(size=10, color="555555")
                c.alignment = center()
            elif ci == 2:
                c.font = Font(name="Calibri", bold=True, size=10,
                              color=GREEN if is_gelir else RED)
                c.alignment = center()
            elif ci == 5:
                c.font = Font(name="Calibri", bold=True, size=10,
                              color=GREEN if is_gelir else RED)
                c.number_format = '#,##0.00'
                c.alignment = center()
            else:
                c.font = body_font(size=10)
                c.alignment = left()
        ws0.row_dimensions[dr].height = 18

    # Totals row
    tot_row = 4 + len(txs) + 1
    ws0.merge_cells(f"A{tot_row}:D{tot_row}")
    tc = ws0.cell(row=tot_row, column=1, value="TOPLAM")
    tc.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    tc.fill = navy_fill(); tc.alignment = center()
    for ci, (v, col) in enumerate([(gelir_total, GREEN),(gider_total, RED)], 5):
        c = ws0.cell(row=tot_row, column=ci, value=v)
        c.font = Font(name="Calibri", bold=True, size=11, color=col)
        c.fill = navy_fill(); c.number_format = '#,##0.00'; c.alignment = center()
    ws0.row_dimensions[tot_row].height = 26

    # ── Sheet 2: Aylık Özet ───────────────────────────────────────────────────
    ws_mo = wb.create_sheet("Aylik Ozet")
    ws_mo.merge_cells("A1:E1")
    c = ws_mo["A1"]
    c.value = "🦔 KIRPI — AYLIK ÖZET"
    c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.fill = navy_fill(); c.alignment = center()
    ws_mo.row_dimensions[1].height = 36
    write_col_headers(ws_mo, 2, ["Yıl","Ay","Gelir","Gider","Net"], [8,14,16,16,16])
    ws_mo.freeze_panes = "A3"

    from collections import defaultdict
    monthly = defaultdict(lambda: {"gelir":0.0,"gider":0.0})
    for row in txs:
        try:
            ym = row["date"][:7]
            monthly[ym][row["type"]] += float(row["amount"])
        except: pass
    MONTHS_TR = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
                 "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]
    for ri, ym in enumerate(sorted(monthly.keys(), reverse=True)):
        dr = 3 + ri
        yr, mo = int(ym[:4]), int(ym[5:7])
        g = monthly[ym]["gelir"]; gz = monthly[ym]["gider"]; net = g - gz
        row_fill = gray_fill() if ri%2==0 else PatternFill("solid", fgColor=WHITE)
        for ci, v in enumerate([yr, MONTHS_TR[mo] if mo<=12 else mo, g, gz, net], 1):
            c = ws_mo.cell(row=dr, column=ci, value=v)
            c.fill = row_fill; c.border = thin_border()
            if ci in (3,4,5):
                c.number_format = '#,##0.00'
                c.font = Font(name="Calibri", bold=(ci==5), size=10,
                              color=(GREEN if (ci==5 and net>=0) else (RED if ci==5 else "1C1C1E")))
                c.alignment = center()
            else:
                c.font = body_font(size=10); c.alignment = center()
        ws_mo.row_dimensions[dr].height = 18

    # ── Sheet 3: Tedarikçi Faturalar ──────────────────────────────────────────
    ws1 = wb.create_sheet("Tedarikci Faturalar")
    data_row = write_sheet_header(ws1, "TEDARİKÇİ FATURALARI",
                                  f"Oluşturma Tarihi: {today_d.strftime('%d.%m.%Y')}")
    cols = ["Tedarikçi","Fatura No","Tutar","Para Birimi","Fatura Tarihi","Vade Tarihi","Durum","Notlar"]
    write_col_headers(ws1, data_row, cols, [22,14,14,10,14,14,12,20])
    data_row += 1
    invoices = db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? ORDER BY due_date", (pid,)
    ).fetchall()
    for ri, row in enumerate(invoices):
        vals = [row["supplier_name"], row["invoice_no"], float(row["amount"]),
                row["currency"], row["invoice_date"], row["due_date"],
                "Ödendi" if row["status"]=="odendi" else "Bekliyor", row["notes"] or ""]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=data_row+ri, column=ci, value=v)
            c.font   = body_font()
            c.fill   = alt_fill(ri)
            c.border = thin_border()
            c.alignment = center() if ci in (3,4,5,6,7) else left()
            if ci == 3:
                c.number_format = '#,##0.00'
            if ci == 7:
                if row["status"] == "odendi":
                    c.font = Font(name="Calibri", size=10, color=GREEN, bold=True)
                else:
                    try:
                        due = date.fromisoformat(row["due_date"])
                        if due < today_d:
                            c.font = Font(name="Calibri", size=10, color=RED, bold=True)
                    except:
                        pass

    # ── Sheet 2: Ödeme Yaşlandırma ────────────────────────────────────────────
    ws2 = wb.create_sheet("Odeme Yaslandirma")
    data_row = write_sheet_header(ws2, "ÖDEME YAŞLANDIRma RAPORU",
                                  f"Rapor Tarihi: {today_d.strftime('%d.%m.%Y')}")
    aging_data = {"0-30 Gün (Güncel)":[],"31-60 Gün":[],"61-90 Gün":[],"90+ Gün (Kritik)":[]}
    aging_totals = {k:0.0 for k in aging_data}
    pending = db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? AND status='bekliyor' ORDER BY due_date",
        (pid,)
    ).fetchall()
    for row in pending:
        try:
            due = date.fromisoformat(row["due_date"])
            days_over = (today_d - due).days
        except:
            days_over = 0
        amt = float(row["amount"])
        item = (row["supplier_name"], float(row["amount"]), row["due_date"], days_over)
        if days_over <= 30:
            aging_data["0-30 Gün (Güncel)"].append(item); aging_totals["0-30 Gün (Güncel)"] += amt
        elif days_over <= 60:
            aging_data["31-60 Gün"].append(item); aging_totals["31-60 Gün"] += amt
        elif days_over <= 90:
            aging_data["61-90 Gün"].append(item); aging_totals["61-90 Gün"] += amt
        else:
            aging_data["90+ Gün (Kritik)"].append(item); aging_totals["90+ Gün (Kritik)"] += amt

    BUCKET_COLORS = {"0-30 Gün (Güncel)":"1D7A46","31-60 Gün":"D4A017",
                     "61-90 Gün":"E07B39","90+ Gün (Kritik)":"C0392B"}
    cur_row = data_row
    for bucket, items in aging_data.items():
        ws2.merge_cells(f"A{cur_row}:F{cur_row}")
        hc = ws2.cell(row=cur_row, column=1, value=f"  {bucket}  —  ₺{aging_totals[bucket]:,.2f}")
        hc.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
        hc.fill  = PatternFill("solid", fgColor=BUCKET_COLORS[bucket])
        hc.alignment = left()
        ws2.row_dimensions[cur_row].height = 22
        cur_row += 1
        if items:
            write_col_headers(ws2, cur_row, ["Tedarikçi","Tutar","Vade","Gecikme (Gün)","",""], [22,14,14,16,1,1])
            cur_row += 1
            for ri, (sup, amt, due_d, days_o) in enumerate(items):
                for ci, v in enumerate([sup, amt, due_d, days_o, "", ""], 1):
                    c = ws2.cell(row=cur_row+ri, column=ci, value=v)
                    c.font = body_font(bold=(ci==4 and days_o>30))
                    c.fill = alt_fill(ri)
                    c.border = thin_border()
                    if ci == 2: c.number_format = '#,##0.00'
                    if ci == 4 and days_o > 90:
                        c.font = Font(name="Calibri", bold=True, size=10, color=RED)
            cur_row += len(items) + 1
        else:
            c = ws2.cell(row=cur_row, column=1, value="  (Bu kategoride bekleyen fatura yok)")
            c.font = Font(name="Calibri", size=9, color="888888", italic=True)
            cur_row += 2

    # ── Sheet 3: Vadeden Kazanç ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Vadeden Kazanc")
    data_row = write_sheet_header(ws3, "VADEDEN KAZANÇ PANOSU",
                                  "Geç ödeme ile elde edilen nakit faiz fırsatı analizi")
    cols3 = ["Tedarikçi","Fatura Tutarı","Vade","Ödeme Tarihi","Gecikme (Gün)","Referans Oran %","Kazanım (₺)"]
    write_col_headers(ws3, data_row, cols3, [22,14,14,14,14,14,14])
    data_row += 1
    paid_inv = db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? AND status='odendi' AND paid_date!=''",
        (pid,)
    ).fetchall()
    total_gain = 0.0
    for ri, row in enumerate(paid_inv):
        try:
            due  = date.fromisoformat(row["due_date"])
            paid = date.fromisoformat(row["paid_date"])
            delayed = (paid - due).days
            if delayed <= 0: continue
            rate = float(row["reference_rate"] or 50)
            gain = float(row["amount"]) * (delayed / 365) * (rate / 100)
            total_gain += gain
        except:
            continue
        vals = [row["supplier_name"], float(row["amount"]), row["due_date"],
                row["paid_date"], delayed, rate, round(gain, 2)]
        for ci, v in enumerate(vals, 1):
            c = ws3.cell(row=data_row+ri, column=ci, value=v)
            c.font = body_font()
            c.fill = alt_fill(ri)
            c.border = thin_border()
            c.alignment = center() if ci > 1 else left()
            if ci in (2, 7): c.number_format = '#,##0.00'
            if ci == 7:
                c.font = Font(name="Calibri", bold=True, size=10, color=GREEN)
    # Total row
    total_row = data_row + len(paid_inv) + 1
    ws3.merge_cells(f"A{total_row}:F{total_row}")
    tc = ws3.cell(row=total_row, column=1, value="TOPLAM KAZANIM")
    tc.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
    tc.fill  = navy_fill()
    tc.alignment = center()
    tv = ws3.cell(row=total_row, column=7, value=round(total_gain, 2))
    tv.font   = Font(name="Calibri", bold=True, size=11, color=ORANGE)
    tv.fill   = navy_fill()
    tv.number_format = '#,##0.00'
    tv.alignment = center()

    # ── Sheet 4: Varlıklar & Amortisman ──────────────────────────────────────
    ws4 = wb.create_sheet("Varliklar Amortisman")
    data_row = write_sheet_header(ws4, "VARLIKLAR & AMORTİSMAN",
                                  f"Türk Vergi Mevzuatına Göre Normal Amortisman ({today_d.year})")
    cols4 = ["Varlık Adı","Tür","Alış Tarihi","Alış Bedeli","Oran %","Yıllık Amortisman","Birikmiş","Defter Değeri"]
    write_col_headers(ws4, data_row, cols4, [22,14,14,14,10,16,14,14])
    data_row += 1
    assets = db.execute("SELECT * FROM assets WHERE profile_id=? AND active=1 ORDER BY name", (pid,)).fetchall()
    for ri, row in enumerate(assets):
        dep = calc_depreciation(float(row["purchase_price"]), row["purchase_date"],
                                float(row["depreciation_rate"]))
        vals = [row["name"], row["asset_type"], row["purchase_date"],
                float(row["purchase_price"]), float(row["depreciation_rate"]),
                dep["annual"], dep["accumulated"], dep["book_value"]]
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(row=data_row+ri, column=ci, value=v)
            c.font = body_font()
            c.fill = alt_fill(ri)
            c.border = thin_border()
            c.alignment = center() if ci > 2 else left()
            if ci in (4,6,7,8): c.number_format = '#,##0.00'

    # ── Sheet 5: Özet ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Ozet")
    write_sheet_header(ws5, "ÖZET RAPOR", f"Tüm modüller — {today_d.strftime('%d.%m.%Y')}")
    summary_data = [
        ("Toplam Bekleyen Tedarikçi Faturası", sum(float(r["amount"]) for r in pending)),
        ("Bekleyen Fatura Adedi", len(pending)),
        ("Toplam Vadeden Kazanım", round(total_gain, 2)),
        ("Aktif Varlık Sayısı", len(assets)),
        ("Toplam Varlık Defter Değeri", sum(
            calc_depreciation(float(r["purchase_price"]), r["purchase_date"],
                              float(r["depreciation_rate"]))["book_value"] for r in assets
        )),
    ]
    for ri, (label, value) in enumerate(summary_data, 1):
        r = 3 + ri
        c1 = ws5.cell(row=r, column=1, value=label)
        c2 = ws5.cell(row=r, column=2, value=value)
        c1.font  = body_font(bold=True, size=11)
        c2.font  = Font(name="Calibri", bold=True, size=12, color=NAVY)
        c1.fill  = alt_fill(ri)
        c2.fill  = alt_fill(ri)
        c1.border = thin_border(); c2.border = thin_border()
        c1.alignment = left(); c2.alignment = center()
        if isinstance(value, float): c2.number_format = '#,##0.00'
        ws5.row_dimensions[r].height = 24
    ws5.column_dimensions["A"].width = 36
    ws5.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"kirpi_rapor_{today_d.strftime('%Y%m%d')}.xlsx"
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── PDF TEMPLATE PALETLERİ ────────────────────────────────────────────────────
PDF_TEMPLATES = {
    "profesyonel":  {"name":"Profesyonel",  "accent":"#2563eb","accent2":"#1d4ed8","bg_head":"#2563eb","txt_head":"#fff","border":"#bfdbfe","row_alt":"#f0f7ff"},
    "minimal":      {"name":"Minimal",      "accent":"#111827","accent2":"#374151","bg_head":"#111827","txt_head":"#fff","border":"#e5e7eb","row_alt":"#f9fafb"},
    "yeşil":        {"name":"Doğa / Yeşil", "accent":"#16a34a","accent2":"#15803d","bg_head":"#16a34a","txt_head":"#fff","border":"#bbf7d0","row_alt":"#f0fdf4"},
    "mor":          {"name":"Modern / Mor", "accent":"#7c3aed","accent2":"#6d28d9","bg_head":"#7c3aed","txt_head":"#fff","border":"#ddd6fe","row_alt":"#f5f3ff"},
    "turuncu":      {"name":"Sıcak / Turuncu","accent":"#ea580c","accent2":"#c2410c","bg_head":"#ea580c","txt_head":"#fff","border":"#fed7aa","row_alt":"#fff7ed"},
    "lacivert":     {"name":"Kurumsal",     "accent":"#1e3a5f","accent2":"#152e4d","bg_head":"#1e3a5f","txt_head":"#fff","border":"#c7d9f0","row_alt":"#f0f4f9"},
    "kirmizi":      {"name":"Dinamik / Kırmızı","accent":"#dc2626","accent2":"#b91c1c","bg_head":"#dc2626","txt_head":"#fff","border":"#fecaca","row_alt":"#fff5f5"},
    "siyah_beyaz":  {"name":"Klasik B&W",   "accent":"#000","accent2":"#333","bg_head":"#000","txt_head":"#fff","border":"#d1d5db","row_alt":"#f3f4f6"},
    "pembe":        {"name":"Şık / Pembe",  "accent":"#db2777","accent2":"#be185d","bg_head":"#db2777","txt_head":"#fff","border":"#fbcfe8","row_alt":"#fdf2f8"},
    "gri":          {"name":"Gümüş / Gri",  "accent":"#475569","accent2":"#334155","bg_head":"#475569","txt_head":"#fff","border":"#cbd5e1","row_alt":"#f8fafc"},
}

@app.route("/api/export/pdf")
@premium_required
def export_pdf():
    pid     = get_pid()
    db      = get_db()
    today_d = date.today()

    MONTHS_TR = ["","Ocak","Şubat","Mart","Nisan","Mayıs","Haziran",
                 "Temmuz","Ağustos","Eylül","Ekim","Kasım","Aralık"]

    # Template seçimi
    tpl_key  = request.args.get("template", "profesyonel")
    tpl      = PDF_TEMPLATES.get(tpl_key, PDF_TEMPLATES["profesyonel"])
    accent   = tpl["accent"]
    accent2  = tpl["accent2"]
    bg_head  = tpl["bg_head"]
    txt_head = tpl["txt_head"]
    border   = tpl["border"]
    row_alt  = tpl["row_alt"]

    # Kullanıcı rapor ayarları (logo, firma adı, iletişim)
    uid = session["user_id"]
    urow = db.execute("SELECT report_name,report_contact,report_logo FROM users WHERE id=?", (uid,)).fetchone()
    report_name    = (urow["report_name"]    if urow else "") or ""
    report_contact = (urow["report_contact"] if urow else "") or ""
    report_logo    = (urow["report_logo"]    if urow else "") or ""

    # Tarih aralığı: önce start/end, yoksa year/month, yoksa bu ay
    year  = request.args.get("year",  today_d.year,  type=int)
    month = request.args.get("month", today_d.month, type=int)
    raw_start = request.args.get("start", "")
    raw_end   = request.args.get("end",   "")
    if raw_start and raw_end:
        try:
            date.fromisoformat(raw_start); date.fromisoformat(raw_end)
            start, end = raw_start, raw_end
            period_label = f"{start} — {end}"
        except ValueError:
            start, end = month_range(today_d.year, today_d.month)
            period_label = f"{MONTHS_TR[today_d.month]} {today_d.year}"
    else:
        start, end = month_range(year, month)
        period_label = f"{MONTHS_TR[month]} {year}"

    profile = db.execute("SELECT name FROM profiles WHERE id=?", (pid,)).fetchone()
    prof_name = profile["name"] if profile else ""

    txns = db.execute(
        "SELECT t.type,t.amount,t.category,t.description,t.date,"
        "t.card_id,t.account_id,"
        "c.bank_name as card_bank,c.card_name as cname,c.card_type,"
        "a.name as acc_name,a.bank as acc_bank,a.type as acc_type "
        "FROM transactions t "
        "LEFT JOIN cards c ON c.id=t.card_id AND t.card_id IS NOT NULL "
        "LEFT JOIN accounts a ON a.id=t.account_id AND t.account_id IS NOT NULL "
        "WHERE t.profile_id=? AND t.date BETWEEN ? AND ? ORDER BY t.date,t.id",
        (pid, start, end)
    ).fetchall()

    gelir_total = sum(r["amount"] for r in txns if r["type"] == "gelir")
    gider_total = sum(r["amount"] for r in txns if r["type"] == "gider")
    hacim_total = gelir_total + gider_total
    net = gelir_total - gider_total

    gelir_cats: dict = {}
    gider_cats: dict = {}
    for r in txns:
        if r["type"] == "gelir":
            gelir_cats[r["category"]] = gelir_cats.get(r["category"], 0) + r["amount"]
        else:
            gider_cats[r["category"]] = gider_cats.get(r["category"], 0) + r["amount"]

    def fmt(n):
        return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def pay_label(r):
        ct = r["card_type"] or ""
        at = r["acc_type"] or ""
        cb = (r["card_bank"] or "") + (" " + (r["cname"] or "") if r["cname"] else "")
        ab = (r["acc_bank"] or "") + (" " + (r["acc_name"] or "") if r["acc_name"] else "")
        if r["card_id"]:
            icons = {"kredi":"💳","banka":"🏧","yemek":"🍽️","hediye":"🎁"}
            return f'{icons.get(ct,"💳")} {html_escape(cb.strip()) or "Kart"}'
        if r["account_id"]:
            icons = {"kmh":"🔄","vadesiz":"🏦","tasarruf":"💰","kredi_karti":"💳"}
            return f'{icons.get(at,"🏦")} {html_escape(ab.strip()) or "Hesap"}'
        return "💵 Nakit"

    def bar_rows(cats, total, color):
        if not cats: return "<p style='color:#888;font-size:.85rem'>Kayıt yok</p>"
        out = ""
        for cat, amt in sorted(cats.items(), key=lambda x: -x[1]):
            pct = (amt / total * 100) if total else 0
            out += f"""
            <div style="margin-bottom:10px">
              <div style="display:flex;justify-content:space-between;font-size:.82rem;margin-bottom:3px">
                <span>{html_escape(cat)}</span>
                <span style="font-weight:700">₺{fmt(amt)} <span style="color:#888;font-weight:400">({pct:.1f}%)</span></span>
              </div>
              <div style="height:7px;background:#eee;border-radius:4px">
                <div style="height:7px;background:{color};border-radius:4px;width:{min(pct,100):.1f}%"></div>
              </div>
            </div>"""
        return out

    tx_rows = ""
    gelir_count = gider_count = 0
    for r in txns:
        sign = "+" if r["type"] == "gelir" else "−"
        color = "#1a7a46" if r["type"] == "gelir" else "#c0392b"
        if r["type"] == "gelir": gelir_count += 1
        else: gider_count += 1
        tx_rows += f"""<tr>
          <td>{r['date']}</td>
          <td>{'Gelir' if r['type']=='gelir' else 'Gider'}</td>
          <td>{html_escape(r['category'])}</td>
          <td style="color:#555">{html_escape(r['description'] or '')}</td>
          <td style="font-size:.78rem;color:#555">{pay_label(r)}</td>
          <td style="text-align:right;font-weight:700;color:{color}">{sign}₺{fmt(r['amount'])}</td>
        </tr>"""

    net_color = "#1a7a46" if net >= 0 else "#c0392b"
    sign = "+" if net >= 0 else "−"

    # Logo HTML
    if report_logo:
        logo_html = f'<img src="{report_logo}" style="max-height:52px;max-width:160px;object-fit:contain">'
    else:
        logo_html = f'<div style="font-size:1.6rem;font-weight:900;color:{txt_head};letter-spacing:-.02em">🦔 Kirpi</div>'

    # Firma/kişi adı
    display_name = html_escape(report_name) if report_name else "Kirpi Nakit Akışı"
    contact_html = f'<div style="font-size:.72rem;opacity:.75;margin-top:2px">{html_escape(report_contact)}</div>' if report_contact else ""

    safe_year  = str(year)
    safe_month = f"{month:02d}"

    html = f"""<!DOCTYPE html>
<html lang="tr"><head>
<meta charset="UTF-8">
<title>Rapor — {period_label}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Helvetica Neue',Arial,sans-serif;color:#1c1c1e;background:#fff;font-size:13px}}
  @media print{{body{{}}@page{{margin:15mm 12mm}}.no-print{{display:none}}}}
  .header{{background:{bg_head};color:{txt_head};padding:22px 32px;display:flex;justify-content:space-between;align-items:center;gap:16px;margin-bottom:24px}}
  .header-right{{text-align:right;font-size:.78rem;opacity:.85;line-height:1.8}}
  .body-wrap{{padding:0 32px 32px}}
  .summary{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:24px}}
  .card{{border:1px solid {border};border-radius:10px;padding:14px;text-align:center}}
  .card-lbl{{font-size:.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:#6d6d72;margin-bottom:5px}}
  .card-val{{font-size:1.3rem;font-weight:900}}
  .section-title{{font-size:.72rem;font-weight:800;text-transform:uppercase;letter-spacing:.1em;color:{accent};margin-bottom:12px;padding-bottom:5px;border-bottom:2px solid {border}}}
  .two-col{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:24px}}
  table{{width:100%;border-collapse:collapse;font-size:.8rem}}
  thead tr{{background:{bg_head};color:{txt_head}}}
  th{{text-align:left;padding:8px 10px;font-weight:700;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em}}
  td{{padding:7px 10px;border-bottom:1px solid {border}}}
  tr:nth-child(even) td{{background:{row_alt}}}
  tr:last-child td{{border-bottom:none}}
  .footer{{margin-top:24px;padding-top:10px;border-top:1px solid {border};font-size:.68rem;color:#aaa;text-align:center;display:flex;justify-content:space-between;align-items:center}}
  .print-btn{{position:fixed;bottom:20px;right:20px;background:{accent};color:#fff;border:none;border-radius:10px;padding:11px 18px;font-size:.88rem;font-weight:700;cursor:pointer;box-shadow:0 4px 16px rgba(0,0,0,.25)}}
</style>
</head>
<body>

<div class="header">
  <div style="display:flex;align-items:center;gap:16px">
    {logo_html}
    <div>
      <div style="font-size:.95rem;font-weight:800;margin-bottom:1px">{display_name}</div>
      {contact_html}
    </div>
  </div>
  <div class="header-right">
    <div style="font-weight:700;font-size:.9rem;margin-bottom:2px">{html_escape(prof_name)}</div>
    <div>{period_label} Finansal Raporu</div>
    <div>Tarih: {today_d.strftime('%d.%m.%Y')}</div>
    <div>{len(txns)} işlem</div>
  </div>
</div>

<div class="body-wrap">
<div class="summary">
  <div class="card" style="border-color:#34c75930;background:#f0fdf4">
    <div class="card-lbl">Toplam Gelir</div>
    <div class="card-val" style="color:#1a7a46">₺{fmt(gelir_total)}</div>
    <div style="font-size:.68rem;color:#888;margin-top:3px">{gelir_count} işlem</div>
  </div>
  <div class="card" style="border-color:#ff3b3030;background:#fff5f5">
    <div class="card-lbl">Toplam Gider</div>
    <div class="card-val" style="color:#c0392b">₺{fmt(gider_total)}</div>
    <div style="font-size:.68rem;color:#888;margin-top:3px">{gider_count} işlem</div>
  </div>
  <div class="card" style="border-color:{border};background:{row_alt}">
    <div class="card-lbl">Net</div>
    <div class="card-val" style="color:{net_color}">{sign}₺{fmt(abs(net))}</div>
    <div style="font-size:.68rem;color:#888;margin-top:3px">Toplam Hacim: ₺{fmt(hacim_total)}</div>
  </div>
</div>

<div class="two-col">
  <div>
    <div class="section-title">Gelir Kalemleri</div>
    {bar_rows(gelir_cats, gelir_total, "#34c759")}
  </div>
  <div>
    <div class="section-title">Gider Kalemleri</div>
    {bar_rows(gider_cats, gider_total, "#ff3b30")}
  </div>
</div>

<div class="section-title">İşlem Detayları ({len(txns)} İşlem)</div>
<table>
  <thead><tr><th>Tarih</th><th>Tür</th><th>Kategori</th><th>Açıklama</th><th>Ödeme</th><th style="text-align:right">Tutar</th></tr></thead>
  <tbody>{tx_rows if tx_rows else '<tr><td colspan="6" style="text-align:center;color:#888;padding:20px">Bu dönemde işlem yok</td></tr>'}
  <tr style="background:#f8f9fa;font-weight:700;border-top:2px solid #dee2e6">
    <td colspan="5" style="font-size:.78rem;color:#495057">Alt Toplam — Gelir</td>
    <td style="text-align:right;color:#1a7a46">+₺{fmt(gelir_total)}</td>
  </tr>
  <tr style="background:#fff5f5;font-weight:700">
    <td colspan="5" style="font-size:.78rem;color:#495057">Alt Toplam — Gider</td>
    <td style="text-align:right;color:#c0392b">−₺{fmt(gider_total)}</td>
  </tr>
  <tr style="background:#f0f4ff;font-weight:800">
    <td colspan="5" style="font-size:.78rem;color:#495057">Toplam İşlem Hacmi</td>
    <td style="text-align:right;color:#333">₺{fmt(hacim_total)}</td>
  </tr>
  </tbody>
</table>

<div class="footer">
  <span>{display_name}</span>
  <span>Kirpi · {datetime.now().strftime('%d.%m.%Y %H:%M')}</span>
</div>
</div>

<button class="print-btn no-print" onclick="window.print()">🖨️ Yazdır / PDF Kaydet</button>
<script>
window.addEventListener('load', function() {{
  document.title = 'Rapor_{safe_year}_{safe_month}';
}});
</script>
</body></html>"""

    return html

@app.route("/api/goals/analysis")
@login_required
def goals_analysis():
    pid = get_pid(); db = get_db()
    today_d = date.today()
    from calendar import monthrange as mr

    # Last 3 months of data
    months_data = []
    for i in range(3):
        if today_d.month - i <= 0:
            m = today_d.month - i + 12; y = today_d.year - 1
        else:
            m = today_d.month - i; y = today_d.year
        _, last = mr(y, m)
        s = f"{y:04d}-{m:02d}-01"; e = f"{y:04d}-{m:02d}-{last:02d}"
        rows = db.execute(
            "SELECT type,SUM(amount) as t FROM transactions WHERE profile_id=? AND date BETWEEN ? AND ? GROUP BY type",
            (pid, s, e)
        ).fetchall()
        gelir = next((float(r["t"]) for r in rows if r["type"]=="gelir"), 0)
        gider = next((float(r["t"]) for r in rows if r["type"]=="gider"), 0)
        months_data.append({"gelir": gelir, "gider": gider, "net": gelir - gider})

    avg_gelir = sum(x["gelir"] for x in months_data) / max(len(months_data), 1)
    avg_gider = sum(x["gider"] for x in months_data) / max(len(months_data), 1)
    avg_net   = avg_gelir - avg_gider

    # Category breakdown for current month
    _, last = mr(today_d.year, today_d.month)
    m_start = today_d.replace(day=1).isoformat()
    m_end   = today_d.replace(day=last).isoformat()
    cat_rows = db.execute(
        "SELECT category,SUM(amount) as t FROM transactions WHERE profile_id=? AND type='gider' AND date BETWEEN ? AND ? GROUP BY category ORDER BY t DESC",
        (pid, m_start, m_end)
    ).fetchall()
    cat_spend = {r["category"]: float(r["t"]) for r in cat_rows}

    # Goals
    goals = db.execute("SELECT * FROM goals WHERE profile_id=?", (pid,)).fetchall()
    total_monthly_goals = sum(float(g["monthly_target"]) for g in goals)
    leftover = avg_net - total_monthly_goals

    # Build commentary
    comments = []

    if avg_gelir == 0:
        comments.append({"icon":"💡","text":"Henüz yeterli veri yok. Birkaç ay gelir ve gider girdikten sonra sana özel tasarruf önerileri sunabilirim.","type":"info"})
        return jsonify({"ok":True,"avg_gelir":0,"avg_gider":0,"avg_net":0,"total_monthly_goals":0,"leftover":0,"comments":comments})

    # Savings capacity comment
    save_rate = round(avg_net / avg_gelir * 100) if avg_gelir > 0 else 0
    if avg_net > 0:
        if save_rate >= 30:
            comments.append({"icon":"🏆","text":f"Harika! Gelirinin %{save_rate}'ini tasarruf ediyorsun. Ortalama aylık ₺{avg_net:,.0f} kenara ayırabiliyorsun. Yatırım hedeflerin için güçlü bir taban var.","type":"success"})
        elif save_rate >= 15:
            comments.append({"icon":"😊","text":f"İyi gidiyorsun! Aylık ortalama ₺{avg_net:,.0f} tasarruf potansiyelin var. Küçük harcamaları kısarak yatırım hedeflerine daha hızlı ulaşabilirsin.","type":"good"})
        else:
            comments.append({"icon":"⚠️","text":f"Bu ay aylık ₺{avg_net:,.0f} tasarruf yapabiliyorsun — gelirinin %{save_rate}'i. Giderlerini %10 azaltsan aylık ₺{avg_gelir*0.1:,.0f} ekstra tasarruf edersin.","type":"warn"})
    else:
        comments.append({"icon":"🔴","text":f"Son 3 ayda ortalama ₺{abs(avg_net):,.0f} açık veriyorsun. Yatırım hedefi koymadan önce giderleri dengelemeyi öneririm.","type":"danger"})

    # Goal allocation comment
    if goals and avg_net > 0:
        if total_monthly_goals <= avg_net:
            if leftover > 500:
                comments.append({"icon":"💰","text":f"Tüm hedeflerini karşılayıp aylık ₺{leftover:,.0f} fazlan kalıyor. Bu fazlayı acil fona ya da yeni bir yatırım hedefine yönlendirebilirsin.","type":"success"})
            else:
                comments.append({"icon":"✅","text":f"Hedeflerin mevcut tasarruf kapasitene sığıyor. ₺{leftover:,.0f} küçük bir buffer var — bir şaşırtıcı gider olursa diye fazla zorlamaya gerek yok.","type":"good"})
        else:
            diff = total_monthly_goals - avg_net
            comments.append({"icon":"⚠️","text":f"Aylık hedeflerinin toplamı (₺{total_monthly_goals:,.0f}) tasarruf kapasiteni ₺{diff:,.0f} aşıyor. Bir veya daha fazla hedefi küçültmeyi düşün.","type":"warn"})

    # Category-specific insights
    TRAVEL_CATS = ["Ulaşım", "Yemek / Restoran", "Eğlence"]
    travel_total = sum(cat_spend.get(c, 0) for c in TRAVEL_CATS)
    if travel_total > avg_gelir * 0.25:
        comments.append({"icon":"✈️","text":f"Bu ay eğlence & ulaşım harcamaların gelirinin %{round(travel_total/avg_gelir*100)}'ini oluşturuyor (₺{travel_total:,.0f}). Bu kategoriyi %20 kıssaydın, döviz ya da altın hedefinize ₺{travel_total*0.2:,.0f} ekleyebilirdin.","type":"warn"})

    food_spend = cat_spend.get("Market / Gıda", 0) + cat_spend.get("Yemek / Restoran", 0)
    if food_spend > avg_gelir * 0.30:
        comments.append({"icon":"🛒","text":f"Market ve yemek harcaman aylık gelirinin %{round(food_spend/avg_gelir*100)}'ini oluşturuyor. Haftalık menü planlaması ile bu kalemi %15 azaltmak mümkün — aylık ₺{food_spend*0.15:,.0f} tasarruf demek.","type":"info"})

    # Goal-type specific tips
    for g in goals:
        gtype = g["goal_type"]
        mt = float(g["monthly_target"])
        if gtype in ("eur","usd","gbp"):
            sym = {"eur":"€","usd":"$","gbp":"£"}[gtype]
            comments.append({"icon":"💶" if gtype=="eur" else ("💵" if gtype=="usd" else "💷"),
                "text":f"'{g['name']}' hedefin için aylık ₺{mt:,.0f} ayırıyorsun. Düzenli alımlar kur dalgalanmalarını dengelemek için ideal — her ay sabit bir günde al.","type":"tip"})
        elif gtype == "gold":
            comments.append({"icon":"🥇","text":f"'{g['name']}' için aylık ₺{mt:,.0f} hedefin var. Altın uzun vadeli bir güvence — gram altın fiyatlarını takip ederek küsürat gramları biriktirebilirsin.","type":"tip"})
        elif gtype == "bitcoin":
            comments.append({"icon":"₿","text":f"'{g['name']}' için aylık ₺{mt:,.0f} ayrılıyor. Bitcoin'de aylık düzenli alım (DCA) volatiliteye karşı en sağlıklı strateji. Tüm birikimini tek seferde yatırmaktan kaçın.","type":"tip"})
        elif gtype == "arsa":
            comments.append({"icon":"🏠","text":f"'{g['name']}' için aylık ₺{mt:,.0f} biriktiriyorsun. Arsa alımında hedef tutarını ve süreyi belirleyip bu tasarrufları vadeli hesapta tutmak getiri sağlar.","type":"tip"})

    # If no goals yet
    if not goals:
        comments.append({"icon":"🎯","text":"Henüz bir yatırım hedefin yok. Euro, dolar, altın, bitcoin veya arsa gibi hedefler ekleyerek aylık ne kadar ayırman gerektiğini görüntüle.","type":"info"})

    return jsonify({
        "ok": True,
        "avg_gelir": round(avg_gelir, 2),
        "avg_gider": round(avg_gider, 2),
        "avg_net": round(avg_net, 2),
        "total_monthly_goals": round(total_monthly_goals, 2),
        "leftover": round(leftover, 2),
        "comments": comments
    })


@app.route("/api/categories")
@login_required
def categories():
    return jsonify({"gelir":GELIR_CATS,"gider":GIDER_CATS,"all":ALL_CATS})

@app.route("/api/init")
@login_required
def api_init():
    """Tek çağrıyla tüm başlangıç verilerini döndür — yükleme süresini azaltır."""
    uid = session.get("user_id"); pid = get_pid()
    db  = get_db()
    row = db.execute("SELECT display_name,username,email,avatar,phone FROM users WHERE id=?", (uid,)).fetchone()
    sub = get_sub_status(uid)
    user_info = {
        "username":     session.get("username"),
        "display":      session.get("display") or (row["display_name"] if row else ""),
        "profile_id":   session.get("profile_id"),
        "profile_name": session.get("profile_name"),
        "profile_type": session.get("profile_type"),
        "email":        row["email"] if row else "",
        "phone":        (row["phone"] if row else "") or "",
        "avatar":       row["avatar"] if row else "",
        "subscription": sub,
    }
    profiles = [row_to_dict(r) for r in db.execute(
        "SELECT * FROM profiles WHERE user_id=? ORDER BY id", (uid,)).fetchall()]
    accounts = [row_to_dict(r) for r in db.execute(
        "SELECT id,name,bank,type,color,initial_balance FROM accounts WHERE profile_id=? AND active=1 ORDER BY name", (pid,)).fetchall()]
    cards_raw = db.execute(
        "SELECT id,bank_name,card_name,card_type,used_,limit_,due_day FROM cards WHERE profile_id=? ORDER BY bank_name", (pid,)).fetchall()
    cards = [row_to_dict(r) for r in cards_raw]
    # Kart hatırlatıcıları
    today_d = date.today(); year, month = today_d.year, today_d.month
    from calendar import monthrange as _mr
    _, last_day = _mr(year, month)
    card_reminders = []
    for c in cards_raw:
        if not (c["used_"] or 0): continue
        due_day  = c["due_day"]
        due_date = f"{year:04d}-{month:02d}-{min(due_day, last_day):02d}"
        days_until = (date.fromisoformat(due_date) - today_d).days
        if 0 <= days_until <= 7:
            card_reminders.append({
                "name": f"{c['bank_name']} {c['card_name']}",
                "amount": float(c["used_"] or 0),
                "due_date": due_date,
                "days_until": days_until,
            })
    tx_count = db.execute("SELECT COUNT(*) as n FROM transactions WHERE profile_id=?", (pid,)).fetchone()["n"]
    return jsonify({
        "user":           user_info,
        "profiles":       profiles,
        "categories":     {"gelir": GELIR_CATS, "gider": GIDER_CATS, "all": ALL_CATS},
        "accounts":       accounts,
        "cards":          cards,
        "card_reminders": card_reminders,
        "tx_count":       tx_count,
    })

@app.route("/api/reminders")
@login_required
def reminders():
    pid     = get_pid(); db = get_db()
    today_d = date.today()
    year, month = today_d.year, today_d.month
    _, last_day = __import__('calendar').monthrange(year, month)
    days_left_in_month = last_day - today_d.day

    # Yaklaşan tekrar eden işlemler (bu ay henüz girilmemiş, önümüzdeki 7 gün içinde)
    recurrings = db.execute(
        "SELECT * FROM recurring WHERE profile_id=? AND active=1 ORDER BY day_of_month",
        (pid,)
    ).fetchall()

    upcoming = []
    for r in recurrings:
        day = r["day_of_month"]
        due_date = f"{year:04d}-{month:02d}-{min(day, last_day):02d}"
        days_until = (date.fromisoformat(due_date) - today_d).days
        if -2 <= days_until <= 7:
            already = db.execute(
                "SELECT id FROM transactions WHERE profile_id=? AND description=? AND date=?",
                (pid, r["description"], due_date)
            ).fetchone()
            if not already:
                upcoming.append({
                    "id": r["id"],
                    "type": r["type"],
                    "amount": float(r["amount"]),
                    "category": r["category"],
                    "description": r["description"] or r["category"],
                    "due_date": due_date,
                    "days_until": days_until,
                })

    # Yaklaşan kart ödeme günleri
    cards = db.execute(
        "SELECT card_name, bank_name, due_day, used_ FROM cards WHERE profile_id=? AND used_>0",
        (pid,)
    ).fetchall()
    card_reminders = []
    for c in cards:
        due_day = c["due_day"]
        due_date = f"{year:04d}-{month:02d}-{min(due_day, last_day):02d}"
        days_until = (date.fromisoformat(due_date) - today_d).days
        if 0 <= days_until <= 7:
            card_reminders.append({
                "name": f"{c['bank_name']} {c['card_name']}",
                "amount": float(c["used_"]),
                "due_date": due_date,
                "days_until": days_until,
            })

    # Onboarding kontrolü
    tx_count = db.execute("SELECT COUNT(*) as n FROM transactions WHERE profile_id=?", (pid,)).fetchone()["n"]
    show_onboarding = tx_count == 0

    return jsonify({
        "upcoming": upcoming,
        "card_reminders": card_reminders,
        "show_onboarding": show_onboarding,
        "tx_count": tx_count,
    })

@app.route("/api/insights")
@premium_required
def insights():
    pid     = get_pid(); db = get_db()
    today_d = date.today()

    # Bu ay ve geçen ay aralıkları
    cur_start, cur_end   = month_range(today_d.year, today_d.month)
    prev_month = today_d.month - 1 if today_d.month > 1 else 12
    prev_year  = today_d.year if today_d.month > 1 else today_d.year - 1
    prv_start, prv_end   = month_range(prev_year, prev_month)

    def period_totals(start, end):
        rows = db.execute(
            "SELECT type, category, SUM(amount) as t FROM transactions "
            "WHERE profile_id=? AND date BETWEEN ? AND ? GROUP BY type, category",
            (pid, start, end)
        ).fetchall()
        gelir, gider, cats = 0.0, 0.0, {}
        for r in rows:
            if r["type"] == "gelir": gelir += r["t"]
            else:
                gider += r["t"]
                cats[r["category"]] = r["t"]
        return gelir, gider, cats

    cur_g, cur_z, cur_cats = period_totals(cur_start, cur_end)
    prv_g, prv_z, prv_cats = period_totals(prv_start, prv_end)

    # Tasarruf oranı
    savings_rate = round((cur_g - cur_z) / cur_g * 100) if cur_g > 0 else 0

    # Skor (0-100)
    score = max(0, min(100, savings_rate + 50))

    # Motivasyon mesajı
    if savings_rate >= 30:
        msg = "Harika gidiyor! Bu ay gelirinin büyük kısmını biriktiriyorsun 🌟"
        emoji = "🟢"
    elif savings_rate >= 10:
        msg = "İyi gidişat! Harcamalarını biraz daha kontrol altına alabilirsin 💪"
        emoji = "🟡"
    elif savings_rate >= 0:
        msg = "Dikkat! Gelirinin neredeyse tamamını harcıyorsun 🔔"
        emoji = "🟠"
    else:
        msg = "Uyarı! Bu ay gelirinden fazla harcama yapıyorsun ⚠️"
        emoji = "🔴"

    # Geçen aya göre kıyaslama
    z_change = round((cur_z - prv_z) / prv_z * 100) if prv_z > 0 else 0
    g_change = round((cur_g - prv_g) / prv_g * 100) if prv_g > 0 else 0

    # Anomali tespiti — önceki aya göre >%25 artan kategoriler
    anomalies = []
    for cat, amt in cur_cats.items():
        prev = prv_cats.get(cat, 0)
        if prev > 0:
            pct = round((amt - prev) / prev * 100)
            if pct >= 25:
                anomalies.append({"cat": cat, "pct": pct, "cur": round(amt), "prv": round(prev)})
    anomalies.sort(key=lambda x: -x["pct"])

    # ── NAKİT LİKİDİTE & NET VARLIK ──────────────────────────────────────────
    days_passed = today_d.day

    # Kredi kartı toplam borcu (KMH kullanılan)
    card_debt = float(db.execute(
        "SELECT COALESCE(SUM(used_),0) as d FROM cards WHERE profile_id=?", (pid,)
    ).fetchone()["d"] or 0)

    # Banka hesapları toplam bakiyesi:
    # Dahil: nakit işlemler + hesaba bağlı işlemler (account_id set) + kart ödemeleri (her ikisi de set)
    # Hariç: saf kart harcamaları (card_id set, account_id NULL) — bunlar zaten card_debt'e dahil
    today_iso = today_d.isoformat()
    acc_bal = db.execute(
        """SELECT
             COALESCE(SUM(a.initial_balance),0) +
             COALESCE((
               SELECT SUM(CASE WHEN t.type='gelir' THEN t.amount ELSE -t.amount END)
               FROM transactions t
               WHERE t.profile_id=%s AND t.date <= %s
                 AND NOT (t.card_id IS NOT NULL AND t.account_id IS NULL)
             ),0) AS b
           FROM accounts a
           WHERE a.profile_id=%s AND a.active=1
             AND a.type NOT IN ('kredi_karti','kmh')""",
        (pid, today_iso, pid)
    ).fetchone()
    total_balance = float(acc_bal["b"] or 0) if acc_bal else 0

    # KMH borcu (accounts tablosundaki KMH kullanımı)
    kmh_debt = float(db.execute(
        """SELECT COALESCE(SUM(t.amount),0)
           FROM transactions t JOIN accounts a ON t.account_id=a.id
           WHERE a.profile_id=%s AND a.type='kmh' AND t.type='gider'""",
        (pid,)
    ).fetchone()[0] or 0)
    kmh_paid = float(db.execute(
        """SELECT COALESCE(SUM(t.amount),0)
           FROM transactions t JOIN accounts a ON t.account_id=a.id
           WHERE a.profile_id=%s AND a.type='kmh' AND t.type='gelir'""",
        (pid,)
    ).fetchone()[0] or 0)
    kmh_balance = max(0, kmh_debt - kmh_paid)

    # Net Likidite = hesap bakiyesi - kredi kartı borcu - KMH borcu
    net_liquidity = round(total_balance - card_debt - kmh_balance)

    # Yatırım değeri (maliyet bazlı — piyasa değeri için ayrı API)
    invest_val = float(db.execute(
        "SELECT COALESCE(SUM(quantity * buy_price),0) as v FROM investments WHERE profile_id=?", (pid,)
    ).fetchone()["v"] or 0)

    # Kıymet değeri
    asset_val = float(db.execute(
        "SELECT COALESCE(SUM(purchase_price),0) as v FROM assets WHERE profile_id=? AND active=1", (pid,)
    ).fetchone()["v"] or 0)

    # Net Varlık = likit varlıklar + yatırımlar + kıymetler
    net_worth = round(total_balance + invest_val + asset_val - card_debt - kmh_balance)

    # Nakit yeterlilik (günlük harcama hızıyla kaç gün dayanır)
    daily_burn = cur_z / days_passed if days_passed > 0 else 0
    days_left  = round(net_liquidity / daily_burn) if daily_burn > 0 and net_liquidity > 0 else 999

    return jsonify({
        "score": score,
        "savings_rate": savings_rate,
        "msg": msg,
        "emoji": emoji,
        "cur_gelir": round(cur_g), "cur_gider": round(cur_z),
        "prv_gelir": round(prv_g), "prv_gider": round(prv_z),
        "gelir_change": g_change, "gider_change": z_change,
        "anomalies": anomalies[:3],
        "days_left": days_left,
        "daily_burn": round(daily_burn),
        "net_liquidity": net_liquidity,
        "net_worth": net_worth,
        "total_balance": round(total_balance),
        "card_debt": round(card_debt),
        "invest_val": round(invest_val),
        "asset_val": round(asset_val),
    })

@app.route("/api/today")
@login_required
def today_summary():
    pid  = get_pid(); db = get_db()
    today_str = request.args.get("date") or date.today().isoformat()

    # All today's transactions
    rows = db.execute(
        "SELECT * FROM transactions WHERE profile_id=? AND date=? ORDER BY id DESC",
        (pid, today_str)
    ).fetchall()

    def row_to_dict(r):
        return {"id": r["id"], "amount": float(r["amount"]),
                "category": r["category"], "description": r["description"] or "",
                "type": r["type"]}

    gelir_list = [row_to_dict(r) for r in rows if r["type"] == "gelir"]
    gider_list = [row_to_dict(r) for r in rows if r["type"] == "gider"]

    # Supplier invoices due today → expected income entries
    try:
        due_invs = db.execute(
            "SELECT * FROM supplier_invoices WHERE profile_id=? AND due_date=? AND status='bekliyor'",
            (pid, today_str)
        ).fetchall()
        for inv in due_invs:
            gelir_list.append({
                "id": None,
                "amount": float(inv["amount"]),
                "category": "Tahsilat",
                "description": "📄 " + (inv["supplier_name"] or "Fatura"),
                "type": "gelir",
                "is_invoice": True
            })
    except Exception:
        pass

    today_gelir = sum(x["amount"] for x in gelir_list)
    today_gider = sum(x["amount"] for x in gider_list)

    # Credit cards
    cards = db.execute("SELECT * FROM cards WHERE profile_id=? ORDER BY bank_name", (pid,)).fetchall()
    card_list = []
    total_avail_sum = 0   # Toplam: sadece pozitif avail'lar toplanır
    total_limit_sum = 0
    total_used_sum  = 0
    for c in cards:
        lim  = float(c["limit_"] or 0)
        used = float(c["used_"]  or 0)
        if lim > 0:
            avail = round(max(0.0, lim - used), 2)   # negatif olmaz
            total_avail_sum += avail
            total_limit_sum += lim
            total_used_sum  += used
        else:
            avail = None   # limitsiz kart — N/A
        card_list.append({
            "id":        c["id"],
            "bank":      c["bank_name"],
            "name":      c["card_name"],
            "card_type": c.get("card_type") or "kredi",
            "limit":     lim,
            "used":      used,
            "avail":     avail,
            "over_limit": lim > 0 and used > lim,   # limit aşıldı mı?
            "due_day":   c["due_day"],
            "min_pct":   c["min_pct"],
            "pct":       round(min(used, lim) / lim * 100) if lim else 0,
        })

    # Banka hesapları + güncel bakiyeler — tek JOIN sorgusu (N+1 giderildi)
    acc_rows = db.execute(
        """SELECT a.*,
             COALESCE(SUM(CASE WHEN t.type='gelir' THEN t.amount ELSE 0 END),0) AS tx_gelir,
             COALESCE(SUM(CASE WHEN t.type='gider' THEN t.amount ELSE 0 END),0) AS tx_gider
           FROM accounts a
           LEFT JOIN transactions t ON t.account_id=a.id AND t.profile_id=a.profile_id
           WHERE a.profile_id=? AND a.active=1
           GROUP BY a.id
           ORDER BY a.bank, a.name""",
        (pid,)
    ).fetchall()
    account_list = []
    for a in acc_rows:
        init  = float(a.get("initial_balance") or 0)
        lim   = float(a.get("limit_") or 0)
        atype = a.get("type","vadesiz")
        g     = float(a["tx_gelir"] or 0)
        z     = float(a["tx_gider"] or 0)
        if atype in ("kredi_karti","kmh"):
            balance = round(init + z - g, 2)
            avail   = round(lim - balance, 2) if lim > 0 else None
        else:
            balance = round(init + g - z, 2)
            avail   = None
        account_list.append({
            "id":      a["id"],
            "bank":    a["bank"],
            "name":    a["name"],
            "type":    atype,
            "color":   a.get("color","#007aff"),
            "balance": balance,
            "limit":   lim,
            "available": avail,
        })

    # Check for recurring income expected today
    today_dt = date.fromisoformat(today_str)
    recurring_gelir_today = False
    recs = db.execute(
        "SELECT * FROM recurring WHERE profile_id=? AND active=1 AND type='gelir'", (pid,)
    ).fetchall()
    for r in recs:
        if r["day_of_month"] == today_dt.day:
            recurring_gelir_today = True
            break

    return jsonify({
        "today_gelir": round(today_gelir, 2),
        "today_gider": round(today_gider, 2),
        "today_net":   round(today_gelir - today_gider, 2),
        "gelir_list":  gelir_list,
        "gider_list":  gider_list,
        "cards":       card_list,
        "accounts":    account_list,
        "total_limit": round(total_limit_sum, 2),
        "total_used":  round(total_used_sum, 2),
        "total_avail": round(total_avail_sum, 2),  # sadece pozitif avail toplamı
        "nakit_bakiye": round(sum(
            a["balance"] for a in account_list
            if a["type"] in ("vadesiz","vadeli","tasarruf")
        ), 2),
        "recurring_gelir_today": recurring_gelir_today,
    })

def generate_notifications(pid, profile_type, db, today):
    """Return list of notification dicts sorted by urgency."""
    notifs = []

    def days_until(target_date):
        return (target_date - today).days

    def urgency(d):
        if d <= 0:   return "urgent"
        if d <= 2:   return "urgent"
        if d <= 5:   return "soon"
        return "normal"

    def add(icon, title, body, d, category=""):
        notifs.append({
            "icon": icon, "title": title, "body": body,
            "days": d, "urgency": urgency(d), "category": category
        })

    # ── CREDIT CARD DUE DATES ──────────────────────────────────────────────
    cards = db.execute("SELECT * FROM cards WHERE profile_id=?", (pid,)).fetchall()
    for c in cards:
        due_day = c["due_day"] or 1
        used = c["used_"] or 0
        if used <= 0:
            continue
        # compute next due date
        if today.day <= due_day:
            due_date = today.replace(day=due_day)
        else:
            # next month
            if today.month == 12:
                due_date = today.replace(year=today.year+1, month=1, day=due_day)
            else:
                due_date = today.replace(month=today.month+1, day=due_day)
        d = days_until(due_date)
        if -3 <= d <= 10:
            min_pay = round(used * (c["min_pct"] or 25) / 100, 2)
            name = c["bank_name"] + (" - " + c["card_name"] if c["card_name"] else "")
            if d <= 0:
                msg = f"Bugün son ödeme günü! 💳 {name} kartının asgari ödemesi ₺{min_pay:,.0f}"
                ico = "🚨"
            elif d == 1:
                msg = f"Yarın son gün! 💳 {name} kartının asgari ödemesi ₺{min_pay:,.0f}"
                ico = "⚠️"
            else:
                msg = f"{d} gün içinde: {name} kartının asgari ödemesi ₺{min_pay:,.0f}"
                ico = "💳"
            add(ico, "Kredi Kartı Ödemesi", msg, d, "kart")

    # ── RECURRING PAYMENTS ────────────────────────────────────────────────
    recurrings = db.execute(
        "SELECT * FROM recurring WHERE profile_id=? AND active=1", (pid,)
    ).fetchall()
    for r in recurrings:
        dom = r["day_of_month"] or 1
        if today.day <= dom:
            due_date = today.replace(day=dom)
        else:
            if today.month == 12:
                due_date = today.replace(year=today.year+1, month=1, day=dom)
            else:
                due_date = today.replace(month=today.month+1, day=dom)
        d = days_until(due_date)
        if -1 <= d <= 7:
            desc = r["description"] or r["category"]
            amt  = float(r["amount"])
            rtype = r["type"]
            if rtype == "gelir":
                if d <= 0:
                    msg = f"Bugün beklenen gelir: {desc} — ₺{amt:,.0f} 🎉"
                    ico = "🎉"
                elif d == 1:
                    msg = f"Yarın beklenen gelir: {desc} — ₺{amt:,.0f} 😊"
                    ico = "😊"
                else:
                    msg = f"{d} gün içinde beklenen gelir: {desc} — ₺{amt:,.0f} 💰"
                    ico = "💰"
            else:
                if d <= 0:
                    msg = f"Bugün ödenmesi gereken: {desc} — ₺{amt:,.0f}"
                    ico = "🔔"
                elif d == 1:
                    msg = f"Yarın ödenmesi gereken: {desc} — ₺{amt:,.0f}"
                    ico = "⏰"
                else:
                    msg = f"{d} gün içinde: {desc} — ₺{amt:,.0f}"
                    ico = "📅"
            add(ico, "Düzenli Ödeme" if rtype=="gider" else "Beklenen Gelir", msg, d, "duzenli")

    # ── BUDGET OVERRUN ─────────────────────────────────────────────────────
    from calendar import monthrange as mr
    _, last_day = mr(today.year, today.month)
    m_start = today.replace(day=1).isoformat()
    m_end   = today.replace(day=last_day).isoformat()
    budgets = db.execute("SELECT * FROM budgets WHERE profile_id=?", (pid,)).fetchall()
    for b in budgets:
        limit = b["limit_"] or 0
        if limit <= 0:
            continue
        spent_row = db.execute(
            "SELECT COALESCE(SUM(amount),0) as s FROM transactions WHERE profile_id=? AND category=? AND type='gider' AND date BETWEEN ? AND ?",
            (pid, b["category"], m_start, m_end)
        ).fetchone()
        spent = float(spent_row["s"] or 0)
        if spent >= limit:
            pct = round(spent / limit * 100)
            msg = f"{b['category']} bütçesi aşıldı! Hedef ₺{limit:,.0f}, harcama ₺{spent:,.0f} (%{pct})"
            add("🎯", "Bütçe Aşımı", msg, 0, "butce")

    # ── ŞİRKET: TURKISH TAX CALENDAR ─────────────────────────────────────
    if profile_type == "sirket":
        year  = today.year
        month = today.month

        def tax(icon, title, body, due_date):
            d = days_until(due_date)
            if -3 <= d <= 14:
                add(icon, title, body, d, "vergi")

        # Monthly deadlines — due the 26th of current month
        monthly_26 = today.replace(day=26) if today.day <= 26 else (
            today.replace(month=month+1, day=26) if month < 12
            else today.replace(year=year+1, month=1, day=26)
        )
        tax("📋", "SGK Primi",
            f"SGK primi bildirimi ve ödemesi — son gün {monthly_26.strftime('%d %B')}",
            monthly_26)
        tax("📄", "Muhtasar Beyanname",
            f"Aylık muhtasar beyanname — son gün {monthly_26.strftime('%d %B')}",
            monthly_26)
        tax("📊", "KDV Beyannamesi",
            f"Aylık KDV beyannamesi — son gün {monthly_26.strftime('%d %B')}",
            monthly_26)
        tax("🪙", "Damga Vergisi",
            f"Damga vergisi beyannamesi — son gün {monthly_26.strftime('%d %B')}",
            monthly_26)

        # Ba-Bs: 5th of each month (for 2 months prior)
        monthly_5 = today.replace(day=5) if today.day <= 5 else (
            today.replace(month=month+1, day=5) if month < 12
            else today.replace(year=year+1, month=1, day=5)
        )
        tax("📑", "Ba-Bs Bildirimi",
            f"Ba-Bs form bildirimi — son gün {monthly_5.strftime('%d %B')}",
            monthly_5)

        # Geçici Vergi: quarterly — 2nd month of each quarter, 14th
        gecici_months = {2: "1. Çeyrek", 5: "2. Çeyrek", 8: "3. Çeyrek", 11: "4. Çeyrek"}
        for gm, label in gecici_months.items():
            gd = date(year, gm, 14)
            tax("💹", f"Geçici Vergi ({label})",
                f"{label} geçici vergi beyannamesi — son gün {gd.strftime('%d %B')}",
                gd)

        # Annual deadlines
        annual = [
            (date(year, 4, 30), "🏛️",  "Kurumlar Vergisi",
             f"Yıllık kurumlar vergisi beyannamesi — son gün 30 Nisan"),
            (date(year, 3, 31), "💼",  "Gelir Vergisi 1. Taksit",
             f"Gelir vergisi 1. taksit ödemesi — son gün 31 Mart"),
            (date(year, 7, 31), "💼",  "Gelir Vergisi 2. Taksit",
             f"Gelir vergisi 2. taksit ödemesi — son gün 31 Temmuz"),
        ]
        for ad, aico, atitle, abody in annual:
            tax(aico, atitle, abody, ad)

    # ── SUPPLIER INVOICE DUE DATES ────────────────────────────────────────────
    sup_invs = db.execute(
        "SELECT * FROM supplier_invoices WHERE profile_id=? AND status='bekliyor'", (pid,)
    ).fetchall()
    for si in sup_invs:
        try:
            due = date.fromisoformat(si["due_date"])
        except:
            continue
        d_val = days_until(due)
        if -5 <= d_val <= 14:
            amt   = float(si["amount"])
            sname = si["supplier_name"]
            if d_val <= 0:
                msg = f"⚠️ Gecikmiş ödeme: {sname} — ₺{amt:,.0f} ({abs(d_val)} gün gecikti)"
                ico = "🚨"
            elif d_val <= 2:
                msg = f"Yarın/bugün ödeme: {sname} — ₺{amt:,.0f}"
                ico = "⚠️"
            else:
                msg = f"{d_val} gün içinde: {sname} tedarikçi ödemesi ₺{amt:,.0f}"
                ico = "🏭"
            add(ico, "Tedarikçi Ödemesi", msg, d_val, "tedarikci")

    # ── ASSET MAINTENANCE & INSURANCE ────────────────────────────────────────
    assets_rows = db.execute(
        "SELECT * FROM assets WHERE profile_id=? AND active=1", (pid,)
    ).fetchall()
    for a in assets_rows:
        name = a["name"]
        if a["maintenance_date"]:
            try:
                mdate = date.fromisoformat(a["maintenance_date"])
                d_val = days_until(mdate)
                if -3 <= d_val <= 21:
                    if d_val <= 0:
                        msg = f"{name} bakım tarihi geçti! Servis planlamanız gerekiyor."
                        ico = "🔧"
                    else:
                        msg = f"{name} için {d_val} gün içinde bakım/servis planlanmış"
                        ico = "🔧"
                    add(ico, "Araç/Varlık Bakımı", msg, d_val, "bakim")
            except:
                pass
        if a["insurance_date"]:
            try:
                idate = date.fromisoformat(a["insurance_date"])
                d_val = days_until(idate)
                if -3 <= d_val <= 30:
                    if d_val <= 0:
                        msg = f"{name} sigortası süresi doldu! Yenilenmesi gerekiyor."
                        ico = "🛡️"
                    else:
                        msg = f"{name} sigortası {d_val} gün içinde yenilenmeli"
                        ico = "🛡️"
                    add(ico, "Sigorta Yenileme", msg, d_val, "sigorta")
            except:
                pass

    # Sort: urgent first, then by days
    notifs.sort(key=lambda x: (0 if x["urgency"]=="urgent" else 1 if x["urgency"]=="soon" else 2, x["days"]))
    return notifs


@app.route("/api/notifications")
@login_required
def api_notifications():
    db   = get_db()
    pid  = get_pid()
    uid  = session["user_id"]
    prof = db.execute("SELECT type FROM profiles WHERE id=?", (pid,)).fetchone()
    ptype = prof["type"] if prof else "sahis"
    today_ = date.today()
    notifs = generate_notifications(pid, ptype, db, today_)
    return jsonify({"ok": True, "items": notifs, "count": len(notifs)})


@app.route("/api/motivation")
@login_required
def motivation():
    db = get_db(); today = date.today()
    year, month = today.year, today.month
    start, end  = month_range(year, month)
    pid = get_pid()
    rows = db.execute("SELECT type,SUM(amount) as t FROM transactions WHERE profile_id=? AND date BETWEEN ? AND ? GROUP BY type",(pid,start,end)).fetchall()
    gelir  = next((r["t"] for r in rows if r["type"]=="gelir"),0) or 0
    gider  = next((r["t"] for r in rows if r["type"]=="gider"),0) or 0
    net    = gelir - gider
    from calendar import monthrange as mr
    _, days_in_month = mr(year,month)
    days_passed = today.day; days_left = days_in_month - days_passed
    daily_spend = gider / max(days_passed,1)
    projected   = daily_spend * days_in_month
    bal = (db.execute("SELECT SUM(CASE WHEN type='gelir' THEN amount ELSE -amount END) as b FROM transactions WHERE profile_id=?",(pid,)).fetchone()["b"] or 0)
    budgets = {r["category"]:r["limit_"] for r in db.execute("SELECT * FROM budgets WHERE profile_id=?",(pid,)).fetchall()}
    top = db.execute("SELECT category,SUM(amount) as t FROM transactions WHERE profile_id=? AND date BETWEEN ? AND ? AND type='gider' GROUP BY category ORDER BY t DESC LIMIT 1",(pid,start,end)).fetchall()
    top_cat = top[0]["category"] if top else None; top_amt = top[0]["t"] if top else 0
    msgs=[]; score=0
    if gelir==0:
        msgs.append({"icon":"💡","text":"Henüz bu ay gelir girmedin. Maaş veya gelirini ekle, analiz yapayım."})
        return jsonify({"score":0,"label":"Veri Bekleniyor","color":"#64748b","msgs":msgs,"projected_month":0,"daily_avg":0,"days_left":days_left,"net":0,"save_rate":0})
    save_rate = net/gelir*100
    if save_rate>=30: score+=40; msgs.append({"icon":"🔥","text":f"Harika! Gelirinin %{save_rate:.0f}'ini biriktiriyorsun. Finansal özgürlük yolundasın."})
    elif save_rate>=15: score+=25; msgs.append({"icon":"✅","text":f"İyi gidiyorsun — gelirinin %{save_rate:.0f}'ini tasarruf ediyorsun."})
    elif save_rate>=0:  score+=10; msgs.append({"icon":"⚠️","text":f"Bu ay gelirinin sadece %{save_rate:.0f}'ini biriktirebildin. Hedef en az %20."})
    else: msgs.append({"icon":"🚨","text":f"Bu ay gelirininden {abs(net):,.0f}₺ fazla harcadın!"})
    if days_left>0:
        if projected < gelir*0.85: score+=20; msgs.append({"icon":"📈","text":f"Ay sonuna {days_left} gün kaldı. Bu hızla {gelir-projected:,.0f}₺ biriktireceksin!"})
        elif projected < gelir:    score+=10; msgs.append({"icon":"📊","text":f"Günlük ort. {daily_spend:,.0f}₺ harcıyorsun. Ay sonu tahmini: {projected:,.0f}₺"})
        else: msgs.append({"icon":"⚡","text":f"Dikkat! Bu hızda ay sonunda {projected:,.0f}₺ harcarsin, gelirine göre fazla!"})
    if budgets:
        over=[]
        for cat,lim in budgets.items():
            sp=(db.execute("SELECT SUM(amount) as t FROM transactions WHERE profile_id=? AND date BETWEEN ? AND ? AND type='gider' AND category=?",(pid,start,end,cat)).fetchone()["t"] or 0)
            if sp>lim: over.append((cat,sp-lim))
        if not over: score+=20; msgs.append({"icon":"🎯","text":"Tüm bütçe hedeflerini tutturuyorsun. Süper disiplin!"})
        else:
            for cat,exc in over[:2]: msgs.append({"icon":"🔴","text":f"'{cat}' bütçeni {exc:,.0f}₺ aştın."})
    else: score+=10; msgs.append({"icon":"💡","text":"Bütçe hedefleri belirle — harcamaların otomatik takip edilsin."})
    if top_cat and top_amt>0 and gelir>0:
        pct=top_amt/gelir*100
        if pct>40: msgs.append({"icon":"🔍","text":f"En büyük giderin '{top_cat}' — gelirinin %{pct:.0f}'i gidiyor."})
        else: score+=10; msgs.append({"icon":"✨","text":f"En büyük gider: '{top_cat}' {top_amt:,.0f}₺ (%{pct:.0f}). Dengeli görünüyor."})
    if bal>=100000: score+=10; msgs.append({"icon":"🏆","text":f"Toplam birikimin {bal:,.0f}₺! 100K kulübüne girdin."})
    elif bal>=50000:
        score+=8; left=100000-bal; mos=left/(net if net>0 else 1)
        msgs.append({"icon":"🎯","text":f"100K₺ hedefine {left:,.0f}₺ kaldı. Bu hızla ~{mos:.0f} ayda ulaşırsın."})
    elif bal>0 and net>0: score+=5; msgs.append({"icon":"🌱","text":f"Aylık {net:,.0f}₺ tasarrufla 1 yılda {net*12:,.0f}₺ yaparsın!"})
    score=min(100,max(0,score))
    if score>=75:   label,color="Finansal Sağlık: Mükemmel","#22c55e"
    elif score>=50: label,color="Finansal Sağlık: İyi","#3b82f6"
    elif score>=25: label,color="Finansal Sağlık: Orta","#f59e0b"
    else:           label,color="Finansal Sağlık: Zayıf","#ef4444"
    return jsonify({"score":score,"label":label,"color":color,"msgs":msgs,
                    "projected_month":round(projected,2),"daily_avg":round(daily_spend,2),
                    "days_left":days_left,"net":round(net,2),"save_rate":round(save_rate,1)})

# ── CSV import ────────────────────────────────────────────────────────────────

def _parse_amount(s):
    s = s.strip().replace(" ","").replace("\xa0","")
    s = re.sub(r"[₺$€£+]","",s)
    s = s.lstrip("-")
    if not s: return None
    dots=s.count("."); commas=s.count(",")
    if   dots==0 and commas==0: pass
    elif dots>=1 and commas==0:
        if dots==1: pass
        else: s=s.replace(".","")
    elif dots==0 and commas==1: s=s.replace(",",".")
    elif dots==0 and commas>1:  s=s.replace(",","")
    elif dots==1 and commas==1:
        if s.index(".")<s.index(","): s=s.replace(".","").replace(",",".")
        else: s=s.replace(",","")
    elif dots>1 and commas==1:  s=s.replace(".","").replace(",",".")
    elif dots==1 and commas>1:  s=s.replace(",","")
    try: return abs(float(s))
    except: return None

def _parse_date(s):
    s=s.strip()
    for fmt in ("%d.%m.%Y","%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%d.%m.%y","%m/%d/%Y"):
        try: return datetime.strptime(s,fmt).strftime("%Y-%m-%d")
        except: pass
    return None

_CAT_MAP = {
    "maaş":"Maaş","maas":"Maaş","salary":"Maaş","ücret":"Maaş",
    "kira":"Kira / Mortgage","mortgage":"Kira / Mortgage","aidat":"Faturalar",
    "elektrik":"Faturalar","doğalgaz":"Faturalar","dogalgaz":"Faturalar","fatura":"Faturalar",
    "internet":"Abonelikler","netflix":"Abonelikler","spotify":"Abonelikler","abonelik":"Abonelikler",
    "market":"Market / Gıda","migros":"Market / Gıda","a101":"Market / Gıda",
    "bim":"Market / Gıda","carrefour":"Market / Gıda","şok":"Market / Gıda",
    "taxi":"Ulaşım","taksi":"Ulaşım","uber":"Ulaşım","metro":"Ulaşım",
    "otobüs":"Ulaşım","akaryakıt":"Ulaşım","benzin":"Ulaşım","shell":"Ulaşım","opet":"Ulaşım",
    "restoran":"Yemek / Restoran","cafe":"Yemek / Restoran","kahve":"Yemek / Restoran",
    "yemeksepeti":"Yemek / Restoran","getir yemek":"Yemek / Restoran",
    "sinema":"Eğlence","konser":"Eğlence","bilet":"Eğlence",
    "eczane":"Sağlık","hastane":"Sağlık","doktor":"Sağlık","ilaç":"Sağlık",
    "giyim":"Giyim","zara":"Giyim","h&m":"Giyim","lcw":"Giyim",
    "okul":"Eğitim","kurs":"Eğitim","udemy":"Eğitim",
    "faiz":"Yatırım / Temettü","temettü":"Yatırım / Temettü",
    "hediye":"Hediye / İkramiye","ikramiye":"Hediye / İkramiye",
}

def _guess_cat(desc, ttype):
    dl=desc.lower()
    for kw,cat in _CAT_MAP.items():
        if kw in dl: return cat
    return "Diğer Gelir" if ttype=="gelir" else "Diğer Gider"

def _detect_cols(header):
    h=[c.lower().strip() for c in header]
    def find(*names):
        for n in names:
            for i,c in enumerate(h):
                if n in c: return i
        return None
    return (find("tarih","date","işlem tarihi","valör"),
            find("açıklama","aciklama","işlem","description","detay","karşı taraf"),
            find("borç","borc","çıkış","debit"),
            find("alacak","giriş","credit"),
            find("tutar","amount","miktar") if find("borç","borc") is None and find("alacak") is None else None)

@app.route("/api/import/preview", methods=["POST"])
@login_required
def import_preview():
    f = request.files.get("file")
    if not f: return jsonify({"ok":False,"error":"Dosya yüklenmedi"}),400
    raw = f.read().decode("utf-8-sig",errors="replace")
    dialect = "excel"
    if raw.count(";")>raw.count(","):
        csv.register_dialect("excel-semicolon",delimiter=";"); dialect="excel-semicolon"
    reader = csv.reader(io.StringIO(raw),dialect)
    rows   = [r for r in reader if any(c.strip() for c in r)]
    if len(rows)<2: return jsonify({"ok":False,"error":"CSV çok kısa"}),400
    header_idx=0
    for i,row in enumerate(rows):
        if any(kw in " ".join(row).lower() for kw in ("tarih","date","tutar","açıklama","işlem")):
            header_idx=i; break
    header=rows[header_idx]; data_rows=rows[header_idx+1:]
    dc,desc_c,deb_c,cred_c,amt_c = _detect_cols(header)
    parsed=[]; skipped=0
    for row in data_rows:
        if not any(c.strip() for c in row): continue
        dt = _parse_date(row[dc]) if dc is not None and dc<len(row) else None
        if not dt: skipped+=1; continue
        desc = row[desc_c].strip() if desc_c is not None and desc_c<len(row) else ""
        if amt_c is not None:
            raw_a=row[amt_c] if amt_c<len(row) else ""
            neg=raw_a.strip().startswith("-"); amt=_parse_amount(raw_a)
            if not amt: skipped+=1; continue
            ttype="gider" if neg else "gelir"
        elif deb_c is not None or cred_c is not None:
            dv=_parse_amount(row[deb_c])  if deb_c  is not None and deb_c <len(row) else None
            cv=_parse_amount(row[cred_c]) if cred_c is not None and cred_c<len(row) else None
            if dv and dv>0:   amt,ttype=dv,"gider"
            elif cv and cv>0: amt,ttype=cv,"gelir"
            else: skipped+=1; continue
        else: skipped+=1; continue
        parsed.append({"date":dt,"description":desc,"amount":round(amt,2),"type":ttype,"category":_guess_cat(desc,ttype)})
    return jsonify({"ok":True,"rows":parsed,"skipped":skipped,"headers":header})

# ── CARDS ─────────────────────────────────────────────────────────────────────

@app.route("/api/cards", methods=["GET"])
@login_required
def list_cards():
    pid = get_pid()
    return jsonify([row_to_dict(r) for r in get_db().execute("SELECT * FROM cards WHERE profile_id=? ORDER BY bank_name",(pid,)).fetchall()])

@app.route("/api/cards", methods=["POST"])
@login_required
def add_card():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    bank = d.get("bank_name","").strip(); card = d.get("card_name","").strip()
    owner = d.get("owner","").strip()
    limit = float(d.get("limit_",0)); used = float(d.get("used_",0))
    due_day = int(d.get("due_day",1)); min_pct = float(d.get("min_pct",25))
    stmt_day = int(d.get("statement_day",20))
    card_type = d.get("card_type","kredi")
    if card_type not in ("kredi","banka","yemek","hediye"): card_type = "kredi"
    if not bank: return jsonify({"ok":False}), 400
    # Yemek kartlarının limiti olabilir (0 da geçerli)
    if card_type not in ("yemek","hediye") and limit <= 0: return jsonify({"ok":False}), 400
    db = get_db()
    cur = db.execute("INSERT INTO cards (user_id,profile_id,bank_name,card_name,owner,limit_,used_,due_day,min_pct,statement_day,card_type,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                     (uid,pid,bank,card,owner,limit,used,due_day,min_pct,stmt_day,card_type,datetime.now().isoformat()))
    db.commit(); return jsonify({"ok":True,"id":cur.lastrowid})

@app.route("/api/cards/<int:cid>", methods=["PUT"])
@login_required
def update_card(cid):
    pid = get_pid()
    d = request.get_json(force=True); fields=[]; params=[]
    for col in ("bank_name","card_name","owner","limit_","used_","due_day","min_pct","statement_day","card_type"):
        if col in d:
            fields.append(f"{col}=?")
            params.append(float(d[col]) if col in ("limit_","used_","min_pct") else int(d[col]) if col in ("due_day","statement_day") else d[col])
    if not fields: return jsonify({"ok":False}),400
    params += [cid, pid]
    get_db().execute(f"UPDATE cards SET {','.join(fields)} WHERE id=? AND profile_id=?",params); get_db().commit()
    return jsonify({"ok":True})

@app.route("/api/cards/<int:cid>", methods=["DELETE"])
@login_required
def del_card(cid):
    pid = get_pid()
    get_db().execute("DELETE FROM cards WHERE id=? AND profile_id=?",(cid,pid)); get_db().commit()
    return jsonify({"ok":True})

@app.route("/api/cards/<int:cid>/pay", methods=["POST"])
@login_required
def pay_card(cid):
    """Kredi kartı ödemesi: gider işlemi oluştur + kart borcunu düş."""
    uid = session["user_id"]; pid = get_pid()
    db  = get_db()
    card = db.execute("SELECT * FROM cards WHERE id=? AND profile_id=?", (cid, pid)).fetchone()
    if not card:
        return jsonify({"ok": False, "error": "Kart bulunamadı"}), 404
    d = request.get_json(force=True)
    amount = float(d.get("amount", 0))
    account_id = d.get("account_id") or None  # ödeme yapılan vadesiz hesap
    if amount <= 0:
        return jsonify({"ok": False, "error": "Geçersiz tutar"}), 400
    current_debt = float(card["used_"] or 0)
    amount = min(amount, current_debt)
    card_type = card.get("card_type") or "kredi"
    cat = "Yemek Kartı Ödemesi" if card_type == "yemek" else "Kredi Kartı Ödemesi"
    desc = d.get("description") or f"{card['bank_name']} {card.get('card_name') or ''} ödemesi".strip()
    # Ödeme işlemini kaydet — vadesiz hesap ve kartla ilişkilendir
    db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,account_id,card_id,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (uid, pid, "gider", amount, cat, desc, date.today().isoformat(), account_id, cid, datetime.now().isoformat())
    )
    # Kart borcunu düş (transaction insert'teki card_id zaten etkisini yapıyor — burada manuel düzeltiriz)
    new_debt = max(0, current_debt - amount)
    db.execute("UPDATE cards SET used_=? WHERE id=?", (new_debt, cid))
    db.commit()
    return jsonify({"ok": True, "paid": amount, "new_debt": new_debt,
                    "limit": float(card["limit_"] or 0),
                    "available": float(card["limit_"] or 0) - new_debt})

# ── TRANSFER ─────────────────────────────────────────────────────────────────

@app.route("/api/transfers", methods=["POST"])
@login_required
def create_transfer():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    from_id = d.get("from_account_id")
    to_id   = d.get("to_account_id")
    amount  = float(d.get("amount", 0))
    dt      = d.get("date", today_str())
    desc    = d.get("description") or "Hesaplar Arası Transfer"
    if not from_id or not to_id or amount <= 0:
        return jsonify({"ok": False, "error": "Geçersiz veri"}), 400
    if int(from_id) == int(to_id):
        return jsonify({"ok": False, "error": "Kaynak ve hedef hesap aynı olamaz"}), 400
    db = get_db()
    ok1 = db.execute("SELECT id FROM accounts WHERE id=? AND profile_id=?", (from_id, pid)).fetchone()
    ok2 = db.execute("SELECT id FROM accounts WHERE id=? AND profile_id=?", (to_id, pid)).fetchone()
    if not ok1 or not ok2:
        return jsonify({"ok": False, "error": "Hesap bulunamadı"}), 404
    now = datetime.now().isoformat()
    db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,account_id,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (uid, pid, "gider", amount, "Hesaplar Arası Transfer", desc, dt, from_id, now))
    db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,account_id,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (uid, pid, "gelir", amount, "Hesaplar Arası Transfer", desc, dt, to_id, now))
    db.commit()
    return jsonify({"ok": True})

# ── RATES & INVESTMENTS ───────────────────────────────────────────────────────

_rates_cache = {"data": None, "ts": 0}
_rates_lock  = threading.Lock()

def _fetch_rates():
    """Fetch USD/TRY, EUR/TRY, Gold/g TRY from free APIs."""
    try:
        # Exchange rates: 1 TRY = X of each
        r = requests.get("https://open.er-api.com/v6/latest/TRY", timeout=8).json()
        rates = r.get("rates", {})
        usd_try = round(1 / rates["USD"], 4) if rates.get("USD") else None
        eur_try = round(1 / rates["EUR"], 4) if rates.get("EUR") else None
        gbp_try = round(1 / rates["GBP"], 4) if rates.get("GBP") else None
    except Exception:
        usd_try = eur_try = gbp_try = None

    gold_try = None
    try:
        # Gold spot in USD/oz → convert to TRY/gram
        g = requests.get("https://api.metals.live/v1/spot/gold", timeout=8).json()
        gold_usd_oz = float(g[0]["price"]) if isinstance(g, list) and g else None
        if gold_usd_oz and usd_try:
            gold_try = round(gold_usd_oz / 31.1035 * usd_try, 2)
    except Exception:
        pass

    return {
        "usd_try": usd_try,
        "eur_try": eur_try,
        "gbp_try": gbp_try,
        "gold_try": gold_try,   # TRY per gram
        "updated": datetime.now().strftime("%H:%M")
    }

def get_rates():
    with _rates_lock:
        if time.time() - _rates_cache["ts"] > 300 or not _rates_cache["data"]:
            _rates_cache["data"] = _fetch_rates()
            _rates_cache["ts"]   = time.time()
        return _rates_cache["data"]

@app.route("/api/rates")
@login_required
def api_rates():
    return jsonify(get_rates())

@app.route("/api/investments", methods=["GET"])
@premium_required
def list_investments():
    pid = get_pid()
    rows = get_db().execute("SELECT * FROM investments WHERE profile_id=? ORDER BY buy_date DESC",(pid,)).fetchall()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/investments", methods=["POST"])
@premium_required
def add_investment():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    name       = d.get("name","").strip()
    itype      = d.get("itype","")
    symbol     = d.get("symbol","").strip().upper()
    quantity   = float(d.get("quantity", 0))
    buy_price  = float(d.get("buy_price", 0))
    buy_date   = d.get("buy_date", today_str())
    note       = d.get("note","")
    account_id = d.get("account_id") or None
    if not name or itype not in ("doviz","altin","fon","hisse") or quantity <= 0 or buy_price <= 0:
        return jsonify({"ok": False, "error": "Geçersiz veri"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO investments (user_id,profile_id,name,itype,symbol,quantity,buy_price,buy_date,note,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (uid, pid, name, itype, symbol, quantity, buy_price, buy_date, note, datetime.now().isoformat()))
    # Nakit çıkışını transaction olarak da kaydet (rapor için)
    _icat = {"doviz":"Döviz Alımı","altin":"Altın Alımı","fon":"Yatırım Fonu","hisse":"Hisse Senedi"}
    cat  = _icat.get(itype, "Yatırım")
    desc = name + (f" ({symbol})" if symbol else "") + (f" — {note}" if note else "")
    total_cost = round(quantity * buy_price, 2)
    db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,account_id,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (uid, pid, "gider", total_cost, cat, desc, buy_date, account_id, datetime.now().isoformat()))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/investments/<int:iid>", methods=["DELETE"])
@premium_required
def del_investment(iid):
    pid = get_pid()
    get_db().execute("DELETE FROM investments WHERE id=? AND profile_id=?", (iid,pid)); get_db().commit()
    return jsonify({"ok": True})

@app.route("/api/investments/<int:iid>/sell", methods=["POST"])
@premium_required
def sell_investment(iid):
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    sell_price = float(d.get("sell_price", 0))
    sell_date  = d.get("sell_date", today_str())
    account_id = d.get("account_id") or None
    if sell_price <= 0:
        return jsonify({"ok": False, "error": "Satış fiyatı geçersiz"}), 400
    db = get_db()
    inv = db.execute("SELECT * FROM investments WHERE id=? AND profile_id=?", (iid, pid)).fetchone()
    if not inv:
        return jsonify({"ok": False, "error": "Yatırım bulunamadı"}), 404
    proceeds = round(inv["quantity"] * sell_price, 2)
    cost     = round(inv["quantity"] * float(inv["buy_price"]), 2)
    profit   = round(proceeds - cost, 2)
    _icat = {"doviz":"Döviz Alımı","altin":"Altın Alımı","fon":"Yatırım Fonu","hisse":"Hisse Senedi"}
    desc = f"{inv['name']} satışı" + (f" (K/Z: ₺{profit:+,.2f})" if profit else "")
    db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,account_id,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (uid, pid, "gelir", proceeds, "Yatırım Geliri / Satış", desc, sell_date, account_id, datetime.now().isoformat()))
    db.execute("DELETE FROM investments WHERE id=? AND profile_id=?", (iid, pid))
    db.commit()
    return jsonify({"ok": True, "proceeds": proceeds, "profit": profit})

@app.route("/api/investments/value", methods=["GET"])
@premium_required
def investment_values():
    pid   = get_pid()
    rows  = get_db().execute("SELECT * FROM investments WHERE profile_id=?",(pid,)).fetchall()
    rates = get_rates()
    result = []
    for row in rows:
        r = row_to_dict(row)
        cur_price = None
        cur_try   = None

        if r["itype"] == "doviz":
            sym = r["symbol"].upper()
            key = f"{sym.lower()}_try"
            rate = rates.get(key)
            if rate:
                cur_price = rate
                cur_try   = round(r["quantity"] * rate, 2)

        elif r["itype"] == "altin":
            if rates.get("gold_try"):
                cur_price = rates["gold_try"]
                cur_try   = round(r["quantity"] * rates["gold_try"], 2)

        elif r["itype"] in ("fon", "hisse"):
            # manual price: user must provide current_price via query
            cp = request.args.get(f"price_{r['id']}", type=float)
            if cp:
                cur_price = cp
                cur_try   = round(r["quantity"] * cp, 2)

        buy_try    = round(r["quantity"] * r["buy_price"], 2)
        profit_try = round(cur_try - buy_try, 2) if cur_try is not None else None
        profit_pct = round((cur_try - buy_try) / buy_try * 100, 2) if cur_try and buy_try else None

        r.update({
            "cur_price":  cur_price,
            "cur_try":    cur_try,
            "buy_try":    buy_try,
            "profit_try": profit_try,
            "profit_pct": profit_pct,
        })
        result.append(r)
    return jsonify(result)

@app.route("/api/investments/<int:iid>/book-income", methods=["POST"])
@premium_required
def book_investment_income(iid):
    """Add investment profit as a gelir transaction for this month."""
    d = request.get_json(force=True)
    amount = float(d.get("amount", 0))
    cat    = d.get("category", "Yatırım / Temettü")
    desc   = d.get("description", "Yatırım Getirisi")
    dt     = d.get("date", today_str())
    if amount <= 0: return jsonify({"ok": False}), 400
    db = get_db()
    uid = session["user_id"]; pid = get_pid()
    cur = db.execute(
        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,created_at) VALUES (?,?,?,?,?,?,?,?)",
        (uid, pid, "gelir", amount, cat, desc, dt, datetime.now().isoformat()))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/tefas/<fon_kod>")
@login_required
def tefas_fon(fon_kod):
    """Fetch latest NAV for a TEFAS fund code."""
    try:
        today_s = date.today().strftime("%d.%m.%Y")
        resp = requests.post(
            "https://www.tefas.gov.tr/api/DB/BindHistoryInfo",
            data={"fontip":"YAT","sfonkod":fon_kod.upper(),"bastarih":today_s,"bittarih":today_s},
            headers={"Referer":"https://www.tefas.gov.tr/"},
            timeout=8)
        data = resp.json().get("data", [])
        if not data:
            # try yesterday
            from datetime import timedelta
            yday = (date.today() - timedelta(days=1)).strftime("%d.%m.%Y")
            resp = requests.post(
                "https://www.tefas.gov.tr/api/DB/BindHistoryInfo",
                data={"fontip":"YAT","sfonkod":fon_kod.upper(),"bastarih":yday,"bittarih":yday},
                headers={"Referer":"https://www.tefas.gov.tr/"},
                timeout=8)
            data = resp.json().get("data", [])
        if data:
            row = data[0]
            return jsonify({"ok":True,"fiyat":float(row.get("FIYAT",0)),"tarih":row.get("TARIH",""),"fon":fon_kod.upper()})
        return jsonify({"ok":False,"error":"Fon bulunamadı"})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

# ── RECURRING ────────────────────────────────────────────────────────────────

@app.route("/api/recurring", methods=["GET"])
@login_required
def list_recurring():
    pid = get_pid()
    rows = get_db().execute("SELECT * FROM recurring WHERE profile_id=? ORDER BY type DESC, day_of_month",(pid,)).fetchall()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/recurring", methods=["POST"])
@login_required
def add_recurring():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    ttype = d.get("type"); amount = float(d.get("amount", 0))
    cat = d.get("category",""); desc = d.get("description","")
    days_raw = d.get("days_of_month", d.get("day_of_month", 1))
    if isinstance(days_raw, list):
        days_list = sorted(set(int(x) for x in days_raw if 1 <= int(x) <= 31))
    else:
        days_list = [int(days_raw)]
    if not days_list: days_list = [1]
    days_str = ",".join(str(x) for x in days_list)
    day_first = days_list[0]
    if ttype not in ("gelir","gider") or amount <= 0 or not cat:
        return jsonify({"ok": False, "error": "Geçersiz veri"}), 400
    db = get_db()
    cur = db.execute(
        "INSERT INTO recurring (user_id,profile_id,type,amount,category,description,day_of_month,days_of_month,active,created_at) VALUES (?,?,?,?,?,?,?,?,1,?)",
        (uid, pid, ttype, amount, cat, desc, day_first, days_str, datetime.now().isoformat()))
    db.commit()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/recurring/<int:rid>", methods=["DELETE"])
@login_required
def del_recurring(rid):
    pid = get_pid()
    get_db().execute("DELETE FROM recurring WHERE id=? AND profile_id=?", (rid,pid)); get_db().commit()
    return jsonify({"ok": True})

@app.route("/api/recurring/<int:rid>", methods=["PUT"])
@login_required
def update_recurring(rid):
    pid = get_pid()
    d = request.get_json(force=True)
    fields, params = [], []
    for col in ("type","amount","category","description","day_of_month","active"):
        if col in d:
            fields.append(f"{col}=?")
            val = float(d[col]) if col == "amount" else int(d[col]) if col in ("day_of_month","active") else d[col]
            params.append(val)
    if "days_of_month" in d:
        days_raw = d["days_of_month"]
        if isinstance(days_raw, list):
            days_str = ",".join(str(x) for x in sorted(set(int(x) for x in days_raw if 1 <= int(x) <= 31)))
        else:
            days_str = str(days_raw)
        fields.append("days_of_month=?"); params.append(days_str)
    if not fields: return jsonify({"ok": False}), 400
    params += [rid, pid]
    get_db().execute(f"UPDATE recurring SET {','.join(fields)} WHERE id=? AND profile_id=?", params)
    get_db().commit()
    return jsonify({"ok": True})

@app.route("/api/recurring/apply", methods=["POST"])
@login_required
def apply_recurring():
    uid = session["user_id"]; pid = get_pid()
    d = request.get_json(force=True)
    today_y     = date.today().year
    year_start  = int(d.get("year_start",  d.get("year", today_y)))
    year_end    = int(d.get("year_end",    year_start))
    month_start = int(d.get("month_start", 1))
    month_end   = int(d.get("month_end",   12))

    if year_end < year_start: year_end = year_start
    if year_end > today_y + 5: year_end = today_y + 5  # max 5 yıl ileri

    db        = get_db()
    templates = db.execute("SELECT * FROM recurring WHERE profile_id=? AND active=1",(pid,)).fetchall()
    created   = 0; skipped = 0

    from calendar import monthrange as _monthrange
    for yr in range(year_start, year_end + 1):
        m_from = month_start if yr == year_start else 1
        m_to   = month_end   if yr == year_end   else 12
        for tpl in templates:
            days_raw = (tpl["days_of_month"] or "").strip()
            days_list = [int(x.strip()) for x in days_raw.split(",") if x.strip().isdigit()] if days_raw else [tpl["day_of_month"]]
            for m in range(m_from, m_to + 1):
                _, last_day = _monthrange(yr, m)
                for d_val in days_list:
                    day = min(d_val, last_day)
                    dt  = f"{yr:04d}-{m:02d}-{day:02d}"
                    exists = db.execute(
                        "SELECT id FROM transactions WHERE profile_id=? AND type=? AND category=? AND description=? AND date=? AND amount=?",
                        (pid, tpl["type"], tpl["category"], tpl["description"], dt, tpl["amount"])
                    ).fetchone()
                    if exists: skipped += 1; continue
                    db.execute(
                        "INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,created_at) VALUES (?,?,?,?,?,?,?,?)",
                        (uid, pid, tpl["type"], tpl["amount"], tpl["category"], tpl["description"], dt, datetime.now().isoformat()))
                    created += 1

    db.commit()
    return jsonify({"ok": True, "created": created, "skipped": skipped})

@app.route("/api/backup/download")
@login_required
def backup_download():
    """Download the full DB as a SQL dump."""
    try:
        dump = _pg_dump_csv()
        now_str  = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"kirpi_backup_{now_str}.csv"
        return dump, 200, {
            "Content-Type":        "application/octet-stream",
            "Content-Disposition": f'attachment; filename="{filename}"',
        }
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500

@app.route("/api/import/confirm", methods=["POST"])
@login_required
def import_confirm():
    uid=session["user_id"]; pid=get_pid()
    rows=request.get_json(force=True)
    if not isinstance(rows,list): return jsonify({"ok":False}),400
    db=get_db(); now=datetime.now().isoformat(); count=0
    for r in rows:
        if not r.get("skip") and r.get("amount",0)>0:
            db.execute("INSERT INTO transactions (user_id,profile_id,type,amount,category,description,date,created_at) VALUES (?,?,?,?,?,?,?,?)",
                       (uid,pid,r["type"],r["amount"],r["category"],r.get("description",""),r["date"],now))
            count+=1
    db.commit(); return jsonify({"ok":True,"imported":count})

# ── HTML ──────────────────────────────────────────────────────────────────────


def AUTH_HTML_render(mode, msg="", token=None):
    ok_modes  = {"forgot_sent", "reset_invalid"}
    is_ok     = msg and (mode in ok_modes or msg.startswith("Kayıt") or msg.startswith("Şifreniz") or msg.startswith("✅"))
    msg_html  = f'<div class="msg {"ok" if is_ok else "err"}">{msg}</div>' if msg else ""

    titles = {
        "register":      "Hesap Oluştur",
        "login":         "Giriş Yap",
        "forgot":        "Şifremi Unuttum",
        "forgot_sent":   "Mail Gönderildi",
        "reset":         "Yeni Şifre Belirle",
        "reset_invalid": "Link Geçersiz",
        "resend_verify": "Doğrulama Maili",
    }
    page_title = titles.get(mode, "Kirpi")

    if mode == "register":
        form = (
    '''    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:24px">
      <div id="reg-sahis" onclick="setRegType(this,'sahis')" style="border:2px solid #6366f1;border-radius:12px;padding:14px 10px;text-align:center;cursor:pointer;background:#3a3a3c;transition:.2s">
        <div style="font-size:1.6rem;margin-bottom:4px">&#128100;</div>
        <div style="font-size:.82rem;font-weight:700;color:#f2f2f7">Bireysel</div>
        <div style="font-size:.7rem;color:#aeaeb2;margin-top:2px">Şahıs hesabı</div>
      </div>
      <div id="reg-sirket" onclick="setRegType(this,'sirket')" style="border:2px solid #48484a;border-radius:12px;padding:14px 10px;text-align:center;cursor:pointer;background:#2c2c2e;transition:.2s">
        <div style="font-size:1.6rem;margin-bottom:4px">&#127970;</div>
        <div style="font-size:.82rem;font-weight:700;color:#aeaeb2">Ticari</div>
        <div style="font-size:.7rem;color:#8e8e93;margin-top:2px">Şirket hesabı</div>
      </div>
    </div>''' + msg_html + '''
    <form method="POST" action="/register">
      <!-- CSRF_FIELD -->
      <input type="hidden" name="profile_type" id="reg-type-val" value="sahis">
      <label id="reg-name-label">Ad Soyad</label>
      <input type="text" name="display_name" id="reg-name" placeholder="Ad Soyad" autocomplete="name">
      <label>Kullanıcı Adı</label>
      <input type="text" name="username" placeholder="kullanici_adi" required autocomplete="username">
      <label>E-posta</label>
      <input type="email" name="email" placeholder="ornek@gmail.com" required autocomplete="email">
      <label>Telefon <span style="color:#8e8e93;font-size:.72rem;font-weight:400">(opsiyonel — şifre sıfırlama için)</span></label>
      <input type="tel" name="phone" placeholder="05XX XXX XX XX" autocomplete="tel">
      <label>Şifre (en az 6 karakter)</label>
      <div style="position:relative;margin-bottom:14px">
        <input type="password" name="password" id="reg-pw" required autocomplete="new-password" style="width:100%;padding-right:44px">
        <button type="button" onclick="togglePw('reg-pw','reg-eye')" id="reg-eye"
          style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;font-size:1.1rem;color:#8e8e93;padding:4px">👁</button>
      </div>
      <label>Şifre Tekrar</label>
      <div style="position:relative;margin-bottom:14px">
        <input type="password" name="confirm" id="reg-pw2" required autocomplete="new-password" style="width:100%;padding-right:44px">
        <button type="button" onclick="togglePw('reg-pw2','reg-eye2')" id="reg-eye2"
          style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;font-size:1.1rem;color:#8e8e93;padding:4px">👁</button>
      </div>
      <div style="display:flex;align-items:flex-start;gap:8px;margin-bottom:14px">
        <input type="checkbox" id="kvkk-check" required style="width:auto;margin:3px 0 0;flex-shrink:0">
        <label for="kvkk-check" style="font-size:.75rem;color:#64748b;cursor:pointer">
          <a href="/kvkk" target="_blank" style="color:#818cf8">KVKK</a>,
          <a href="/gizlilik" target="_blank" style="color:#818cf8">Gizlilik</a> ve
          <a href="/kullanim-kosullari" target="_blank" style="color:#818cf8">Kullanım Koşulları</a>\'nı okudum, kabul ediyorum.
        </label>
      </div>
      <button class="btn" type="submit" id="reg-btn">Bireysel Hesap Oluştur</button>
    </form>
    <div class="link">Zaten hesabın var mı? <a href="/login">Giriş Yap</a></div>
    <script>
    function togglePw(inputId,btnId){
      var inp=document.getElementById(inputId);
      var btn=document.getElementById(btnId);
      if(!inp)return;
      var show=inp.type==="password";
      inp.type=show?"text":"password";
      if(btn) btn.textContent=show?"🙈":"👁";
    }
    function setRegType(el,t){
      document.getElementById("reg-type-val").value=t;
      document.getElementById("reg-sahis").style.borderColor=t==="sahis"?"#6366f1":"#2a2f45";
      document.getElementById("reg-sahis").style.background=t==="sahis"?"#3a3a3c":"#2c2c2e";
      document.getElementById("reg-sirket").style.borderColor=t==="sirket"?"#6366f1":"#48484a";
      document.getElementById("reg-sirket").style.background=t==="sirket"?"#3a3a3c":"#2c2c2e";
      if(t==="sahis"){
        document.getElementById("reg-name-label").textContent="Ad Soyad";
        document.getElementById("reg-name").placeholder="Ad Soyad";
        document.getElementById("reg-btn").textContent="Bireysel Hesap Oluştur";
      } else {
        document.getElementById("reg-name-label").textContent="Şirket Ünvanı";
        document.getElementById("reg-name").placeholder="Şirket A.Ş.";
        document.getElementById("reg-btn").textContent="Ticari Hesap Oluştur";
      }
    }
    </script>''')

    elif mode == "login":
        form = msg_html + '''
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-bottom:24px">
      <div id="login-bir" onclick="setLoginType('bireysel')" style="border:2px solid #6366f1;border-radius:12px;padding:12px 10px;text-align:center;cursor:pointer;background:#3a3a3c;transition:.2s">
        <div style="font-size:1.3rem;margin-bottom:3px">&#128100;</div>
        <div id="login-bir-lbl" style="font-size:.82rem;font-weight:700;color:#f2f2f7">Bireysel</div>
      </div>
      <div id="login-tic" onclick="setLoginType('ticari')" style="border:2px solid #48484a;border-radius:12px;padding:12px 10px;text-align:center;cursor:pointer;background:#2c2c2e;transition:.2s">
        <div style="font-size:1.3rem;margin-bottom:3px">&#127970;</div>
        <div id="login-tic-lbl" style="font-size:.82rem;font-weight:700;color:#aeaeb2">Ticari</div>
      </div>
    </div>
    <form method="POST" action="/login">
      <!-- CSRF_FIELD -->
      <input type="hidden" name="remember_me" id="remember_me_val" value="0">
      <label id="login-uname-label">Kullanıcı Adı</label>
      <input type="text" name="username" required autocomplete="username">
      <label>Şifre</label>
      <div style="position:relative;margin-bottom:14px">
        <input type="password" name="password" id="login-pw" required autocomplete="current-password" style="width:100%;padding-right:44px">
        <button type="button" onclick="togglePw('login-pw','login-eye')" id="login-eye"
          style="position:absolute;right:10px;top:50%;transform:translateY(-50%);background:none;border:none;cursor:pointer;font-size:1.1rem;color:#8e8e93;padding:4px">👁</button>
      </div>
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;margin-top:-6px">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-size:.82rem;color:#6d6d72;margin-bottom:0">
          <div id="remember-toggle" onclick="toggleRemember()" style="width:40px;height:24px;border-radius:12px;background:#c7c7cc;position:relative;cursor:pointer;transition:background .2s;flex-shrink:0">
            <div id="remember-knob" style="width:20px;height:20px;border-radius:50%;background:#fff;position:absolute;top:2px;left:2px;transition:left .2s;box-shadow:0 1px 3px rgba(0,0,0,.2)"></div>
          </div>
          Beni Hatırla <span id="remember-lbl" style="color:#aaa">(30 gün)</span>
        </label>
        <a href="/forgot-password" style="font-size:.8rem;color:#007aff;text-decoration:none">Şifremi unuttum</a>
      </div>
      <button class="btn" type="submit" id="login-btn">Bireysel Giriş Yap</button>
    </form>
    <div class="link">Hesabın yok mu? <a href="/register">Kayıt Ol</a></div>
    <script>
    var _rem=false;
    function toggleRemember(){
      _rem=!_rem;
      document.getElementById("remember_me_val").value=_rem?"1":"0";
      document.getElementById("remember-toggle").style.background=_rem?"#007aff":"#c7c7cc";
      document.getElementById("remember-knob").style.left=_rem?"18px":"2px";
      document.getElementById("remember-lbl").style.color=_rem?"#007aff":"#aaa";
    }
    function togglePw(inputId,btnId){
      var inp=document.getElementById(inputId);
      var btn=document.getElementById(btnId);
      if(!inp)return;
      var show=inp.type==="password";
      inp.type=show?"text":"password";
      if(btn) btn.textContent=show?"🙈":"👁";
    }
    </script>
    <script>
    function setLoginType(t){
      var isBir=t==="bireysel";
      document.getElementById("login-bir").style.borderColor=isBir?"#6366f1":"#2a2f45";
      document.getElementById("login-bir").style.background=isBir?"#3a3a3c":"#2c2c2e";
      document.getElementById("login-bir-lbl").style.color=isBir?"#f2f2f7":"#aeaeb2";
      document.getElementById("login-tic").style.borderColor=!isBir?"#6366f1":"#48484a";
      document.getElementById("login-tic").style.background=!isBir?"#3a3a3c":"#2c2c2e";
      document.getElementById("login-tic-lbl").style.color=!isBir?"#f2f2f7":"#aeaeb2";
      document.getElementById("login-btn").textContent=isBir?"Bireysel Giriş Yap":"Ticari Giriş Yap";
      document.getElementById("login-uname-label").textContent=isBir?"Kullanıcı Adı":"Kullanıcı Adı / Vergi No";
    }
    </script>'''

    elif mode == "forgot":
        form = f"""
    <h2 style="font-size:1.1rem;font-weight:700;margin-bottom:8px;text-align:center">Şifremi Unuttum</h2>
    <p style="text-align:center;color:#64748b;font-size:.82rem;margin-bottom:20px">Kullanıcı adın, email veya telefon numaranla şifre sıfırlama linki alabilirsin.</p>
    {msg_html}
    <form method="POST" action="/forgot-password">
      <!-- CSRF_FIELD -->
      <label>Kullanıcı Adı, Email veya Telefon</label>
      <input type="text" name="identifier" required placeholder="kullanici_adi / ornek@gmail.com / 05XX..." autocomplete="username email tel">
      <div style="font-size:.72rem;color:#64748b;margin-top:-10px;margin-bottom:14px">📱 Telefon numarası girebilirsin — hesabına kayıtlı email'e link gönderilir</div>
      <button class="btn" type="submit">Sıfırlama Linki Gönder</button>
    </form>
    <div class="link"><a href="/login">← Giriş ekranına dön</a></div>"""

    elif mode == "forgot_sent":
        form = f"""
    <div style="text-align:center;padding:20px 0">
      <div style="font-size:3rem;margin-bottom:16px">📬</div>
      <h2 style="font-size:1.1rem;font-weight:700;margin-bottom:12px">Mail Gönderildi</h2>
      {msg_html}
      <p style="color:#64748b;font-size:.82rem;margin-top:12px">Spam/Junk klasörünü de kontrol etmeyi unutma.</p>
    </div>
    <div class="link"><a href="/login">← Giriş ekranına dön</a></div>"""

    elif mode == "reset":
        form = f"""
    <h2 style="font-size:1.1rem;font-weight:700;margin-bottom:20px;text-align:center">Yeni Şifre Belirle</h2>
    {msg_html}
    <form method="POST" action="/reset-password/{token}">
      <!-- CSRF_FIELD -->
      <label>Yeni Şifre <span style="color:#64748b">(en az 6 karakter)</span></label>
      <input type="password" name="password" required autocomplete="new-password">
      <label>Şifre Tekrar</label>
      <input type="password" name="confirm" required autocomplete="new-password">
      <button class="btn" type="submit">Şifremi Güncelle</button>
    </form>"""

    elif mode == "resend_verify":
        form = f"""
    <h2 style="font-size:1.1rem;font-weight:700;margin-bottom:8px;text-align:center">Doğrulama Maili Gönder</h2>
    <p style="text-align:center;color:#64748b;font-size:.82rem;margin-bottom:20px">Kayıt emailini girerek yeni doğrulama linki alabilirsin.</p>
    {msg_html}
    <form method="POST" action="/resend-verify">
      <!-- CSRF_FIELD -->
      <label>Email Adresi</label>
      <input type="email" name="email" required placeholder="ornek@gmail.com" autocomplete="email">
      <button class="btn" type="submit">Doğrulama Maili Gönder</button>
    </form>
    <div class="link"><a href="/login">← Giriş ekranına dön</a></div>"""

    else:  # reset_invalid
        form = f"""
    <div style="text-align:center;padding:20px 0">
      <div style="font-size:3rem;margin-bottom:16px">⚠️</div>
      <h2 style="font-size:1.1rem;font-weight:700;margin-bottom:12px">Link Geçersiz</h2>
      {{msg_html}}
    </div>
    <div class="link"><a href="/forgot-password">Yeni link al →</a></div>"""

    return (AUTH_HTML
        .replace("__MODE_TITLE__", page_title)
        .replace("  <!-- __FORM_CONTENT__ -->", form)
        .replace("<!-- CSRF_FIELD -->", _csrf_input()))


ICON_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 512 512">
  <defs>
    <linearGradient id="bg" x1="0%" y1="0%" x2="100%" y2="100%">
      <stop offset="0%" style="stop-color:#4f46e5"/>
      <stop offset="100%" style="stop-color:#7c3aed"/>
    </linearGradient>
    <linearGradient id="body" x1="0%" y1="0%" x2="100%" y2="100%">
      <stop offset="0%" style="stop-color:#6d28d9"/>
      <stop offset="100%" style="stop-color:#4c1d95"/>
    </linearGradient>
  </defs>

  <!-- background rounded square -->
  <rect width="512" height="512" rx="110" fill="url(#bg)"/>

  <!-- SPINES — sharp triangles rising from back -->
  <g fill="#c4b5fd" stroke="#a78bfa" stroke-width="2">
    <polygon points="256,52 240,130 272,130"/>
    <polygon points="300,62 282,138 318,142"/>
    <polygon points="342,82 320,154 358,162"/>
    <polygon points="378,112 352,178 390,190"/>
    <polygon points="404,148 376,208 414,224"/>
    <polygon points="418,192 390,244 428,262"/>
    <polygon points="212,62 230,138 194,142"/>
    <polygon points="172,84 196,156 158,164"/>
    <polygon points="140,116 166,182 128,194"/>
  </g>

  <!-- BODY — dark rounded ellipse -->
  <ellipse cx="268" cy="330" rx="175" ry="130" fill="url(#body)"/>

  <!-- FACE / HEAD — lighter round bulge to left -->
  <ellipse cx="148" cy="310" rx="98" ry="90" fill="#ddd6fe"/>

  <!-- inner face -->
  <ellipse cx="138" cy="318" rx="72" ry="66" fill="#ede9fe"/>

  <!-- EYE -->
  <circle cx="118" cy="292" r="22" fill="#1e1b4b"/>
  <circle cx="112" cy="284" r="8" fill="white"/>
  <circle cx="110" cy="283" r="4" fill="white" opacity="0.6"/>

  <!-- NOSE — shiny dark oval -->
  <ellipse cx="75" cy="322" rx="22" ry="16" fill="#4c1d95"/>
  <ellipse cx="70" cy="316" rx="8" ry="5" fill="#818cf8" opacity="0.5"/>

  <!-- SMILE -->
  <path d="M85 340 Q105 362 130 348" stroke="#6d28d9" stroke-width="7" fill="none" stroke-linecap="round"/>

  <!-- CHEEK blush -->
  <ellipse cx="148" cy="350" rx="28" ry="16" fill="#a78bfa" opacity="0.35"/>

  <!-- BELLY lighter patch -->
  <ellipse cx="295" cy="348" rx="90" ry="60" fill="#7c3aed" opacity="0.5"/>
  <ellipse cx="295" cy="355" rx="60" ry="38" fill="#c4b5fd" opacity="0.15"/>

  <!-- FEET -->
  <ellipse cx="195" cy="440" rx="52" ry="28" fill="#4c1d95"/>
  <ellipse cx="305" cy="448" rx="52" ry="28" fill="#4c1d95"/>
  <ellipse cx="408" cy="438" rx="46" ry="24" fill="#4c1d95"/>

  <!-- SMALL SPARKLE top-right -->
  <circle cx="420" cy="80" r="10" fill="white" opacity="0.3"/>
  <circle cx="448" cy="56" r="6" fill="white" opacity="0.2"/>
</svg>"""

MANIFEST = json.dumps({
    "name": "Kirpi — Nakit Akış Yönetimi",
    "short_name": "Kirpi",
    "description": "Gelir, gider, kart borçları ve bütçenizi tek yerden yönetin. Şahıs ve şirket profillerini ayrı tutun.",
    "start_url": "/",
    "scope": "/",
    "display": "standalone",
    "display_override": ["standalone", "minimal-ui"],
    "background_color": "#0a0c12",
    "theme_color": "#6366f1",
    "orientation": "portrait-primary",
    "lang": "tr",
    "categories": ["finance", "productivity"],
    "screenshots": [
        {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png", "form_factor": "narrow"},
    ],
    "shortcuts": [
        {"name": "Dashboard", "url": "/", "description": "Ana sayfa"},
    ],
    "icons": [
        {"src": "/icon.svg", "sizes": "any", "type": "image/svg+xml", "purpose": "any"},
        {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
        {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
    ],
    "prefer_related_applications": False,
}, ensure_ascii=False)

SW_JS = """
const CACHE='kirpi-v3';
const ASSETS=['/manifest.json'];
self.addEventListener('install',e=>{e.waitUntil(caches.open(CACHE).then(c=>c.addAll(ASSETS)));self.skipWaiting();});
self.addEventListener('activate',e=>{e.waitUntil(caches.keys().then(keys=>Promise.all(keys.filter(k=>k!==CACHE).map(k=>caches.delete(k)))));self.clients.claim();});
self.addEventListener('fetch',e=>{
  if(e.request.method!=='GET')return;
  if(e.request.url.includes('/api/'))return;
  e.respondWith(fetch(e.request).catch(()=>caches.match(e.request)));
});
"""

@app.route("/sw.js")
def sw(): return SW_JS, 200, {"Content-Type": "application/javascript"}

@app.route("/manifest.json")
def manifest(): return MANIFEST, 200, {"Content-Type": "application/manifest+json"}

@app.route("/health")
def health():
    try:
        with pg_connect() as con:
            con.execute("SELECT 1")
        return jsonify({"ok": True, "db": "ok"}), 200
    except Exception as exc:
        return jsonify({"ok": False, "db": str(exc)}), 503

# ── PREMIUM SAYFASI ───────────────────────────────────────────────────────────

@app.route("/premium")
def premium_page():
    uid = session.get("user_id")
    if uid:
        sub = get_sub_status(uid)
        if sub["is_premium"] and sub["status"] == "premium":
            current_info = f'<div style="background:#22c55e18;border:1px solid #22c55e40;border-radius:12px;padding:16px 20px;margin-bottom:24px;color:#22c55e;font-size:.9rem;font-weight:600">✅ Aktif Premium üyeliğiniz var — {sub["days_left"]} gün kaldı.</div>'
        elif sub["is_premium"] and sub["status"] == "trial":
            current_info = f'<div style="background:#6366f118;border:1px solid #6366f140;border-radius:12px;padding:16px 20px;margin-bottom:24px;color:#818cf8;font-size:.9rem;font-weight:600">🎁 Ücretsiz deneme süreniz: {sub["days_left"]} gün kaldı.</div>'
        else:
            current_info = '<div style="background:#ef444418;border:1px solid #ef444440;border-radius:12px;padding:16px 20px;margin-bottom:24px;color:#f87171;font-size:.9rem;font-weight:600">⚠️ Deneme süreniz dolmuş. Aşağıdan Premium\'a geçin.</div>'
    else:
        current_info = ""

    required = request.args.get("required")
    req_banner = '<div style="background:#f59e0b18;border:1px solid #f59e0b40;border-radius:12px;padding:14px 20px;margin-bottom:20px;color:#fbbf24;font-size:.85rem">🔒 Erişmek istediğiniz özellik Premium üyelik gerektirir.</div>' if required else ""

    csrf = get_csrf_token() if uid else ""

    return f"""<!DOCTYPE html>
<html lang="tr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Premium — Kirpi</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0a0c12;color:#e2e8f0;font-family:'Inter',system-ui,sans-serif;padding:40px 20px;min-height:100vh}}
.wrap{{max-width:760px;margin:0 auto}}
.back{{display:inline-flex;align-items:center;gap:8px;color:#818cf8;text-decoration:none;font-size:.85rem;margin-bottom:32px}}
h1{{font-size:2rem;font-weight:900;margin-bottom:8px;background:linear-gradient(135deg,#818cf8,#a78bfa,#f472b6);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.sub{{color:#64748b;font-size:.9rem;margin-bottom:36px}}
.plans{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-bottom:36px}}
@media(max-width:600px){{.plans{{grid-template-columns:1fr}}}}
.plan{{background:#111318;border:1px solid #1e2233;border-radius:16px;padding:28px 24px}}
.plan.featured{{border-color:#6366f1;background:linear-gradient(145deg,#111318,#16182a)}}
.plan-name{{font-size:.8rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase;color:#64748b;margin-bottom:12px}}
.plan.featured .plan-name{{color:#818cf8}}
.plan-price{{font-size:2.4rem;font-weight:900;color:#e2e8f0;margin-bottom:4px}}
.plan-price span{{font-size:1rem;font-weight:400;color:#64748b}}
.plan-period{{font-size:.8rem;color:#64748b;margin-bottom:24px}}
.plan-features{{list-style:none;margin-bottom:28px}}
.plan-features li{{font-size:.85rem;color:#94a3b8;padding:6px 0;display:flex;align-items:center;gap:8px;border-bottom:1px solid #1a1d26}}
.plan-features li:last-child{{border-bottom:none}}
.plan-features li .icon{{flex-shrink:0}}
.btn-plan{{width:100%;padding:13px;border:none;border-radius:10px;font-size:.92rem;font-weight:700;cursor:pointer;transition:.2s}}
.btn-free{{background:#1e2233;color:#94a3b8}}
.btn-premium{{background:linear-gradient(135deg,#6366f1,#a855f7);color:#fff}}
.btn-premium:hover{{opacity:.9}}
.contact-box{{background:#111318;border:1px solid #1e2233;border-radius:12px;padding:24px;text-align:center}}
.contact-box h2{{font-size:1rem;font-weight:700;color:#c7d2fe;margin-bottom:10px}}
.contact-box p{{font-size:.85rem;color:#64748b;margin-bottom:16px;line-height:1.6}}
.contact-box a{{color:#818cf8;font-weight:600}}
.badge{{display:inline-block;background:#6366f120;color:#818cf8;border:1px solid #6366f140;border-radius:20px;padding:3px 10px;font-size:.7rem;font-weight:700;margin-left:6px}}
</style></head>
<body><div class="wrap">
<a class="back" href="/">← Uygulamaya Dön</a>
{req_banner}
{current_info}
<h1>🦔 Kirpi Premium</h1>
<p class="sub">İlk {TRIAL_DAYS} gün ücretsiz — tüm özellikler açık. Sonrasında istediğiniz planı seçin.</p>

<div class="plans">
  <div class="plan">
    <div class="plan-name">Ücretsiz</div>
    <div class="plan-price">₺0 <span>/ay</span></div>
    <div class="plan-period">Sonsuza kadar ücretsiz</div>
    <ul class="plan-features">
      <li><span class="icon">✅</span> Gelir / Gider takibi</li>
      <li><span class="icon">✅</span> Bütçe yönetimi</li>
      <li><span class="icon">✅</span> Hesaplar &amp; kartlar</li>
      <li><span class="icon">✅</span> Hedefler &amp; görevler</li>
      <li><span class="icon">✅</span> 1 profil (Şahıs)</li>
      <li><span class="icon">❌</span> Yatırım takibi</li>
      <li><span class="icon">❌</span> Kıymet yönetimi</li>
      <li><span class="icon">❌</span> Tedarikçi / fatura</li>
      <li><span class="icon">❌</span> Excel &amp; PDF export</li>
      <li><span class="icon">❌</span> Telegram entegrasyonu</li>
      <li><span class="icon">❌</span> Gelişmiş AI analiz</li>
    </ul>
    <button class="btn-plan btn-free" disabled>Mevcut Plan</button>
  </div>

  <div class="plan featured">
    <div class="plan-name">Premium <span class="badge">ÖNERİLEN</span></div>
    <div class="plan-price">₺49 <span>/ay</span></div>
    <div class="plan-period">veya ₺499/yıl (%15 indirim)</div>
    <ul class="plan-features">
      <li><span class="icon">✅</span> Tüm ücretsiz özellikler</li>
      <li><span class="icon">✅</span> Yatırım takibi (hisse, fon, altın…)</li>
      <li><span class="icon">✅</span> Kıymet &amp; amortisman</li>
      <li><span class="icon">✅</span> Tedarikçi &amp; fatura yönetimi</li>
      <li><span class="icon">✅</span> Excel &amp; PDF raporu</li>
      <li><span class="icon">✅</span> Telegram bot entegrasyonu</li>
      <li><span class="icon">✅</span> Gelişmiş AI finansal analiz</li>
      <li><span class="icon">✅</span> Çoklu profil (Şahıs + Şirket)</li>
      <li><span class="icon">✅</span> Öncelikli destek</li>
    </ul>
    <button class="btn-plan btn-premium" onclick="requestPremium()">Premium'a Geç →</button>
  </div>
</div>

<div class="contact-box">
  <h2>Nasıl Satın Alırım?</h2>
  <p>Şu anda online ödeme sistemi kurulum aşamasındadır.<br>Premium almak için aşağıdaki adrese e-posta gönderin:</p>
  <p><a href="mailto:destek@kirpiapp.com">destek@kirpiapp.com</a></p>
  <p style="font-size:.78rem;margin-top:12px">Konu: Premium Üyelik Talebi · Kullanıcı adınızı belirtin</p>
</div>

</div>
<script>
function requestPremium(){{
  window.location.href = 'mailto:destek@kirpiapp.com?subject=Premium%20Üyelik%20Talebi&body=Merhaba%2C%0AKullanıcı%20adım%3A%20{html_escape(session.get("username","") if uid else "")}%0APremium%20üyelik%20almak%20istiyorum.';
}}
</script>
</body></html>"""

# ── ABONELİK API ──────────────────────────────────────────────────────────────

@app.route("/api/subscription/status")
@login_required
def api_sub_status():
    return jsonify(get_sub_status(session["user_id"]))

@app.route("/api/subscription/activate", methods=["POST"])
@admin_required
def api_sub_activate():
    """Admin only: activate premium for a user."""
    if not validate_csrf():
        return jsonify({"ok": False, "error": "CSRF"}), 403
    data     = request.get_json(force=True)
    username = (data.get("username") or "").strip().lower()
    months   = int(data.get("months") or 1)
    if months < 1 or months > 120:
        return jsonify({"ok": False, "error": "Geçersiz süre"}), 400
    with pg_connect() as con:
        user = con.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone()
        if not user:
            return jsonify({"ok": False, "error": "Kullanıcı bulunamadı"}), 404
        uid = user["id"]
        expires = (datetime.now() + timedelta(days=30 * months)).isoformat()
        now_iso = datetime.now().isoformat()
        existing = con.execute("SELECT id FROM subscriptions WHERE user_id=%s", (uid,)).fetchone()
        if existing:
            con.execute(
                "UPDATE subscriptions SET status='active', expires_at=%s, note=%s WHERE user_id=%s",
                (expires, f"Admin activated {months}mo", uid)
            )
        else:
            con.execute(
                "INSERT INTO subscriptions (user_id,plan,status,started_at,expires_at,note,created_at) VALUES (%s,'premium','active',%s,%s,%s,%s)",
                (uid, now_iso, expires, f"Admin activated {months}mo", now_iso)
            )
    return jsonify({"ok": True, "expires_at": expires})

@app.route("/api/subscription/revoke", methods=["POST"])
@admin_required
def api_sub_revoke():
    """Admin only: revoke subscription."""
    if not validate_csrf():
        return jsonify({"ok": False, "error": "CSRF"}), 403
    data     = request.get_json(force=True)
    username = (data.get("username") or "").strip().lower()
    with pg_connect() as con:
        user = con.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone()
        if not user:
            return jsonify({"ok": False, "error": "Kullanıcı bulunamadı"}), 404
        con.execute("UPDATE subscriptions SET status='cancelled' WHERE user_id=%s", (user["id"],))
    return jsonify({"ok": True})

@app.route("/icon.svg")
def icon_svg(): return ICON_SVG, 200, {"Content-Type": "image/svg+xml"}

ICON_192_B64 = "iVBORw0KGgoAAAANSUhEUgAAAMAAAADACAYAAABS3GwHAAASI0lEQVR4nO2dW68kVRXHV18MFsGAThR0aIhCJdpvGnM+QUVDjoYPMERDCARe0EfjByA8MvPihAkxGuYDGDRkzPkExsDbhHBiDBk1oiIgl1bmnNM+9Kme3bv3Ze29175VrV9ycqrrsqu6+//fa+1LdU2gIJ565sN17mtg4vPKy/dOcl9DT9YLYcEzAHkNkfTELHgGQ0pDJDkRlfC7tvkWADiXdXS8eovi/ExaUhgh6glChG8Ru2+52/fLpqiHmEaIUrCv8Lu2+aZidTQTiK/ZEOUTwwikBfoIX1HTq8owlUthhH55wkYoH0ojkBXkKn5FbS8f72IE0mhw/p/NUDBUJiApxEX8kvBtos8RDfrX2z82QplQmCCoAEfhi6mOq/AxxsBss6FMi2BjgrcDymUiEmIE7wOJan1T7p8jJQIwp0VTNkKZ+JrA6yCs+A21Plb4oSnRmfECzUyF5Z1oAGyEIvExgfMBDuJX1fpUwlddg0nsVG2DqbB+cnS8+pNnuUwkXE3gtLOH+DG1fkhKlCMt6penwEYoEhcToHfEiB+R8lAIX67pc/YU9f+nsEmL2AiFgDUBaqcA8dtE7pIS+QyWpYoG/f/Z0fHqzwFlM4RgTDCnOJGD+F0jgK22zzFusIY7gu+P377u2uYh2KRF73iUzSTG6hBb7W8Rv28EMNX4JZigxzaSfCugbIYAWxQwbowgftdaP0T4JbQN2AgFYDLBVLchkfjFmv4M7oh/rdlHtw3zWrcOs83011+3/PoMAE4B4LRrm4td2yxAwZOP36NazRBi0rLWAAhs4peFpVsHgKv1fU2hWkdtgl74qte9GU66trkITFEoDYCo/eV+/pAIYKv1bcs2U2CNgdkWEg3OYBMNvtq1zdeASYpO084RIFD8osjOEMfYlnXbMK9zRoPTrm0e6NMfToPysWcAU+1/nvcDmEUPiH1UKY8sPHm7atklMpiig7wOFQ0Wy9m/dNvAHg1C5iltYfPgUWnbNQJozSFtx4jfJnhMBMBGBvE1ZTQwHWMyQv9/S9c2DwCTnB0DWGp/bOqDFT8o9gs1hG4b5jUmGvgYQBa+aAqRk65tvsJGiIuscVQEQKQ+FOJXrcNuV+2PNYW8zlbjAwCcLZaz9wEAzv+r0hvV3+n5n1weCNtPurb5MiDgNkQ4WwNYen5MovURv0rcrhEANOtttb18nK12x+Tu2CjQ/532BupZLGf/Fso/7drmAtYIjBui1q1zgTRTm11QiR8I15mWXbb16ARuwvezkTmF3enW665tLsBmkt0/iM7BCGAbwb6pj4v4MRFAtc5kBExK1F+nXLvbUiUZ1yigoo84a7iTLp0AwO2ubb6kOYYJwGgAyw9VuYofU4a8zsUkpjRIXFalN/J2Y1q0WM4+Et+M8JrCAKq0q28bfLE3gpz3+7YDxt5+mAKg838XIfvu6xMVXJYBdrslYbGcfago22YCFboyXAxgnU4BGyPca7gOBkGveW0EEHp+REwCN+3nK37VNtN207IuxRGXda9168CwXSv+xXL2qaqA8/W6ATRx+VRzDU6MvfYHMKdAOiHJ6+T/urzfR/wUEcAmfFUKI++riyag2ddmAhNy7a+LBmSM2Qi2RrBJpKp11OJX7eOy3Vbjg7Rd3KZ8bai9V9K+OiNhDSDX/tvXumsYs5B9maryf8vPF+rW+eBjEHmdars80U65LDVmdaLVvVZhqvF9DKCLBlq6tmEXIHnqmQ/X2BRItS609veNDrb9XKZXb1ksZ5/AvshUr22YTACL5ewz08Hn23UmEI2t47Rrm7ttRuBoscF1HEBeltfFFr9qm2hA076mZRmfmltVvtYIiHJ0f7b8X+w2vRt5vtGyZwCp98cmHlPNmEL8PZjp1cYIgDzGKGCp9paPIzHAYjk7sRwr34Dz+a5tGtsJxxoRVBFA/pJstaZptNeGi/hVqYgp5VGVZ0uDVqb9FsvZ/4zvZvcYXTTAYIoARhbL2W3YTZn6+5Lv6iPCWMWuQjcXyDXcy+tcanuX48Vl233EPqmP7/6m433K8TmnyBls5hT15fS/ULHu2uYuAMAYeRT4jAPYxKbbFvpfXPa5o2xnu64rUbO/XKYJXfoDi+UM1X+/WM5CDdBHgLWwLE7HZs5xuSMMW8vGMIO4rBO/ap0tGuxxnuao0hcUi+Wsn+8vHx8qah9jiOc+A0C1IawMKYXaMYBi+oP8gWNDO4UZVMeYxL+GzRf8MbgZQoccMW4jjpGPDRa9DwajGK9nSMLGIrcBXHJXl5rYVI4O+Tib+HXH+1yD7zUPnqGZBNMLJK+zpRMxUiCs+Ptc+xPD/rpr3CL0pPT7jsYIXdug0uKhGEH3Zm1fuE+NG1P8KpGqUh5XMXsJX5WCuObv8v4EDWNsGeuubSZYI9ROaCPYtIxZ59MI1u4j1vyaPv1+G7Y/H2CT/1fXc0LRkyQbYSi1vgjFXCCXfUPbARiD2MpBXYOiN2cUKIyz7tqG7IHqpWG7Kd5HAFhjYFOfEPGj834DJCa4dfM0q4hCI8IQa38Av1+Hxv6Ss7zsYgab+E1lbBFSnaRtgPNzk0UPyrKYXTAGcGkQu9b+2G22fbHpjxOL5Wxds/hiX/sQogL2GWGqHhbda9/eINt2l/bAlsVy9tmtm6efM+0TQorUxnQOjMhDjVBzJWDD5SF5NhO4Hi+uc837XSLAzrl8pgLkzt9NhJpDdxz2PXdtMzk6XlVrEMq+3ljtAdT/xXL2X+S1jYZUxq153MD1ouXZjK6pkGpf7xwde6xQ64/SCCE4RJG97tIa2ggUrsU0QHXrQlMfF0Yp/tT5e21jBjsGODpevYU4Bitqn33EbaRmCBnYGnIj0IRvA7trm0kNtT+A/5PiKbpGdbW/qpwYEWEUqNoBKXqOakFlAPFWOhdC2gPyNp8IYQV7R1bJvT4UyO+PWuxieU8+fg/88jcfUxZPii4CYExgS4Uw2zBPjYleyw9d8DZ8o8QQ8E2BeijaA5h1ZCnQ2MUeG12boNSxgr1eIKEhTCUUlzEBr3ECl9sVx1KzlUapvUO2blDMRWNqc6zwVce4bmMCCb15R0eJJtAZwPVCXdsDqn1Y8JVTY3Q1RYCJ9D8U8cPRTamW10VpHNf4RaUi9mdTWhRQGgA5ICbjmwrZtnNEqACTceRtJZnAFgFcowBlKiQuY8zgBEcBNbdunk76v5jnKcUEWgMookCICUQwd5SZyuXaPxGpzJAT2zjA9kdVA8/jOkCG7iFy+cWGIX+Rsek/O8rIWcL4ANYAvQmmgPtxVYoBMmx7wAiLnhbx86QwQ24TGMcBztMgWUBU7QHXZXmd8kMTwzaLPy5D+Hwx9wOIUSAUnwlz1i5RFnw96KZK5LgWAIQBpCiQqlcIfSyLPh+UI8a5TIC6I+zoePU20KZCuu0ubQPuBaqIUrudXWaDTsGtISxiSn3k1z7tgSq4fuUa2fVeev7pqn9pTkWOBvFE9aBsHV3bPAob8YvPrJUf4qz7A8W+oFnWbdOtKwZKkfuSwhyhBtClrqkN4Ho/wBT2fxkCO05guxvLFiXkddmFBlCG4GXka6I2RMx0JnUUcDLA0fHq7a5tHoE76RD1AJnr+EByShS8jdiGqBmnFKina5tvAMAJ7KZAtlRIfpK7Lt0xpUbychJqFD2WEDOERAJb712qKOB7S6RqigQ2FXJ9Y1lq/yGLXkR8n65moB4VzoFXBAAA6Nrm6+DWIBa3gWK7vA4sy1EYi/BNpIgKpUSAEAM8AgC3YdcAJhOYeoB8TLDl4HDu/B7+8LuT7RfAotfjawaMEUowgbcBALZR4ARwJlCtA8Q6kJd9BG/jJ9//BXWRg8LVCKMwAABA1zYPwSYVsjWIsQbQiT7Zg+rYDHqwRqAwAEB8E1AY4GHYbwuoBK9LjUC37uBw7vJ09iiwGdSYjEDVDgCIb4DgX4c+Ol69A+oZo95zhQ4O55+VIH4AgMs3noPLN57LfRnFMZR2U3AE6Ona5kEwp0K2cQI4OJybHnJRBBwR9hGjgUt36CAigIDqngF0w6kG8QMARwMFMaNB7GnSZBEAAKBrm4twpy0gRwFlBDg4nH9Kdf7UcDTY52dXn3XaP3cUoH6uk1ie9eaZmsUPwNFAxYvPXs19CU6QGuDoeHXrvExrKnRwOP+E8ty5YBPsU5MJyJ/sd3S8+itY7iM+OJyX+8QED9gE+1CaIGY7INajLVXdoVMAgIPD+UeRzpkVNsE+NUSCKAY4Ol79DRQR4OBw/h+qc9x3Af+XCjbBPqWbIObDjcW2wOTgcP5+SGEhok5pCDbBPiYT5J5GHc0AYhQ4OJy/51tODNHGNgKbYJ8Xn71a5E/YRH28/dHx6u8Hh/N/+hyboraOeQ42wT7Xr1xb+/6AWayGcFQD+JA6b891zrHSjxqX8kt+UQ3w859+wWn/3CKkPj9HATXi1IncJohmABfxl1QDU18Lm0BNKbNJQ58THAxWbN/+zg/31r35xmvEV3OH+y4AfODddGdqgXQyXA+29seIXyV8mZhGoDIBT5xT43KrZYxJcdkawS7if+cvv93+mfaLQSmp2VDJnQqRGwBT+7uKX6RWE3BboEyK6wYF0Iu/J4cJmHjkjALJDVBrSlHrdddCLhOQGsCW/lA1emMebyLUBJwGlUeRKVDPww/+wGk9Uzc5okAyA/jWnrLYc4ufU6FhkX0gDENu0TPDhSwCmPL/odWaIe+H2wFmUqdBxbUBQkd1Y44KM8OjOAMwTEqStAF+/F31PJhf/VGdDrz5xmte3Zlc+w+D61eureU5QrF+HCuaATCTv0RjyGZwNUFq8fNs0WEQJQXymfmoihJYUXPNz/hCboCQab86E+gEbtrGMBiqGAcA4Fp+bKjaATEgjQAUN33oGszMeKnp16EZpirYAMyoITPACy99RDLMrxsbGAp8b3BZcARgiqamZ4QBQNhkr5pqfx4EGwZRukEv33jOOdRTif/3vz65ZNr+vR/Nr5OciBkE5L8LpJoWrTNDKtHrCDGDTwTg/N+NS88/PYmdAiUZCIuZ2viKvz+WI0K5xBY/QCIDfPAe/U0xIcJXleNiBM7/hwN5I/iFlwb5CLBgOP0pk2TdoJS1JlXt71Mm1/5puP+xJ5KcJ4oBOArswrV/uSQdCKu99qz9+pl9ohlAFwUoRBSj58ZWpu91c+1fNlEjwNhTIRa/H6nyf4BMN8RQdIv2NXZogxgTTTj1GS7R2wAxUyGAsHQopvi59q+DKI9IUqH75bjUA2QpBrxY/P6kTH8AEhoAIJ0JKGDx5yG1AZJ2g5rSoVLy7JBrYfHXR/IbYkw9Q7lNEHJ+Fn84qWt/gEx3hNlMkNoIoedk8ddLtlsibWMEKYxAcQ4WPw05an+AzPcEYwbKYhiBqkwWPw25xA9QwC/D9SawPWBPFqxLzxG1gVj4wyFpN6gNzEO2c8PipyVn7Q9QmAF6SjQCC5+e3OIHKCAFUoFNi1LAwh82RUYAFSnNwKKPTwm1P0BFBuh59/VXozxpkUWfjlLED1ChAQA2JhDxMQQLPg8liR+gUgMA7JuAKZ/SxA9Q8Y/jlvhhMnpK/b6qNQBAuR8qs0vJ31PVBgAo+8Nlyv9+qm0DyHCboD5KMMdgDADAJqidHIaYAACwCZiSSGWEV16+d1J9G0CmhLDKhPHu668mq8gGZwCAjQnYCPWTwgSDNAAzHGKbYAqwyYWiniUD3BYYDjG+y17zHAGYKohVoQ3SAFz7M1gGaQCGwbI1wBDbAQyjQtQ6RwBGyVi6kXcMMJQoMJYvLzYlfY5U1yJrnCMAY6QEE8S8hj0DcBRgZHJ+lpTnVml70BGATeCH6nPL8VmmOKe2tucZouPFJrzYn2cM4esymyJ/GIua+x97IpsJxC+zBiNixNfvQ/1+st0PoGNIUaAnlQhdvsxSjBEqQNf3kXLev26btcE7RBP0lFSD5TbBUNtLtk4dVI/PkE0gkrsGKyFNGxKYHs1RtAGw5BZCrNzadr4xg+7zH0sUKI2S0rSawI5nOQ16sQnykztNqwGXwVznUV82AVMyrjMZnEeChzJVghkePtr0mgrBJmBKw1eT3nOB2ARMKYRokUTE3C5gckBRCZPMBuVowKSGSnNk06HZBEwqKLUWRbScEjExiFHJRq212QgMBTGziyRpCxuB8SFFWp00b2cjMBhStiezNlzZEAxA3g6Uonpu2BDjoKQew/8Dln/QyG6DDBAAAAAASUVORK5CYII="
ICON_512_B64 = "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAA6VUlEQVR4nO3d3Y4kyXXY8dPTbYNlkeaKK2stmS1LtsqW+04C0U9QMEG0BT7AEiaIBRfcG5KXBh9A4KV2b5bgghBscB9AoA1ijX4CwyDvBpTaH6SGtkVbK+2SKzW/ptsXU9mTXZ0f8XEi4kTE/wc0pqcqMjMqu6rOiZORmUcCU1559f3b0n0AgBS+8fWPHpXuA57jj5EZAR4AppEg5MXOTohgDwBxSArSYccqIuADQFokBHrYkREI+ABQFglBOHacJ4I+ANhEMuCHneWAoA8AdSEZWMcOWkDgB4C6kQjMY8ccIOgDQJtIBu5jZ+wR+AGgDyQCz3S/E3oJ/Lvt5l+JSJLXenl1/b0U6wWAlHpPBLp98a0F/pQBPhYJAgDLek0EunvRLQT+3Xbze6X7oIHEAIAlvSUC3bzYWgN/4pF9yn3i/d4iIQBgQS+JQPMvssbArzjCt/zaV997JAQASmo9EWj2xdUU+BVG+Rqv1UI1YLYdyQCAUlpNBJp8UTUE/8hRfsjrs7xPlt6Hk8+READIqcUkoKkXZD3wR4z0fZcxvR8czb037z1OIgAgp5YSgSZeSAWB33e07/N6upggOLMcyQCAIlpIBKp/AVaDf8Bo37Wt1us1ud/2QuYM3P1OIgAgh9qTgKo7bzH4e472cwR96wmDz3vQda4AyQCALGpOAqrseCeBP/Q1tjZBUMR/kuCDZIBEAEBKNSYC1XXYUvD3LPOnCvo9ThAUcZwkOPo/iQCApGpLAqrqrJXgrxz4Uwb9lPurukmCo/8fkQgASKGmJKCKjloJ/CJepX7NwM8EwfuiJgkKiQCAxGpIBMx30ErwVwr8Kcr1sfunlasIrrVdSgZIBACos54EmO6cheBfIPBbmyAYs1wO3lcRlJlqgJAIAFBmOQkw27HSwd/jOL9G4GeCoL61swMOHxsfFvizNF0C0COrSYDJThkI/i6j/hyBnwmCD8W8Z50nCQqJAABFFpMAcx0qGfwrDfwlAr71SoHaJEEhEQCgxFoSYKYzFYz6Uwd+JgiGyTVJ8M+9egUAM6wkAiY6UfGoP3fgZ4KgP+1JgiQCAKJZSAKKd8B48LcQ+JkgmEbwJEEhEQCgoHQSUHTjpYK/4wz/mOBvbZ5AyDI13mZYRO8qglOPHSYBV4HbAgARKZsEFNtwweCfctRvMfBzFcGHXN73S2cMkAgAUFMqCTgpsdFSEgb/2MBfIuhbmCCouZ4pcx+qqW0etnXdh0e77eZ3ReQRhwUA1KRI1pF79O9Q8m8h8KcO+jnmGZSgPUnwv6n0CkBXSlQBsm+wQPDPPerPGfgtnRVQQ7B35Xqr4fFjJAIAouROArJurJLgX3LU3/sEwdTrFsk7SfC/B24LQKdyJgHZNmQs+Fsb9TNB0I6Qqwge/p9EAECwXElAlo3kDP6JjveXDPxMEExH6yqCh88/uJKgkAgA8JAjCUi+gQqCf6pRv6XAb3GCYOw2c/AJ+nOPH1YD/kd0rwB0IXUS0MxpgBmDf0wwZYJgXeb6fTTz/Nzjd+vbbTe/IyQCAAxIml3kGv0nON5vbY6A6zo01xO6TI8TBKeWXTtjgIoAgFUpqwDJVmw4+Ocu+aeuGLi2tT5BUHtdqWhPEpTLq+vvR/YJQMNSJQFJVtpY8Ld8ZkCNEwS11qG5nim+nw2NSYL/03ObADqRIgmodg6AcvDPPeq3FPiZIBi+/fEHcqr92vPj5452281vC4kAgEzUM4oco38Dwb/2wM8EwbQ0riT4fe1OAaibdhVAdWWNBP+a5gi4Lu/atuYJgjnWn3OSoMizGwx9P2KbABqjmQSorajj4J9y1G818DNBcF6KSYJ/EdspAO3QSgKqmQOwP89/Turgb2nUX+p0wFxBv4YJgiJ+txqean978Nzce/Jot938lgiJAABdKllE6tH/ykV+LAb/Vs4MYIJgnJBbDR8+N3XtgB/EdgxA3TSqAI80OpKSYvC/9Wy/tP5U1QLfvvv2yzWgu64rpIrgs+xh+5BtlrTU/6XXNbfcjYjcDBUBDZ/79Ie1VgWgMtEJQIZj/1rBf+qxXHMEYg4VaCQOPsF27XnfCkJMwNc0F1Q1f0L7MvX44WM3o39vdtvNxzUTAQB10Yi9UQlAhtL/3KQ/jeA/J7RSEJowxCQNMYmDa9uY4ObaLjbgawbnGKF9cO3/4WMkAkDHYmOw2UMABYP/1GO+y6xtJ2a5tWV9R9ohz4e2jw3IWsE9ZRXAdb/6/g2mlrurCuy2m4977QF5Xv7nMADQp+AEIOXo31jwXxKaMIQsF7tsTN+W2vqMbH3EBPrQkbgmnz6EPD9ODG5E5HZfDfBOBADUKyYWmzsNcOV0v0Olgr/PdlMu57KsxvJa29JeLlcwT2Gu72u3Gh4/dzv1+G67+Sfy7CJCT2I7CaBdQRWAxMf+XYNkTcE/dOTuUiJeEru8a1uNUb5PW+2R/NpoXeMnpi9Tjx+2vzl4/uk+EQDQuNCYbKoC4FH61xq1pk4Wlp5PtdxaG80Rf0jAT9E257pCufTB9eJChxcOmqsayG67+U15Vg34ocP2AXTE+0ICqUb/SsE/pq12e81lXJZba+Mz2o95PtV2c61Dcz1zUt5q+PD/R3L/3gI/FJme+Pcnf/qBZ7cAWOJ7cSATZwEsHPePGTGWDP5LJf8lKYK/b4k99Pmptq6HDUICbmiZXbNcH8q3DyHPjw8dDIcHnu4rAgDglwAkPPYfO0q0Fvx9Hh+em9uOS9IQuuxaO5+g6Bv0fYQE6SLB/fTs+K+UVrXWf5eE4HBewK08O2XwN5T6mB2nLALzfGN08TkACqX/mCpByHo1k4XQ7YT0Q3t517ahQVezD7FCEhZfrqW7qXVP3UzocF7Arcc2AHTAuQKQYvSfKPinbJsj+LuO+uceD60YuC7v2tZ3xO0zUtcc1c+NtLNVDRT6MdVurWoAoEE+sdrEHIADPgHIddlUbYfnciyztlyOwL8WQHwDTM6ArxXcZwP06dnxuyIi+39dAnpsoF97P0y2PT07/tHUyq2X17lyIaDLKQHIPPo/FJMQzD2uGfynHvNdZm07sctpBf7Q56fa+gR9X6GBPiY4a/Lth+/zs3bbzT+O7DuAwlxjdpE5AAlK/5aC/xKtZVyWi112rZ1vMAx9nRrr1tyWFXOvYe1Kgmuv/Wa33bwkIkeXV9d/Gdo5APatVgAy3O530HLwXyrVLkk16l+jUTUYt11bV0gAD6kkpBjF+4zKQ39C+zNe3nd/3ey2m5dqPmMA6JlL7M4+B8Cj9D8lxSgxR/D3eXx4bm47Icu5LKuxvGtb38AWE/BjOAfn07PjvxkvePj/XP1wXN6n7dN9RaCow+P+zAMA4mVNADwu+OMTaGLWVyL4uwRHn8ddlssR+NcCjG9QzhnwNUbePuuN+Yl6Dadnx3+91OH9dQyGZe+uI7Dbbn7dQiIAQM9iApCg/O8SQLWDuu9I0+c5zfZLy4SMpnMH/tDnp9r6BH1fMYE+dASuyTdJCOnnuO04EbjZJwJMFAQqsBbDs1UAlEv/GsE/pm2K9loJg8tya8vGBv6YgLPWxjfY+gZAC0E+VuxrWFr+Rp4dFvh15T4DyGw2Acg0+c8lCPt8aWlud+k53+Dpu36XZbRH/ZqB34VLUNIYpfu0VX3Pn54dvzezDe0fbePR/2wisNtu/lGORIDj/UC4pViepQIwM/r3CcKa7VIGf5/Hl9avnTCsrbdE4F97PjTou7bTCJy3InK7D/QlOSUJrv3cT2A83E+LiYDS6/BCYgDEKX4vgBWuwTq0nc+2fdebKoj7LhOzvbU2WkmbbzDWThZTryPFuubMXe9fax+P7ztwO2xvt938mjy7dsD/89wOgEImKwCa5f+I0X+O4J+7re96xs+FBP+l5WIrBrFVA5/1jNv6jPJ9TY6cI2muy3d7Wq9laX23cn+S4K+VqggAmDYX05MeAlg47W8s9osptJ2V4O8SJH0e11guZeD3DUgpg35IgPQKrKdnxz/2WI/2T4y5da1t80b2iUDk9kWEMj+QUuo5ACGj1qk2MYHep12J4L8kNGEIWW6pP5qB34VLEAsJdL4BcjWoLgR4r/UkErPNpcA/t/57icBuu3lRKxGYQ4IAhHuQAGiV/yNG/67BP3U7rbah7bUSBtflQpOGtXYhgX/t+dCg79rO53VX4fTs+Cee7YfExrXSUDQRALBsKranrACEjv5D1t1a8Pd5fHhOs1qwtj3X5X1H2SHPL7V3TSZUqgGRXAJtzI9Gn1z6On7+Zvz7brt5MbAfABIoeRaAS2CM+ZJNsWyJ4O8ycvV5PHa5tTY+gTrm+RTbDNl2ifWVMvU6jmYeHz93O/y+224+Js/OFng3TRcBuEpSAXCc+R8qNEnIWSXI0X5pmRTVgrU2viPpped8R/subTSrAV5Oz44/0F5nJi77xbcSMUwS/NjaYYGaj+/X3Hf0414CUPDWv1OPxVQIUrdbapurfe5qQY7A78I3oMe0WeyH73F13/Ur/6TqU2i/h0sKfyy2UwRbwM1hjFevACiN/mO+sLSDv8uya8/5Bj7f9bssEzPq91mnTzvvwH96dvz+yrpSBX3NgJorYIuIyOnZ8d8GLveBQ1+W+u2UCOy2m1/VSASsICFBLbLeDngvZOQ61SZ0PT7tXLe79JxvwA5Zv3bCsLbeXIF/LSCmDPraATl2XSGj/hR9X1v33POTwX/8+z4RaGaiIIkArFNNAGZO/Qv54l17TDvQuwa6lME/JGC6BHGf7ayt13VZrcA/aV8F0A76yUbgIquj8BxBPJRvwJ/6zBw+N5cIDIcFfpXgCaR3lwAoHf8PDUgaXIN1aDufbfuu1zexcFkmNGGISRpSBv6l8v942dCg7+M2cGKfteAewiVJcU0Ixv8fJwE3vp2ykjBY6QcwZxzrU58GmGv0nyP4524b0t5327HPxS6bevsa7UOXSa2WPk2dJjjcsOh2pR2AhNQOAcxM/hvTGP3HfEHEBKOag//aKDNm1J50xL+0jv2V6nz+prmqAxqWRtlefTo9O76O6cjp2fHfRfZlqq3rYwASyn0lQN/2rkEkZLux7SwF/7kvzRTVgrXtJQ38HlIG/aD+7YPp1LZig5/mupbW77t9l7aHj3lfttgyDgnAMpUEwOG6/6lG/67BP3U7rbZa7V2WCQnga1/sa0IDv08wSxH07wWtg0DuK+S1xIzANfn2JbRNkNLBtvT2AV+PRFQmAKZYXnP04rLeVoK/SzDzedxlOZdlfZdfDQijkaJv8PAN+qmVDuyaQoL+4XMAEhpifqpJgKHBZO75mNG/C5/A4bpsqeC/RGsZl+XW2nitd6EsHPO3C11XjBq3Mbe+o5nHfddB0AcKiD4EEDj5z+f5GKFJQmw7a8F/aZS9JKZaEDri12ZppC+nZ8c/jVzF3Aj7wes4PTv+WeS21tbj0hfX98qtiNzGTloshfI/apRiEqCV0b9r8E/dbqmtb/uQQB4SxGOCtG/f1/qyaOZ8/LXgExL0cyYKocG0VJ+8grzjMkF2280/TLHeGCQHsKr0dQA0vgRi1qEd/F2WXXtOK7HQXCbVcrkqMr7rmW2fYISq+f4tybUv48MG6v0/PTv+8U42H5Fntxz+sfb6gZZEVQAcZv8fSjH6T7ken3au2116TrOqoJkwaAd/1xGgRjvfEWeO0XZMnyxUA2LkeC3DLYc/orzeSYzwUavYQwCpRsYx2wvtU2w7S8Hf5/HhudCEwSdp0A78IjJ7nX3toK8WqFaOzVsN9HOBe+nHa/1acxb22x6SgA/nSgSA2li6EJD28j7rjg3qsYlPyuDvEth8Hl9bZ8w+mlyv53X3QwL/2vOp3ovawf5uPadnxz+PXNc9p2fHv4jpz8JPCuP13yUCibYFVOuR0k2AXLiMJn2eDx395wj+uduGtF9aJjRhCB31awY/13Zrr2MqEYmdwX+4nZhlS1QKQqoA2ZOs/TyNIfjfJQK77eZXSiUCHCaANa+8+v5tcAVg4vS/2FFw6LpSrMNnBOm6rKXgvxSsl/j2VyPwO/0tHK7OFxT0CysV6FPIXQ24Hf17706DmokAgR01S3UIwPfDnGL0H7oel+V821kL/j6PD8/5BPmQv0VoG5d1+FYtSgkJiDmThJgKQO4EZm7bw+2Gb3bbza9k6gtgUso5AGO3M7+vtXV9PiRBcA2E2u202mq210oYNEf9MYFiLdiEBlg1+2P0PoGx6WrA6dnxL3NsZ/T4kAT8AxIB9CooAQg4/W9JbEKgqcfgv0Rz1J888I9uXbu0bt/AX0pwNSBRMJXTs+OnKdabyOF+G1cAHlQEQhIByv+oXWgFYCkAugZHl/XG9sWnTej2fNrVEvx9grxrYF97frLNzCl+PlRG+4qnqLls26W/paoBc5WI2B9V+7MWlioAh8nAXSKg3ZcBCQOsSX0lwEOxH3SNL4rYUb3r+kK3u7bO2MQiZP2+23V93nd7PjQTz1TBVfs9hofG++7wxkPD/29Hvx/ttpsPicijy6vrmNs+A+blmgMwR2PkuNY+JAjFJgmh7Zbazq3Xd90hyUWK4L+03uCANzoFbM3a6DP5yHqhVF9iZF/M6dnxTaJVh1Qi7s4Y2G03H5qrCDCaRwu0E4CYLyvtZCC0jeuy2pWEtbY+6w1JLFyXcUkiQp/PEexi+59KbMDPlTCEBNW1n1wOtznXlweJQMY+Atl4JwCe5//fzvzuIlfA16oi+LRz3a5G25D2IeuJeT5n4PJ+Xvuqeq7bdVzu3vIJR9PD+qupSOwnLc4lGmsVgHv/31cDNpm6DmRR8hCAT3IQ8nxIYNcO9BpVgtzB36ca4LMd71G/w8V95pYbX7FvbdulRvw+So6aW7K0H9eSgRsRebrbbj4UU/7n0AEsKT0HYI7G6F+LawDWbhfbVqt98Mh5pQ+hy7lyCfyu60lm4dS6LoJ9horCWuLkkgwcJgJAEzQTgKVgFTrSdmmbcvSfOvgviW2rFfx9t+vyfKkEzzWglgi+MSP8LpIFJUv7eeq5wySgygSAygOmaN8OuNQ6Ytcf04eYkaRPgNRo69M+JkivjbZ8l/N2cN6+b+DPLSbg3y2b6/h8TfMAHLgmBDLTBqiWVwKgfAXAOalGnbFttEvGtQZ/lyDu87jr8yGiA3/gbXBTYJQfICBZmUsIbkXktrKrIYoIo3/M860AhHz5pBxdlzw8kKOdVluf9iHBfW19odWC1Jb6li3YzszcjzkkUIVKKgnqfwMCMqwocTfAw6xaa70h64vRWvB3Gtmfnh1/MPOc1mGEYaT104nnvCycvmci8C9su4bA2KPq/i7jZIPEA4e0LgWs9cHwSQ5Kj/5dxBwOSBX8tQL12nMpDhPEiqlupFZ6+1hRScUCcGb1NMAp2qP/kHkAc4/HHvcP3e7aOmOD/+RodFQFCC35+y4TK2jEn+queinkDk41BcOa+grkFJMAhATQmPK/z7ZDjynHvqac7Zbauq43xfF+7WWiVHgYAFBByR9rUlQAQr84tZID323NPaaZoGhXEtbaJgn+4+P/E7fp9a4krCyTisnA39sotbbXq91fgjMsqOUQQIr5ABr9iAlqMcG+5Mh/qr12yT9FcDAZ+NGv3Xbz91KtmwQDLpwTgIVrAIR8ceYu/689HxI0Sx33LxX8c5X8bw8u4hNt4Vx+Aj9Kerrbbo5zbpDEAGM+FYDYL0rX4BIyAo15XlOp+QGubTWD/9LNekJL/rn+VquBv6YLvpQqp9dUxo/pa8LXeSsit7vt5lHuRAAQSXsIIEXZ3fLoP8ekvznZg79ne5eSf87gr9FGXU0BNUYvr9PR8N6/2W03Kt/HjPLhqpY5AHO0R5Spg0PM4YCYRCFZ8D+oApgd9e9P6XPpC8EJWUwkQre77eZIKxEA1mi/0UKCUUz532fboRP2Yl9T7LqX2sXOVXB+bmLWf0hfNJZJgcAPS4ITAUb/8BGaAJQo16beptXSf+7gHxIMp7YbUvK/XThvP5UH/Zi5Nj8qF3LoofDhitvddnNUcPtoXK5SU4oPkc/osvRIs6bgP/XY2oS5a8ft+m5b1cHEPpOjft+AU/p4em39rdG+GkAiAHUlbgZkcXs5Jwf69MG1Xcrg7yO05J87GSMIoTprhwVcy/8cJsAgNgHQDLyacwMsTw5M3a5I8D+oAkytq3QVxgXJAe7kqFYEbIPDAlCjdTdAV1ofKJ/koIbJgdrtXJZdelzzi69YyX/K6dnxzZPHT6cS39v981UlAE8ePyUYeDg9O75tYZ8NScDl1XVV71fYkjsByCHH6D8mEIeuK+Zwg6Xgb+0Ly1p/mglSh2pLrmowJALsW4TIMQkwxfF5rWVTBqjc8wM0RtqTbReu+HfY7qcr61qbTJjzDADK/ZhFQEUPNBOAHKPgw+dSf0hTTg503Z5Lu5Dgr1nFcNnuWl+yfeFymh9aQrKCUCWvOJXiTesz4s8x691n2dDSf+rgn+KQScz6gaaUCOCcCQCR+i8FvCTFoQetiX+pg/+SlMFfJu7kV03wtzSSstQXDZZej6W+ACXlTAC0g5X2aN365ECNZTXOGtDoh9b6gxEEMKem90ZNfYU9KRKA2EAa+ob2SQ6sTw6MWU5jxn/Ua91XAZbWPfVa+CIDgIxaPA1wipXRf6nj/sHBf+UCP75m+7G/W183WjzNb4nL62U064f9hVi1JgApS+spR/8uzAR/ZaX36wPj8+01v0x7C+5aciYJc9daKBVUW732A2wrlQBoTaabei71Bzjl6F+777HBP7g/p2fHP3/y+OnfT7HukviSLotKAqAndQKQ87S60HWuJQ8pXkOu4/6u7XLOedBctyqCext6SBI0+v+5T39Ynjx+esTlhPvV8mmAU2Lf6KHBTKPiEVv6zx78Z67sZyohGL5ICf59sVT+L42bC/WrdALAqYHh6/V5vMTI38L2gOakSFRIAvrknABcXl1/L2I7KUbevsu1eGqgy7pyLRuy7uEMgKcJtwtUoXQFYrfdHJEI9KV0BWBOjg+CldG/dunftd3qdg5u7hPk9Oz4FzN9MDPyL/3Fi/J4DzxHEtAPqwlAbjlG/5ZL/7lPc+TLFgiQK1EhCeiDpQQgdqKc5twAX1qTA2MmFJoN/qOL/BD4AUOWEoqQQwLcZKguJRIATg2cfiwmMGufVui7/ZBtm0IJGNaVeo9SDWhXaALQ0hui1KmBvuv0WW+pSZfemACI0iwlf5b6MuaSBDD6r4+lQwADTg3026526T/JF9BMoDf5ZQfgISoB7SmZAHBq4HT7mNK/WvDf39EvJXPB3+roC7Dy3pxLAsajfyoB9fBKACauBZAzI7TwAUgx+k+9nqIj/7FRFcDC3xIwE1hrQiWgHRYPAZQUWx1IMfrPddw/l1r6CZhgMUkZJwGM+Otl/XbAPZ0aaKL0H9CXpnBb1r5YvCdA7vdg6PZGSUC33xe1s5IAWHgDWTg10KVN6tJ/0r/F6dnxzZPHTx8Nv6fcFhDCYlJgFfulbpoJwJHkC+QlRu+xy6c65z7mLIPuP7yM9uHi8H2SK/ARYJFSTAKQM+CH0Ay4OU4NdGmTerSe60vtrgqQE8EeWqgS9Pd6W+SdAFxeXX9vt938XorOLNCYCOf6XO5TA7UvJBRd+j89O/55xPZdJf/yIOAjJ+2kgACL1FLMAYipDKQYSaeWq88hZwyYOO6fA8EeFoUmBdaD/1r/PvfpD8uf/OkHubqDQBoJgPVDAbFSnhro2j4kYJs+7h/zBUewR81KzSdYwtkvfbJyFkBqFk8NtHL8vviXzxK+lNC62uYTuPZtt90cXV5dm30dCLwQ0OiKgCW+nGNGv9rbynFqYFel/yePnx6Nf0r3B0A4rhpom+UKgLX5AJqnBqYcrWtfFCgpgjzQNioBdmmfinX4Zd7yl3uOUwO1qh3OZ1Gcnh3/wrEtgARaKP+jDloJgNVDAYftfI7/+7zRYw8VhG4jtM9mPsR8oQDt41CATcEJwMSdAVOxcCggdR+0rwWwtB4TpX8AD1lOiGP7RhJgj0YFwOWP2tofXvPCQS7L5J4cCKAz2snH1PpIAmxJdTnWGv7IWuX/0G1qbkvzyoHZWR71AND9jJIE2BGVAGQ8DLAm5amBvslB7PH+kJF96HoAGNFTIkwSYINmBeDo4N8ULMwHCF1/DaP/Yl9APX35ATVJ9dkkCSivxFkALf/RNa8VMLc+rdH/gzanZ8e/dFg3AGW9JsAkAWXluCVrzj+wdsld69RArdG/5sS/2L4AaFyviUkvohOAhXkAljO70qV3l/WlOLtAY9lk+LIByirxGaQKUI72HIBcf0hrcwEOt5Hj1ECXNpoXDwKQgNXEV6NfPjcOit0W/KkkABNVgKk/Zm2XCU51IR6NZXJfIjgbq1+GANIiCcgvxxyAUrRPlUt52eBuAz7QM6sJr9V+QVeKmwH1nsVpXwfA5fno0j9nAAAojSpAXmoJwMxkwB6vDeAr5WQ/rT5kxegDvXjy+OnR8FO6L1aQBOST8hBAzdcGyHFqoNbzTPwDGmAlGbCQgJME5KGaAOyrAC6HASxeG0D71ECNY/IpPohVJAMWvoSAUsbJQOmEAO3KPQmw50MBVkb/ACrTYzJAFSC9k8TrH/6At/vf14KRS5uSau+b+dF/T19wQIjxZ6SGStnp2fFt6Od6t90cXV5dm3+NtVKvAHgcBhDHNrlpnBroG4zXRu5a9wTwXT6LHkc3gIYePjtUAtJJXQGwplTQS73dqIl/p2fHT3W7s6zlLyuglNoqAygvyRyAURVApMy1AVqaD6BxXL946b+HkQpgRWufN6oAaeSYA3B78H8R9zkBlrieGpg7+TA78a+VLx+gZoefw1qrA8wH0JcsAbi8uv7ebrv5lxGrSJkglDo1UKO96dE/QR+wjUMFGKQ+DfDo4GetrTaLhwJ8qgUapwb6rC9Ia+VGoBe1fWY5FKAraQIwc3ngQ7X/QWOv/R+6La32QX0m6AP1y1UB0NwOSYCeHGcBxM4DsDBXILSMHhOws4z+T8+Ob9Y6JVLfSAEAsCz5lQAPrgtg8doAOZMLa9WCxTaM8oF2WTr+79sXqgA6cl0HwOIovnT7peVDZvb7jv4n2xPsAeRkKRHpTZZ7AUxcHfCwGpDjtsEplLrSnuqZAk8eP320/6lt/wPoFFWAeDlvBjQX5K3cNjj1qYEprxUQPPp/8vjpo4V2ABpkZdQd2w+SgDjZEoCDqwO6svDHbeHUQAAwzUpS0pOstwO+vLr+M3G7NoCFkX7pdWps2/XYPx88oBNWAq1WP6gChMuaAOxpBH3rf3DXywb7rovRP4AqWUk88Fz2BGBUBRibmhA493wOyS6gs7IezQ8Io38A91gJwtr9oAoQptTtgA/L/ybelA5qu2sg8MDbb7xV7L3y8hc/zxc1YMTRK6++X+TLYLfd/AsRuZFngWv8IxO/+/wb8pjm7z7tQtcTu52lx1ChkkE9FZIFXRZG/08ePz1a6kfsqcjcLdBPqQqAXF5d//luu9nu/ztcFnj4XeT5xYMsXERoSehlgg/b5r7wECrSYoBfs/aaSRCAOMUqACLeVYC1EX+KCoDLiHwtAQgd4TP671CPgV4bicFDFkb/rqgC5FOsAiByrwrg8wdLURFg9I3sCPZpTO1XkgLgoaIJwN7U9QCmvhitHwqYEjPCD13vUtvQ7SESwb6snpOCmkb/GnbbzRFVADfFE4CVKoBP0LeQIJQ+NbD068ceAd++w79RLwkBMCg6B2Bst938rtyfDyArv/v86/PYWluX5ZaWyfWcTxtEIuC3p4WEoMbRv9ZNyagCrCteARixcm2A3NtNUTXQXjcOEPDbR4UArTNTARCZrAKUqABonSXg2i7n7H8zf+saEfQxqC0ZqKkSoHlbcqoAyyxVAESeXZp4HDQtXhvAdbsp+hc7+Q+eCPqYMn5f1JAMjIOq9WTg9Oz4VjMJwDxTFQAR7yrA2og/RQXAdQSe8zlG/4oI+ghVQzIwsJwIUAXIw1wCICKy227+uTxMAmThd9f/S+Bja22X2ruua22Z0G0vPYY9gj60kQyEIwHIw2oCMFcFWJsHkDMByH38P2b0P/dY1wj6yKWWZMBKIqB9CIAkYJrJBEDEqQrgkgAs/evz2Np6XNpPtaP8nxlBH6XVkAyUTgRIAPIwmwCIiOy2m38mdq8NEBqg55Z3Xc6n3dJjXSHww5oaEgGRMskACUAe1hOAqSpAiQTAJ0FY+j1keZ91ry3bFYI+alFDMpAzEUhxFgBJwEOmEwCRySqAxQQgJGCTACRC4EetSASeowqQXg0JwLgKcLN/eCoRkIXHZOL5pTZzj7k8P9fGdRmf5yj/jxD40YoaEgGRtMkAVYD0zCcAIt5VgLWAH5MA5CjzM/r3QNBH62pIBlIkAiQA6VWRAIiI7Lab3xFb1waICebVjf7PL07U3yf/5T/9MvgDTuBHb6wnAtpJQKqrAZIEPGftUsBLDm8WJDId4NYuE5z6MsJW3lze/UgR5EO3N5ccEPjRq+G9bzERKH3aIMJUUwEQCaoC+P7r+ljIKD32+L9q+T93sNfwpU++WboLgBmWEoFaDgGIUAEYqy0BGOYClL42QKoEIEn5v8Zg74qkAL0rnQikGv2nvCEQScAzVSUAIiK77ea35eFkwJoTgBTH/29bDvpzSAbQs1KJQG1nAoiQAAyqSwBE7pIAl0MBS49N/ev6mEugjh3lLz032e784uRGICIkA+hXzkQg9bF/EoC0ak0AxnMBSlwbIKacH5IAzLY5vzh5KlhEMoDe5EoCak0AREgCRCpNAERmqwBzQX8t4FtMAGZ/J+iHIxlAT1ImAjlm/pMApFXTaYCHhtMCXf6IqU/907Dav/OLk1/m6EjLXn/nNREhEUAfLJ86iPKqrQCILFYB1o77u4zuNSsAURMAzy9OfiFIgkQAPdFKBGq9H8Ch3qsAj0p3IMbl1fX3Z54aXzRo6g2U6k2l+mY6vzj5BcE/rdffee2uKgC0jgtpYazqCoDIvSpAzmsDhI7unY7/n1+c/FxQBBUB9CK0GpD7qn/MA0in+gRARGS33fyWrB8GMJ8AEPjtIBFAD3yTgBKX/OUwQDpNJAAid0mAxrUBDp+faqOaAJxfnPxMYBKJAHrgmgiQALSl6jkAE+aO/eeeAev0hjq/OPkZwd825gigBy5zA7jhT3uaSQAur67/QpYn/s0pcnrM+cXJT0tsF2FIAtC6t99465ZJgn1pJgHYm7pl8FLbmOeDnF+c/JTgXyeqAejBVBLQ8uh/t910e42EphKAy6vrH8jzJGD8Myj2hz6/OLkm8LeBRACtoxLQh2YmAY7ttpuPi98lgjXOEJh9/vzi5NrzJaASTBJE6/7d175QugtMBEykqQrAyCMJmw+gjuDfNioBaN1Xv/C10l1IrtfDAE1WAERmqwDZKgDnFyd/F9N/1IdqAFpWshKQugIg0mcVoNUKgMj9KsBgbk6A6oRAgn+fqAagZV/9wte6qAb0pNkEYH9aoMj8mQFLQT042zy/OPnb0GVRP5IAtI4koB3NHgIY7A8FTN0xUBZ+X/v/g7L/+cXJB6odR/U4JICW5TwkkOMQgEh/hwGarQAECjoUQPDHFKoBaBmVgPo1nwBcXl3/UMKvDbCadRL8sYQkAC3LlQS0fCGikppPAPbmXmfURYLOL05+EtYd9IQkAC1jcmC9ukgALq+un8j8jYJCHBH84YMkAK0jCahPFwmAiMjl1fX/2v8aferf+cXJj1U6ha6QBKB1tScBvV0QqJsEYM/n2gCTCP6IQRKA1tWeBPSkqwRgPyFQJPDaAOcXJ+/r9wq9IQlA60gC6tBVArA3fs3O5Z7zi5P39LuCXpEEoHUkAfY1fyGgKbvt5jfl/sWBROYvFCTnFyd/k7uPub3wYvw63ns3fh294WJBaJ3mBYO4J4CuLhMAEZHddvMbsn6zIDm/OPnrIh1MQCPIhyI5mEcSgNZpJQFcEVBXj4cABnNvpLvHaw/+L7x4/4e+2MThALSOwwE2dZsAXF5d/29ZuDZAjcG/piBbU19zIAlA6776ha9lG8HDTbcJgIjI5dX1/9n/qnWBoOxaCaKtvA4A895+463bJ4+fHpEI2NB1ArD34JTA84sT80esWw6WLb+2JVQB0IO333jrVuTZ8XwSgbK6TwAOqgBH5xcnf1WyP0t6GyX39npFSALQhyEJEHmeCJAM5Nd9ArD3SETk/OLk/5buyJTeguCUnvYBSQB6ME4CBmuJAHcF1EUCIPcmBJrSU9BzxT4B2le6ItDLPQFIAPbOL07+snQfBgS5da3vI6oA6MFUFWCsdCLQOhIAEfnKlz9Sugt3Wg5qKbS8v0gC0IO1JECERCCV7hMAK8G/9RFtSi3vO5IA9MAlCRDJdyXAXnSfAJTWcvDKjX0J1Ms1CYCerhOA0qN/glUare1XqgDoBUlAXielO9Cj1gKURcM+5iZEADCt2wpAqdE/wT+vVvY3VQD0gipAPl0mAAT/vrSy30kC0AuSgDy6TABKaCUI1Yr9D9SldBLQw8WAupsDUGL0nyP4/P4f/OFqm+9+51vpO2LYCy/WPyfg9Xdeky998s3S3QDQgO4SgJxSB36XoD/XvtdkgMmBQD3efuOt25e/+PnmR+KlHL3y6vvdHGvJOfpPFfx9g76LXpOBmpMAqgDoSakk4PLquun4yByABGoK/inXax3zAgD0rJsKQK7Rf4qgkjNA91gNqLUSQBUAPSlRBaACAGe1B/8S27OASgBgX+mzAlrURQKQY/SfO/j/4If/8cFPju22qsYkgOsCAIjRRQJQo7kgvBTsNROBHpMAALZRBdDVfAJQ4+h/Kfi7IAkIQxUAQE+aTwBSsxb8Q9vPIQkAYAlVAD0kABFyBYvQYK45L6AnJAGAbSQBOppOAFKW/1uY8e/Kar9SqikJ4DAAgBBNJwA10Sr9ay8/6DEJAGAXVYB4JAABahodIhx/ZwAtazYBSFX+Jyj0pZa/N4cB0COqAHGaTQBqUkt5vZZ+AgDWkQB4qGU0CF383QG7qAKEazIBSFH+LxUE/unH/03R5fFMDUkAhwEA+GgyAahJbWX12voLoH1UAcKQADgoPfoLHcUz+tdV+n0AAJqaSwByXPu/BN9gTvDvE4cB0CuqAP6aSwC0WRr1uQZ1gn86lt4PABDjpHQH4GcI7lNX+CPwAwBckQAssDzaI9iX88KLIu+9W7oXAA69/cZbty9/8fNHpftRCw4BAADQoaYSgFYnAAI+mAiInjEZ0F1TCYAmy+V/lMf7A2jb5dV184kECUBh3/3Ot0p3wUtt/QUATCMBmMDoDi54nwA2cRjADQkAAAAdIgEwoJayei39BACsayYB0DoDgLIufFh9v3AmAHrHYYB1zSQAAADAHQmAEdbL69b7BwDwQwJgiNUga7VfAIBwJAAjVo/nwjbeN4BNofMAergIkAg3A5r12U+86dz23/9XvQlX3/3Ot+T3/+AP1dYXi9E/ALSJBGDvS590D/iHDpOF2ITAShJA8AeAdnWfAMQE/jlDQhCTCJROAgj+ANC2bucAfOmTbyYJ/mOf/cSbXocSDpUKwgR/AK3gegDzuqsApA76U2IqArkrAQR/AOhDVxWAEsF/LLQa8N3vfCt5YM6xjZZxJgCA2nSTAJQO/gOLhwQI/ADwTC+nAIp0cgjASvAffPYTbwZPEBwH65hDAwR9AOhb8xUAa8F/EFMJGAxle9dg7tseAFrARMBpTVcArAb/QUwl4BBBHQDgo/kKAAAAeKjZBMD66H+gcSgAABCvpwmAIg0nAAAAYF6TCUAto/8BVQAAQG7NJAB/9Mc/Kd0FwIzakmAA+TWTAAAAMIdTAR8iAQAAdK+3CYAiJAAAAHSJBAAAgA6RAAAA0KGmEoDhTIDX39G5vG4uWpcDRjnvvVu6B89xBgDgp8fj/yKNJQAAAMBNswlALVUARv8AgBKaTQAAAFjTa/lfpPEEwHoVgNE/AKCU5hKAw0sCW00CCP5IgQmAAFw1lwBMsZYEEPzbYukMAABw1UUCIGInCSD4A4ANPR//FxE5Kd2BFP7oj38iX/nyRx48/vo7rxUtkbYU/P/zf/jly77L/Ot/e/J2ir7gGcr/AHw0mQAsGSoBOb8saw/8IcHedT0kBQBQRncJwCBHIlBr4NcK+CHbqi0h4Pg/UKfey/8iIkevvPp+szth6jDAHM1EgMAfr5ZEwEoCQPkfWPfyFz9/NPxOAtBxBeDQeJLgCy+KfPYT7l+otQb8gaXAPxj6VEsiAAC1IQGY8N679Qd1FxYD/yHLiYCV0T8AP4z+n2n6NMDDiwLhuRqC/1ht/c2J8j+wblz+xzNNJwAxWh7d1RpMLfW75fcHgD6QAHTGUhANUXv/AZRF+f+55hMADgM810rwbOV1aKD8DyBU8wlAjJbKvK0FzZKvp6X3BYB+dZEA9F4FaC34D1p9Xa4Y/QN+KP/f10UCEIPRHsZ4PwBoRTcJQO9VALSF0T/g7uUvfv6I0f9D3SQAMWoe9bVeJs/5+mp+HwDAoa4SAKoAaAGjf8APo/9pXSUAMRj99Y2/P4DWdJcAxFQBCAJ9svR3Z/QPQEt3CQAAoB8vfeozpbtgVpcJQE9VAIt30dOU+vVZ+nsz+gegqcsEQKSvJABhLP2dCf4AtHWbAAAA0LOuE4BeqgCtHgZI+bos/X0Z/QNhOP6/rOsEIJalILGmtSSgl+APAKl0nwDEXhyopmDRShLQU/Bn9A8gle4TABGSgJoQ/AFABwnAXk+XCa41Cai13yEI/kAcjv+vIwFQYm30uKa2YNrT+f4AkMPRK6++z00SRr7y5Y9ELf/Ci0odycjyHQNzJCrWgj+jfyAeFYB1JAATekwCRGwlArkqFAR/oD0EfzckADN6TQJEyiYCOQ9NEPyBNpEAuDkp3YFWvfduvUnAYRBOmRCUmotgLfgDQG5UABbEVgEGtSYCS0KSAgsTD60Gfkb/gA5G/+5IAFaQBLSD4A+0jwTAHacBrtC6PoDV4NMLq/uf4A/oIfj7IQFwQBJQN6v7neAPoCQSAEckAXWyur8J/gBKIwHwoJkEWA1MrbC8jwn+gD7K//5IADxp3jPAaoCqneX9SvAHYAUJQADtJMBywKqJ9X1J8AfSYPQfhgQgkPbdA60HL8tq2HcEfwDWkABESHELYeuBzJoa9hfBH0iH0X84EoBIqZKAGgJbSbXsI4I/AKtIABSkSAJE6glyOdW0Twj+QFqM/uNwMyAlQxKgdengsSHg9Xw54VqCvgiBH0AdqAAoS1UNEHk++q0pGMao8fUS/IE8GP3HIwFIIGUSMKgtMPqo9bUR/AHUhLsBJpbikMCcmg8R1BjwBwR+IC9G/zqoACSWoxowGJfMrQfUmvq6hOAP5EXw10MCkEHOJGDMUpC11BctBH8ANeMQQGY5Dwn40Dh80EpgX0PgB8pg9K+L0wAzS3m6YIxegncMAj+AlnAIoJBShwUQhuAPlMXoXx+HAAywVg3AcwR+oDyCfxocAjDA6mGBnhH4AbSOCoBBJALlEPgBWxj9p0MCYBiJQD4EfsAegn9aHAIwjEMD6RH4AfSKCkBFSAT0EPgB2xj9p0cCUCESgXAEfsA+gn8eJACVIxlYR9AH6kHwz4c5AJUbX1CIZOA5gj4ALKMC0KAfffub8vo7r5XuRnYEfaBujP7zIgFo1I++/c2731tOBgj6QBsI/vmRADRsnAQcqjEpINgDbSL4l0EC0LilJOCQpaSAYA/0geBfDglAB3ySgDkpkgOCPNA3gn9ZnAUAJwRrAGjLo9IdQHpk2QCs4XupPBKATvBhA2AF30c2kAB0hA8dgNL4HrKDBKAzfPgAlML3jy0kAB3iQwggN7537CEB6BQfRgC58H1jEwlAx/hQAkiN7xm7SAA6x4cTQCp8v9hGAgA+pADU8b1iHwkARIQPKwA9fJ/U4dE3vv7Ro9KdgA18aAHE4nukDt/4+kePqADgHj68AELx/VEXbgaEB1761GdU7iAIoC+h3xskDmWQAGDS8IEkEQCQ2tz3DIlBWiQAWEQ1AEApU989JAV6mAOAVXzgAFjxo29/8+4HcUgA4IQkAIA1JANxSADgjCQAgFUkA/7urgHwyqvv35bsCOrChwyAdQxapg3X/6ECgCB8sABYR0VgGQkAgpEEAKgBicA0EgBEeelTnyERAFAFkoD7SACggiQAQA2oBjx3lwBwUyDEIgkAUItek4BxrKcCAFUcEgBQi16TgAEJANT1/qECUI+ev69IAKCq5w8TgDr1+r11LwFgHgBi9PohAlC/Hr6/DmM8FQCo6OHDAwAtIQEAAED6G8iQACBabx8aAO3q6fvsQQLAPAAAANoyFdupACBKT9kyALSEBAAAgJFeBjaTCQCHAQAAaMNcTKcCgGC9ZMkA0CISAAAAOjSbAHAYAACAui3FcioAAAB0aDEBoAoAAGVxe22EWovhJ7k6AgBwcxj0X/rUZ5h0m1EvSReHAADAkF6CD8pbTQA4DIA5fFEBgE0usZsKAAAYMlfqJ+HOo6f97JQAUAUAgPJ6Ck4ltLJ/XWM2FQBEaeUDAwC9cU4AqAIAQHkk3Wm0sl99YjUVAERr5YMD1ILPnK5e96dXAkAVAABs6DVoaWtpP/rGaCoAUNHShwioBZ+7OL3vP+8EgCoA5vT+YQI0+H6O+NyFaW2/hcRmKgBQ1dqHCqjBS5/6DJ89R+yr54ISAKoAWMKHCyiDz96yVvdPaEymAoAkWv2gASlpfG4Y4T7EPpkWNZJ/5dX3b7U6gjZxBzPAXYog1fNnsIegH1ORpwKApHr4AAIaUn1Wehz99viaQ0Qfy6cKAFc9j0SANTkDVoufxR4Dfux8PJXJfCQBcNXiF08uS19w7Ne6lQxeNb93egz6A43J+CQAKKLmL53ctL/k2Pe2WApiNbw3LO2vkswkACIkAQhTwxdOKbm/6Phb5FdDMCv5vqhh/5SgdSo+CQBMIPg8Rzm4D7UHN833Su37IjdzCYAISQB09BiELH0B9rj/c7P090ZdNC/Ep35FP5IAaOkhEFkNBD3s+1Ks/s1hn/ZVeJNc0pckANpaCki1BICW9rkVtfztYU+KS/CfaK8QSOHwi7Om4FTrl/7Q75r2tVW1vgfQtmQ39aEKgNwsBKpWv+gt7NtatfqeQD6pbsCX9K5+JAGwghnLOkgE3PX8PoGelHffTX5bX5IAoD0kAvMI/NCSMviLMAcAQADmBzxE4EdtklcARKgCAD3oMRkg6COV1KN/kUwJgAhJANCLHhIBAj9SyhH8RTImACIkAUBvWkoGCPrIIVfwF8mcAIiQBAA9qykhIOAjt5zBX6RAAiBCEgDgOQtJAcEepeUO/iKFEgARkgAA67h+A3pQIviLcBogAMMI2kA6j0ptuFTGAwCAFSVjYbEEQIQkAADQr9IxsGgCIFJ+BwAAkJuF2Fc8ARCxsSMAAMjBSswz0Ykxzg4AALTISuAfmKgAjFnbQQAAxLIY28wlACI2dxQAACGsxjSTCYCI3R0GAIAry7HMbAIgYnvHAQCwxHoMM50AiNjfgQAAHKohdpnv4BhnCAAALKsh8A/MVwDGatqxAIC+1BajqkoAROrbwQCA9tUYm6rr8BiHBAAAJdUY+AfVVQDGat7xAIC61R6Dqk4AROr/AwAA6tNC7Kn+BYxxSAAAkFILgX/QzAsZIxEAAGhqKfAPqj8EMKXFPxQAoIxWY0qTL2qMagAAIESrgX/Q9IsbIxEAALhoPfAPuniRYyQCAIApvQT+QVcvdoxEAAAg0l/gH3T5osdIBACgT70G/kHXL36MRAAA+tB74B+wEw6QCABAmwj897EzFpAMAEDdCPrz2DEOSAQAoC4E/nXsIE8kAwBgE0HfDzsrAskAAJRF0A/HjlNEQgAAaRHw9bAjEyIhAIA4BPx02LGZkRQAwDSCfV7sbGNIEAC0igBvy/8HohCO5wohXCAAAAAASUVORK5CYII="

@app.route("/icon-192.png")
def icon_192():
    import base64
    return base64.b64decode(ICON_192_B64), 200, {"Content-Type": "image/png"}

@app.route("/icon-512.png")
def icon_512():
    import base64
    return base64.b64decode(ICON_512_B64), 200, {"Content-Type": "image/png"}

_BANK_DOMAINS = {
    'Garanti BBVA':'garantibbva.com.tr','İş Bankası':'isbank.com.tr',
    'Akbank':'akbank.com','Yapı Kredi':'yapikredi.com.tr',
    'Ziraat Bankası':'ziraatbank.com.tr','Halkbank':'halkbank.com.tr',
    'Vakıfbank':'vakifbank.com.tr','QNB Finansbank':'qnbfinansbank.com',
    'Denizbank':'denizbank.com','ING':'ingbank.com.tr','TEB':'teb.com.tr',
    'HSBC':'hsbc.com.tr','Multinet':'multinetup.com','Edenred':'edenred.com.tr',
    'Ticket Restaurant':'ticket.com.tr','Sodexo':'sodexo.com.tr',
    'Pluxee':'pluxee.com','Paye':'paye.com.tr','Enpara':'enpara.com',
    'Papara':'papara.com','Param':'param.com.tr','Tosla':'tosla.com',
    'Ininal':'ininal.com','Albaraka Türk':'albaraka.com.tr',
    'Kuveyt Türk':'kuveytturk.com.tr','Türkiye Finans':'turkiyefinans.com.tr',
    'Revolut':'revolut.com','Wise':'wise.com','Paycell':'paycell.com.tr',
}
_bank_logo_cache = {}

@app.route("/bank-logo/<path:bank>")
def bank_logo_route(bank):
    from flask import Response as _Resp
    if bank in _bank_logo_cache:
        ct, data = _bank_logo_cache[bank]
        return _Resp(data, mimetype=ct, headers={"Cache-Control":"public,max-age=604800"})
    domain = _BANK_DOMAINS.get(bank)
    if not domain:
        return "", 404
    try:
        r = requests.get(f"https://logo.clearbit.com/{domain}?size=64", timeout=4,
                         headers={"User-Agent":"Mozilla/5.0"})
        if r.status_code == 200 and r.headers.get("content-type","").startswith("image"):
            ct = r.headers["content-type"].split(";")[0].strip()
            _bank_logo_cache[bank] = (ct, r.content)
            return _Resp(r.content, mimetype=ct, headers={"Cache-Control":"public,max-age=604800"})
    except Exception:
        pass
    return "", 404

# Digital Asset Links for TWA (Google Play) — set ASSETLINKS_JSON env var after PWABuilder
@app.route("/.well-known/assetlinks.json")
def assetlinks():
    content = os.environ.get("ASSETLINKS_JSON", "[]")
    return content, 200, {"Content-Type": "application/json"}

@app.route("/")
@login_required
def index():
    display = session.get("display","")
    first_name = display.split()[0] if display.strip() else display
    return HTML.replace("__USER_DISPLAY__", str(html_escape(first_name)))

# ── LEGAL PAGES ───────────────────────────────────────────────────────────────



@app.route("/kvkk")
def kvkk():
    body = """
<h2>1. Veri Sorumlusu</h2>
<p>Bu aydınlatma metni, 6698 sayılı Kişisel Verilerin Korunması Kanunu (KVKK) kapsamında Kirpi Nakit Akışı uygulaması tarafından hazırlanmıştır.</p>

<h2>2. İşlenen Kişisel Veriler</h2>
<div class="box">
<ul>
<li><strong>Kimlik verileri:</strong> Ad soyad, kullanıcı adı</li>
<li><strong>İletişim verileri:</strong> E-posta adresi</li>
<li><strong>Finansal veriler:</strong> Kullanıcının bizzat girdiği gelir, gider, yatırım ve kart bilgileri</li>
<li><strong>Teknik veriler:</strong> Giriş tarihleri, IP adresi (sunucu logları)</li>
</ul>
</div>

<h2>3. Kişisel Verilerin İşlenme Amaçları</h2>
<ul>
<li>Kullanıcı hesabının oluşturulması ve yönetimi</li>
<li>Şifre sıfırlama hizmetinin sunulması</li>
<li>Uygulama güvenliğinin sağlanması</li>
<li>Yasal yükümlülüklerin yerine getirilmesi</li>
</ul>

<h2>4. Kişisel Verilerin Aktarılması</h2>
<p>Kişisel verileriniz <strong>üçüncü şahıslarla paylaşılmamakta</strong>, satılmamakta veya kiralanmamaktadır. Yalnızca yasal zorunluluk halinde yetkili kamu kurumlarıyla paylaşılabilir.</p>

<h2>5. Kişisel Verilerin Saklanma Süresi</h2>
<p>Verileriniz hesabınızın aktif olduğu süre boyunca saklanır. Hesap silme talebinde bulunduğunuzda tüm verileriniz kalıcı olarak silinir.</p>

<h2>6. KVKK Kapsamındaki Haklarınız</h2>
<ul>
<li>Kişisel verilerinizin işlenip işlenmediğini öğrenme</li>
<li>İşlenmişse bilgi talep etme</li>
<li>İşlenme amacını ve amacına uygun kullanılıp kullanılmadığını öğrenme</li>
<li>Yurt içinde veya yurt dışında aktarıldığı üçüncü kişileri bilme</li>
<li>Eksik veya yanlış işlenmişse düzeltilmesini isteme</li>
<li>Silinmesini veya yok edilmesini isteme</li>
<li>İşlemenin otomatik sistemler vasıtasıyla gerçekleştirilmesi halinde aleyhine bir sonucun ortaya çıkmasına itiraz etme</li>
</ul>
<p>Başvurularınız için: <strong>kvkk@kirpiapp.com</strong></p>
"""
    return legal_page("KVKK Aydınlatma Metni", body)

@app.route("/gizlilik")
def gizlilik():
    body = """
<h2>1. Gizlilik Taahhüdümüz</h2>
<p>Kirpi, kullanıcılarının gizliliğini en üst düzeyde korumayı taahhüt eder. Bu politika, hangi verileri topladığımızı ve nasıl kullandığımızı açıklar.</p>

<h2>2. Topladığımız Veriler</h2>
<div class="box">
<ul>
<li><strong>Hesap bilgileri:</strong> Kullanıcı adı, ad soyad, e-posta</li>
<li><strong>Finansal içerik:</strong> Yalnızca siz tarafından girilen işlem, bütçe ve yatırım verileri</li>
<li><strong>Teknik veriler:</strong> Sunucu erişim logları (güvenlik amaçlı)</li>
</ul>
</div>

<h2>3. Verilerinizi Nasıl Kullanıyoruz</h2>
<ul>
<li>Hizmeti size sunmak ve hesabınızı yönetmek</li>
<li>Şifre sıfırlama e-postası göndermek</li>
<li>Uygulama güvenliğini sağlamak</li>
</ul>

<h2>4. Veri Güvenliği</h2>
<p>Şifreleriniz <strong>bcrypt</strong> algoritmasıyla hashlenerek saklanır. Ham şifreniz hiçbir zaman kaydedilmez. Verileriniz şifreli bağlantı (HTTPS) üzerinden iletilir.</p>

<h2>5. Çerezler</h2>
<p>Uygulama yalnızca oturum yönetimi için zorunlu bir oturum çerezi kullanır. Reklam veya takip çerezi kullanılmaz.</p>

<h2>6. Üçüncü Taraf Hizmetler</h2>
<ul>
<li><strong>Railway.app:</strong> Uygulama barındırma (ABD merkezli)</li>
<li><strong>Zoho/Gmail SMTP:</strong> E-posta bildirimleri</li>
<li><strong>open.er-api.com:</strong> Döviz kurları (anonim)</li>
<li><strong>tefas.gov.tr:</strong> Fon fiyatları (anonim)</li>
</ul>

<h2>7. İletişim</h2>
<p>Gizlilik konularındaki sorularınız için: <strong>gizlilik@kirpiapp.com</strong></p>
"""
    return legal_page("Gizlilik Politikası", body)

@app.route("/kullanim-kosullari")
def kullanim_kosullari():
    body = """
<h2>1. Kabul</h2>
<p>Kirpi uygulamasını kullanarak bu kullanım koşullarını kabul etmiş sayılırsınız.</p>

<h2>2. Hizmet Tanımı</h2>
<p>Kirpi, kişisel ve kurumsal nakit akışı takibi için tasarlanmış bir web uygulamasıdır. Uygulama yalnızca bilgi amaçlıdır; <strong>finansal tavsiye niteliği taşımaz.</strong></p>

<h2>3. Kullanıcı Sorumlulukları</h2>
<ul>
<li>Hesap güvenliğinizi korumak sizin sorumluluğunuzdadır</li>
<li>Şifrenizi kimseyle paylaşmayınız</li>
<li>Uygulamayı yasalara aykırı amaçlarla kullanamazsınız</li>
<li>Başkalarının hesaplarına erişmeye çalışamazsınız</li>
</ul>

<h2>4. Hizmet Sürekliliği</h2>
<p>Kirpi, hizmet kesintisi yaşanmayacağını garanti etmez. Verilerinizi düzenli olarak yedeklemenizi öneririz. Uygulama içinden veri yedeği alabilirsiniz.</p>

<h2>5. Sorumluluk Sınırlaması</h2>
<p>Kirpi, kullanıcı tarafından girilen verilerin doğruluğundan sorumlu değildir. Uygulama finansal kararlarınız için kullanılan tek kaynak olmamalıdır.</p>

<h2>6. Hesap Sonlandırma</h2>
<p>Koşulları ihlal eden hesaplar önceden bildirim yapılmaksızın askıya alınabilir veya silinebilir.</p>

<h2>7. Abonelik ve Ücretlendirme</h2>
<div class="box">
<ul>
<li><strong>Ücretsiz Deneme:</strong> Yeni kullanıcılar kayıt tarihinden itibaren 30 gün boyunca tüm Premium özelliklere ücretsiz erişir.</li>
<li><strong>Ücretsiz Plan:</strong> Deneme süresi sonunda temel özellikler (işlem takibi, bütçe, hesaplar, hedefler) ücretsiz kullanılmaya devam eder.</li>
<li><strong>Premium Plan:</strong> Yatırım takibi, kıymet yönetimi, tedarikçi/fatura modülleri, Excel/PDF dışa aktarma, Telegram entegrasyonu ve gelişmiş AI analizi Premium üyelik gerektirir.</li>
<li><strong>Ücret:</strong> Premium aylık ₺49, yıllık ₺499 olarak uygulanır (KDV dahil). Fiyatlar değiştirilebilir; mevcut abonelere 30 gün önceden bildirim yapılır.</li>
<li><strong>Abonelik İptali:</strong> İstediğiniz zaman aboneliğinizi iptal edebilirsiniz. İptal sonrası mevcut dönem sonuna kadar erişim devam eder; kısmi iade yapılmaz.</li>
<li><strong>İade Politikası:</strong> İlk 7 gün içinde yapılan ücretli abonelik taleplerinde tam iade yapılır. Sonrasında iade yapılmaz.</li>
</ul>
</div>

<h2>8. Değişiklikler</h2>
<p>Bu koşullar zaman zaman güncellenebilir. Önemli değişiklikler e-posta ile bildirilir.</p>

<h2>9. Uygulanacak Hukuk</h2>
<p>Bu sözleşme <strong>Türkiye Cumhuriyeti hukuku</strong> kapsamında yorumlanır. Uyuşmazlıklarda İstanbul Mahkemeleri yetkilidir.</p>

<h2>10. İletişim</h2>
<p>Sorularınız için: <strong>destek@kirpiapp.com</strong></p>
"""
    return legal_page("Kullanım Koşulları", body)

# ── ADMIN PANEL ───────────────────────────────────────────────────────────────

@app.route("/admin")
@admin_required
def admin_panel():
    db = get_db()
    total_users     = db.execute("SELECT COUNT(*) AS n FROM users").fetchone()["n"]
    total_profiles  = db.execute("SELECT COUNT(*) AS n FROM profiles").fetchone()["n"]
    sahis_count     = db.execute("SELECT COUNT(*) AS n FROM profiles WHERE type='sahis'").fetchone()["n"]
    sirket_count    = db.execute("SELECT COUNT(*) AS n FROM profiles WHERE type='sirket'").fetchone()["n"]
    total_txn       = db.execute("SELECT COUNT(*) AS n FROM transactions").fetchone()["n"]
    total_cards     = db.execute("SELECT COUNT(*) AS n FROM cards").fetchone()["n"]
    total_invest    = db.execute("SELECT COUNT(*) AS n FROM investments").fetchone()["n"]
    premium_count   = db.execute("SELECT COUNT(*) AS n FROM subscriptions WHERE status='active'").fetchone()["n"]
    recent_users    = db.execute(
        "SELECT u.id, u.username, u.display_name, u.email, u.created_at, "
        "s.status AS sub_status, s.expires_at "
        "FROM users u LEFT JOIN subscriptions s ON s.user_id=u.id "
        "ORDER BY u.id DESC LIMIT 20"
    ).fetchall()
    txn_by_day      = db.execute("SELECT LEFT(created_at, 10) as d, COUNT(*) as c FROM transactions GROUP BY d ORDER BY d DESC LIMIT 14").fetchall()
    new_users_week  = db.execute("SELECT COUNT(*) AS n FROM users WHERE created_at >= (NOW() - INTERVAL '7 days')::TEXT").fetchone()["n"]
    new_users_month = db.execute("SELECT COUNT(*) AS n FROM users WHERE created_at >= (NOW() - INTERVAL '30 days')::TEXT").fetchone()["n"]

    rows_html = ""
    for u in recent_users:
        safe_uname = html_escape(u['username'])
        sub_s = u['sub_status'] or ""
        exp   = (u['expires_at'] or "")[:10]
        # determine badge
        sub_info = get_sub_status(u['id'])
        if sub_info['status'] == 'trial':
            badge = f'<span style="background:#6366f120;color:#818cf8;border:1px solid #6366f140;border-radius:4px;padding:2px 6px;font-size:.7rem">Deneme ({sub_info["days_left"]}g)</span>'
        elif sub_info['status'] == 'premium':
            badge = f'<span style="background:#22c55e20;color:#22c55e;border:1px solid #22c55e40;border-radius:4px;padding:2px 6px;font-size:.7rem">Premium · {exp}</span>'
        else:
            badge = '<span style="background:#64748b20;color:#64748b;border:1px solid #64748b40;border-radius:4px;padding:2px 6px;font-size:.7rem">Ücretsiz</span>'
        rows_html += f"""<tr>
        <td>{safe_uname}</td>
        <td>{html_escape(u['display_name'])}</td>
        <td style="font-size:.78rem;color:#64748b">{html_escape(u['email'])}</td>
        <td style="color:#64748b;font-size:.8rem">{u['created_at'][:10]}</td>
        <td>{badge}</td>
        <td style="display:flex;gap:6px;align-items:center;flex-wrap:wrap">
          <button onclick="activateSub('{safe_uname}')" style="background:#6366f120;color:#818cf8;border:1px solid #6366f140;border-radius:6px;padding:4px 8px;font-size:.72rem;cursor:pointer">+Premium</button>
          <form method="POST" action="/admin/delete-user/{safe_uname}" onsubmit="return confirm('Bu kullanıcıyı ve tüm verilerini silmek istediğinize emin misiniz?')" style="display:inline">
            <input type="hidden" name="csrf_token" value="{get_csrf_token()}">
            <button type="submit" style="background:#ef444420;color:#ef4444;border:1px solid #ef444440;border-radius:6px;padding:4px 8px;font-size:.72rem;cursor:pointer">Sil</button>
          </form>
        </td>
        </tr>"""

    chart_labels = [r['d'] for r in reversed(list(txn_by_day))]
    chart_data   = [r['c'] for r in reversed(list(txn_by_day))]

    return f"""<!DOCTYPE html>
<html lang="tr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Admin — Kirpi</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0a0c12;color:#e2e8f0;font-family:'Inter',system-ui,sans-serif;padding:32px 20px}}
.wrap{{max-width:1100px;margin:0 auto}}
.back{{display:inline-flex;align-items:center;gap:8px;color:#818cf8;text-decoration:none;font-size:.85rem;margin-bottom:28px}}
h1{{font-size:1.6rem;font-weight:900;margin-bottom:28px;background:linear-gradient(135deg,#818cf8,#a78bfa);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
.stats{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:16px;margin-bottom:32px}}
.stat{{background:#111318;border:1px solid #1e2233;border-radius:12px;padding:20px;text-align:center}}
.stat-val{{font-size:2rem;font-weight:900;color:#818cf8}}
.stat-lbl{{font-size:.78rem;color:#64748b;margin-top:4px}}
.stat-sub{{font-size:.72rem;color:#22c55e;margin-top:2px}}
.card{{background:#111318;border:1px solid #1e2233;border-radius:12px;padding:24px;margin-bottom:24px}}
h2{{font-size:1rem;font-weight:700;margin-bottom:16px;color:#c7d2fe}}
table{{width:100%;border-collapse:collapse;font-size:.82rem}}
th{{text-align:left;color:#64748b;font-weight:500;padding:8px 12px;border-bottom:1px solid #1e2233}}
td{{padding:9px 12px;border-bottom:1px solid #0d0f18;color:#cbd5e1}}
tr:hover td{{background:#13161f}}
canvas{{width:100%!important;height:200px!important}}
.modal-bg{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:9999;align-items:center;justify-content:center}}
.modal{{background:#111318;border:1px solid #1e2233;border-radius:16px;padding:32px;min-width:320px;max-width:440px;width:90%}}
.modal h3{{color:#c7d2fe;font-size:1rem;margin-bottom:20px}}
.modal input{{width:100%;background:#0a0c12;border:1px solid #1e2233;color:#e2e8f0;padding:10px 12px;border-radius:8px;font-size:.9rem;margin-bottom:12px}}
.modal label{{font-size:.78rem;color:#64748b;display:block;margin-bottom:5px}}
.modal-btns{{display:flex;gap:10px;margin-top:4px}}
.btn-ok{{flex:1;padding:10px;background:#6366f1;color:#fff;border:none;border-radius:8px;font-size:.88rem;font-weight:700;cursor:pointer}}
.btn-cancel{{flex:1;padding:10px;background:#1e2233;color:#94a3b8;border:none;border-radius:8px;font-size:.88rem;cursor:pointer}}
</style></head>
<body><div class="wrap">
<a class="back" href="/">← Uygulamaya Dön</a>
<h1>🛡️ Admin Paneli</h1>

<div class="stats">
  <div class="stat"><div class="stat-val">{total_users}</div><div class="stat-lbl">Toplam Kullanıcı</div>
    <div class="stat-sub">+{new_users_week} bu hafta · +{new_users_month} bu ay</div></div>
  <div class="stat"><div class="stat-val">{total_profiles}</div><div class="stat-lbl">Profil</div>
    <div class="stat-sub">👤 {sahis_count} şahıs · 🏢 {sirket_count} şirket</div></div>
  <div class="stat"><div class="stat-val">{premium_count}</div><div class="stat-lbl">Premium Üye</div>
    <div class="stat-sub" style="color:#f59e0b">{total_users - premium_count} ücretsiz / deneme</div></div>
  <div class="stat"><div class="stat-val">{total_txn}</div><div class="stat-lbl">Toplam İşlem</div></div>
  <div class="stat"><div class="stat-val">{total_cards}</div><div class="stat-lbl">Kayıtlı Kart</div></div>
  <div class="stat"><div class="stat-val">{total_invest}</div><div class="stat-lbl">Yatırım Kaydı</div></div>
</div>

<div class="card">
  <h2>📈 Son 14 Gün — Günlük İşlem</h2>
  <canvas id="chart"></canvas>
</div>

<div class="card">
  <h2>👥 Kullanıcılar &amp; Abonelikler</h2>
  <table>
    <thead><tr><th>Kullanıcı Adı</th><th>Ad Soyad</th><th>Email</th><th>Kayıt</th><th>Plan</th><th>İşlem</th></tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>

<!-- Premium Aktivasyon Modal -->
<div class="modal-bg" id="subModal">
  <div class="modal">
    <h3>Premium Abonelik Aktiflestir</h3>
    <label>Kullanıcı Adı</label>
    <input type="text" id="subUsername" readonly>
    <label>Süre (ay)</label>
    <input type="number" id="subMonths" value="1" min="1" max="120">
    <div class="modal-btns">
      <button class="btn-ok" onclick="confirmActivate()">Aktifleştir</button>
      <button class="btn-cancel" onclick="closeSubModal()">İptal</button>
    </div>
  </div>
</div>

</div>
<script>
var labels={chart_labels};
var data={chart_data};
var canvas=document.getElementById('chart');
var ctx=canvas.getContext('2d');
canvas.width=canvas.offsetWidth; canvas.height=200;
var max=Math.max(...data,1);
var pad={{l:40,r:20,t:20,b:30}};
var W=canvas.width,H=canvas.height;
var bw=(W-pad.l-pad.r)/Math.max(labels.length,1);
ctx.fillStyle='#0a0c12'; ctx.fillRect(0,0,W,H);
[0.25,0.5,0.75,1].forEach(function(f){{
  var y=pad.t+(H-pad.t-pad.b)*(1-f);
  ctx.strokeStyle='#1e2233'; ctx.lineWidth=1;
  ctx.beginPath(); ctx.moveTo(pad.l,y); ctx.lineTo(W-pad.r,y); ctx.stroke();
  ctx.fillStyle='#475569'; ctx.font='10px system-ui'; ctx.textAlign='right';
  ctx.fillText(Math.round(max*f),pad.l-4,y+4);
}});
data.forEach(function(v,i){{
  var x=pad.l+i*bw+bw*0.1;
  var bh=(v/max)*(H-pad.t-pad.b);
  var y=H-pad.b-bh;
  var grd=ctx.createLinearGradient(0,y,0,H-pad.b);
  grd.addColorStop(0,'#6366f1'); grd.addColorStop(1,'#4338ca44');
  ctx.fillStyle=grd;
  ctx.beginPath();
  ctx.roundRect(x,y,bw*0.8,bh,4);
  ctx.fill();
  if(i%3===0){{
    ctx.fillStyle='#64748b'; ctx.font='9px system-ui'; ctx.textAlign='center';
    ctx.fillText(labels[i]?labels[i].slice(5):'',x+bw*0.4,H-pad.b+14);
  }}
}});

function activateSub(username) {{
  document.getElementById('subUsername').value = username;
  document.getElementById('subMonths').value = 1;
  var m = document.getElementById('subModal');
  m.style.display = 'flex';
}}
function closeSubModal() {{
  document.getElementById('subModal').style.display = 'none';
}}
function confirmActivate() {{
  var username = document.getElementById('subUsername').value;
  var months   = parseInt(document.getElementById('subMonths').value) || 1;
  fetch('/api/subscription/activate', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{username: username, months: months, csrf_token: '{get_csrf_token()}'}})
  }}).then(r => r.json()).then(d => {{
    closeSubModal();
    if(d.ok) {{ alert(username + ' için Premium aktifleştirildi. Bitiş: ' + d.expires_at.slice(0,10)); location.reload(); }}
    else {{ alert('Hata: ' + d.error); }}
  }});
}}
</script>
</body></html>"""

@app.route("/admin/delete-user/<username>", methods=["POST"])
@admin_required
def admin_delete_user(username):
    if not validate_csrf():
        return "Geçersiz istek", 403
    with pg_connect() as con:
        user = con.execute("SELECT id FROM users WHERE username=%s", (username,)).fetchone()
        if not user:
            return "Kullanıcı bulunamadı", 404
        uid = user["id"]
        for tbl in ["telegram_links","telegram_link_codes","telegram_pending",
                    "card_daily_balance","asset_maintenance","assets",
                    "supplier_invoices","suppliers","invoices","accounts",
                    "investments","recurring","todos","goals","budgets",
                    "transactions","cards","profiles","password_reset_tokens"]:
            con.execute(f"DELETE FROM {tbl} WHERE user_id=%s", (uid,))
        con.execute("DELETE FROM users WHERE id=%s", (uid,))
    return redirect("/admin")

if __name__ == "__main__":
    init_db()
    t = threading.Thread(target=_daily_backup_loop, daemon=True)
    t.start()
    app.run(debug=os.environ.get("FLASK_DEBUG") == "1", port=5001)

init_db()
_backup_thread = threading.Thread(target=_daily_backup_loop, daemon=True)
_backup_thread.start()

def _register_tg_webhook():
    if not TELEGRAM_TOKEN or not APP_URL:
        return
    try:
        import time as _time
        _time.sleep(3)  # gunicorn'un tam ayağa kalkmasını bekle
        wh_payload = {"url": f"{APP_URL}/api/telegram/webhook"}
        if TELEGRAM_WEBHOOK_SECRET:
            wh_payload["secret_token"] = TELEGRAM_WEBHOOK_SECRET
        r = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/setWebhook",
            json=wh_payload,
            timeout=10
        )
        log.info("Telegram webhook: %s", r.json())
    except Exception as e:
        log.warning("Telegram webhook kayıt hatası: %s", e)

_tg_thread = threading.Thread(target=_register_tg_webhook, daemon=True)
_tg_thread.start()
